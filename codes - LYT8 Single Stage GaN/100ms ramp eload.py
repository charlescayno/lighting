"""IMPORT DEPENDENCIES"""
from re import A
from time import time, sleep
import sys
import os
import math
from tracemalloc import start
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.default_scope_settings import *
import shutil
import numpy as np
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	
import getpass
username = getpass.getuser().lower()
####################################################################################################################################################################

test = "100ms ramp time"
operation = "CV"
condition = input(">> Type test condition: ")
excel_name = f"{condition}"
waveforms_folder = f'C:/Users/{username}/Desktop/APPS EVAL/Single Stage GAN/Test Data ({operation})/{date}/{test}/{condition}/'
path = path_maker(f'{waveforms_folder}')

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.101"
sig_gen_address = '11'
dimming_power_meter_address = 21
eload_channel = 5

"""USER INPUT"""
vin_list = [90,115,130,180,230,265,277,300]
vin_list = [90,115,230,265,277,300]
# vin_list = [90]
# vin_list = [130]
iout_nom = 1
vout = 50
# iout_list = [0*iout_nom,0.25*iout_nom,0.5*iout_nom,0.75*iout_nom,iout_nom]
iout_list = [iout_nom]
percent_list = [0,10,25,50,75,100]
percent_list = [0,50,100]
# percent_list = [25,100]
# percent_list = [25,50]
cr_load = vout/iout_nom

"""SCOPE CHANNEL"""
iout_channel = 1
# vin_channel = 2
ids_channel = 3
vout_channel = 2

"""SCOPE SETTINGS"""
trig_channel = 1
trig_level = 0.4
trig_edge = 'POS'

time_position = 20
time_scale = 0.1

"""ZOOM SETTINGS"""
zoom_enable = False
zoom_pos = 20
zoom_rel_scale = 1


"""
MEASUREMENT SETTINGS OPTIONS: "MAX,MIN,RMS,MEAN,PDELta"
"""

ch1_enable = 'ON'
ch2_enable = 'ON'
ch3_enable = 'ON'
ch4_enable = 'ON'

# """CHANNEL 1"""
# ch1_scale = 0.4
# ch1_position = -4
# ch1_bw = 20
# ch1_rel_x_position = 20
# ch1_label = "IOUT"
# ch1_measure = "MAX,MIN"
# ch1_color = "GREEN"
# ch1_coupling = "DCLimit"

# """CHANNEL 2"""
# ch2_scale = 300
# ch2_position = 2
# ch2_bw = 20
# ch2_rel_x_position = 40
# ch2_label = "VIN"
# ch2_measure = "MAX,RMS"
# ch2_color = "YELLOW"
# ch2_coupling = "DCLimit"

# """CHANNEL 3"""
# ch3_scale = 2
# ch3_position = -4
# ch3_bw = 500
# ch3_rel_x_position = 60
# ch3_label = "PRI_IDS"
# ch3_measure = "MAX,MIN"
# ch3_color = "LIGHT_BLUE"
# ch3_coupling = "DCLimit"

# """CHANNEL 4"""
# ch4_scale = 5
# ch4_position = -5
# ch4_bw = 20
# ch4_rel_x_position = 80
# ch4_label = "VOUT"
# ch4_measure = "MAX,MIN,MEAN"
# ch4_color = "PINK"
# ch4_coupling = "DCLimit"
# ch4_offset = 10


"""CHANNEL 1"""
ch1_scale = 0.2
ch1_position = -4
ch1_bw = 20
ch1_rel_x_position = 20 #from left
ch1_label = "DC-DC IOUT" 
ch1_measure = "MAX,MIN,MEAN"
ch1_color = "GREEN"
ch1_coupling = "DCLimit"

"""CHANNEL 2"""
ch2_scale = 5
ch2_position = -4
ch2_bw = 20 #20, 200, 500 in Mhz
ch2_rel_x_position = 40
ch2_label = "LYT8 VOUT"
ch2_measure = "MAX,MIN,MEAN"
ch2_color = "YELLOW"
ch2_coupling = "DCLimit"
ch2_offset = 20

"""CHANNEL 3"""
ch3_scale = 2 # 0.2
ch3_position = -4
ch3_bw = 20
ch3_rel_x_position = 60
ch3_label = "PRI_IDS"
ch3_measure = "MAX,MIN,MEAN"
ch3_color = "LIGHT_BLUE"
ch3_coupling = "DCLimit"

"""CHANNEL 4"""
ch4_scale = 5
ch4_position = -4
ch4_bw = 20
ch4_rel_x_position = 80
ch4_label = "DC-DC VOUT"
ch4_measure = "MAX,MIN,MEAN"
ch4_color = "PINK"
ch4_coupling = "DCLimit"
ch4_offset = 20

"""DO NOT EDIT BELOW THIS LINE"""
####################################################################################################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)


"""GENERIC FUNCTIONS"""

def discharge_output(times):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(1)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    # INCLUDE
    
    scope.channel_settings(state=ch1_enable, channel=1, scale=ch1_scale, position=ch1_position, label=ch1_label,
                            color=ch1_color, rel_x_position=ch1_rel_x_position, bandwidth=ch1_bw, coupling=ch1_coupling, offset=0)
        
    scope.channel_settings(state=ch2_enable, channel=2, scale=ch2_scale, position=ch2_position, label=ch2_label,
                            color=ch2_color, rel_x_position=ch2_rel_x_position, bandwidth=ch2_bw, coupling=ch2_coupling, offset=ch2_offset)
    
    scope.channel_settings(state=ch3_enable, channel=3, scale=ch3_scale, position=ch3_position, label=ch3_label,
                            color=ch3_color, rel_x_position=ch3_rel_x_position, bandwidth=ch3_bw, coupling=ch3_coupling, offset=0)
    
    scope.channel_settings(state=ch4_enable, channel=4, scale=ch4_scale, position=ch4_position, label=ch4_label,
                            color=ch4_color, rel_x_position=ch4_rel_x_position, bandwidth=ch4_bw, coupling=ch4_coupling, offset=ch4_offset)
    
    if ch1_enable != 'OFF': scope.measure(1, ch1_measure)
    if ch2_enable != 'OFF': scope.measure(2, ch2_measure)
    if ch3_enable != 'OFF': scope.measure(3, ch3_measure)
    if ch4_enable != 'OFF': scope.measure(4, ch4_measure)

    scope.record_length(50E6)
    scope.time_position(time_position)
    scope.time_scale(time_scale)

    if zoom_enable:
        scope.remove_zoom()
        scope.add_zoom(rel_pos=zoom_pos, rel_scale=zoom_rel_scale)
    
    trigger_channel = trig_channel
    trigger_level = trig_level
    trigger_edge = trig_edge
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1

def operation():

   

    ac.voltage = 230
    ac.turn_on()
    sleep(2)

    scope.run_single()

    eload.channel[eload_channel].cc = 0
    eload.channel[eload_channel].turn_on()
    sleep(2)


    start = 0
    end = 1
    step = 100E-6
    steps = (end-start)/step
    ramp_time = 100E-3
    time_per_step = ramp_time/steps

    print(time_per_step)


    for i in np.arange(0, 1.1, 0.01):
        eload.channel[eload_channel].cc = i
        eload.channel[eload_channel].turn_on()
        # sleep(0E-6)
            
            

def main():
    global waveform_counter
    discharge_output(times=1)
    # scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)