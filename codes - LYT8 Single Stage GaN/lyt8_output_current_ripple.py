## UPDATED AUGUST 08, 2022

"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.default_scope_settings import *
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	

##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.129"
sig_gen_address = '11'
dimming_power_meter_address = 21
eload_channel = 5

"""USER INPUT"""


vin_list = [90,115,130,180,230,265,277,300]
iout_nom = 1.5
iout_list = [0.25*iout_nom,0.5*iout_nom,0.75*iout_nom,iout_nom]
iout_list = [iout_nom]
iout_channel = 1

test = "OUTPUT CURRENT RIPPLE (Cout=3000uF)"
operation = "CV"
waveforms_folder = f'C:/Users/ccayno/Desktop/APPS EVAL/Single Stage GAN/Test Data ({operation})/{test}'
path = path_maker(f'{waveforms_folder}')


"""SCOPE SETTINGS"""
trig_channel = 1
trig_level = 1.4
trig_edge = 'POS'

time_position = 50
time_scale = 0.002

"""
MEASUREMENT SETTINGS OPTIONS: "MAX,MIN,RMS,MEAN,PDELta"
"""

ch1_enable = 'ON'
ch2_enable = 'OFF'
ch3_enable = 'OFF'
ch4_enable = 'OFF'

"""CHANNEL 1"""
ch1_scale = 0.015
ch1_position = 0
ch1_bw = 20
ch1_rel_x_position = 20
ch1_label = "OUTPUT CURRENT"
ch1_measure = "MAX,MIN,RMS,MEAN,PDELta"
ch1_color = "GREEN"
ch1_coupling = "DCLimit"
ch1_offset = iout_nom

# """CHANNEL 2"""
# ch2_scale = 100
# ch2_position = -4
# ch2_bw = 500
# ch2_rel_x_position = 40
# ch2_label = "PRI_VDS"
# ch2_measure = "MAX"
# ch2_color = "YELLOW"
# ch2_coupling = "DCLimit"

# """CHANNEL 3"""
# ch3_scale = 2
# ch3_position = -4
# ch3_bw = 500
# ch3_rel_x_position = 60
# ch3_label = "PRI_IDS"
# ch3_measure = "MAX"
# ch3_color = "LIGHT_BLUE"
# ch3_coupling = "DCLimit"

# """CHANNEL 4"""
# ch4_scale = 10
# ch4_position = -4
# ch4_bw = 20
# ch4_rel_x_position = 80
# ch4_label = "VOUT"
# ch4_measure = "MAX"
# ch4_color = "PINK"
# ch4_coupling = "DCLimit"


"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output():
    ac.turn_off()
    for i in range(1):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(2)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    
    scope.channel_settings(state=ch1_enable, channel=1, scale=ch1_scale, position=ch1_position, label=ch1_label,
                            color=ch1_color, rel_x_position=ch1_rel_x_position, bandwidth=ch1_bw, coupling=ch1_coupling, offset=ch1_offset)
        
    scope.channel_settings(state=ch2_enable, channel=2, scale=ch2_scale, position=ch2_position, label=ch2_label,
                            color=ch2_color, rel_x_position=ch2_rel_x_position, bandwidth=ch2_bw, coupling=ch2_coupling, offset=0)
    
    scope.channel_settings(state=ch3_enable, channel=3, scale=ch3_scale, position=ch3_position, label=ch3_label,
                            color=ch3_color, rel_x_position=ch3_rel_x_position, bandwidth=ch3_bw, coupling=ch3_coupling, offset=0)
    
    scope.channel_settings(state=ch4_enable, channel=4, scale=ch4_scale, position=ch4_position, label=ch4_label,
                            color=ch4_color, rel_x_position=ch4_rel_x_position, bandwidth=ch4_bw, coupling=ch4_coupling, offset=0)
    
    if ch1_enable != 'OFF': scope.measure(1, ch1_measure)
    if ch2_enable != 'OFF': scope.measure(2, ch2_measure)
    if ch3_enable != 'OFF': scope.measure(3, ch3_measure)
    if ch4_enable != 'OFF': scope.measure(4, ch4_measure)

    scope.record_length(50E6)
    scope.time_position(time_position)
    scope.time_scale(time_scale)

    # scope.remove_zoom()
    # scope.add_zoom(rel_pos=50, rel_scale=0.5)
    
    trigger_channel = trig_channel
    trigger_level = trig_level
    trigger_edge = trig_edge
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()

def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1

def find_trigger(channel, trigger_delta):

    scope.edge_trigger(channel, 0, 'POS')

    # finding trigger level
    scope.run_single()
    scope.trigger_mode('AUTO')
    soak(5)
    scope.trigger_mode('NORM')
    

    # get initial peak-to-peak measurement value
    labels, values = scope.get_measure(channel)
    max_value = float(values[0])
    max_value = float(f"{max_value:.4f}")

    # set max_value as initial trigger level
    trigger_level = max_value
    scope.edge_trigger(channel, trigger_level, 'POS')

    # check if it triggered within 3 seconds
    scope.run_single()
    soak(3)
    trigger_status = scope.trigger_status()

    # increase trigger level until it reaches the maximum trigger level
    while (trigger_status == 1):
        trigger_level += trigger_delta
        scope.edge_trigger(channel, trigger_level, 'POS')

        # check trigger status
        scope.run_single()
        soak(3)
        trigger_status = scope.trigger_status()

    # decrease trigger level below to get the maximum trigger possible
    trigger_level -= 1*trigger_delta
    scope.edge_trigger(channel, trigger_level, 'POS')

    scope.run_single()
    soak(3)
    trigger_status = scope.trigger_status()
    while (trigger_status != 1):
        # decrease trigger level below to get the maximum trigger possible
        trigger_level -= 1*trigger_delta
        scope.edge_trigger(channel, trigger_level, 'POS')

        # check trigger status
        scope.run_single()
        soak(3)
        trigger_status = scope.trigger_status()

def create_output_excel_result(df):
    
    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{test}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    df_to_excel(wb, "Sheet1", df, 'B2')
    wb.save(dst)


def collect_data(iout, v_input):
    
    """COLLECT DATA"""
    load = iout
    vac = v_input
    labels, values = scope.get_measure(iout_channel)
    iout_pkpk = float(values[2])
    iout_mean = float(values[3])
    iout_ripple = 100*iout_pkpk/iout_mean
    iout_flicker = 100*iout_pkpk/(2*iout_mean)
    output_list = [load, vac, iout_pkpk, iout_mean, iout_ripple, iout_flicker]
    
    return output_list

def operation():
    df_header_list = ['LED', 'Vin', 'Ipk-pk (A)', 'Imean (A)', '%Ripple', 'Flicker(%)']
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list

    for iout in iout_list:
        
        for vin in vin_list:

            if iout == 0:
                eload.channel[eload_channel].turn_off()
            else:
                eload.channel[eload_channel].cc = iout
                eload.channel[eload_channel].turn_on()

            scope.channel_offset(iout_channel, iout)



            ac.voltage = vin
            ac.turn_on()
            sleep(5)

            find_trigger(channel=iout_channel, trigger_delta=0.001)
            
            scope.run_single()
            sleep(3)
            

            

            discharge_output()
            percent = int(iout*100/iout_nom)
            filename = f"Iout Ripple - {vin}Vac, {iout}A ({percent}percent CC load).png"
            screenshot(filename, path)


            output_list = collect_data(iout, vin)
            df.loc[len(df)] = output_list

            print(output_list)

            create_output_excel_result(df)

    
    


def main():
    global waveform_counter
    discharge_output()
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)