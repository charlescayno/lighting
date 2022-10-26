"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from tracemalloc import start
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Tektronix_SigGen_AFG31000
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.default_scope_settings import *
import shutil
import numpy as np
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	
import getpass
username = getpass.getuser().lower()
####################################################################################################################################################################

test = "PWM Dimming Transient"
operation = "CC"
# condition = input(">> Type test condition: ")

waveforms_folder = f'C:/Users/{username}/Desktop/APPS EVAL/Single Stage GAN/Test Data ({operation})/{test}/'
path = path_maker(f'{waveforms_folder}')

"""COMMS"""
ac_source_address = 9
source_power_meter_address = 2
load_power_meter_address = 10
eload_address = 12
scope_address = "10.125.10.210"
sig_gen_address = '11'
dimming_power_meter_address = 20
eload_channel = 5
dc_source_address = '4'

vr_channel = 1
vout_channel = 2
dim_channel = 3
iout_channel = 4
# print("CH1: VR | CH2: VOUT | CH3: DIM+ | CH4: IOUT")

"""USER INPUT"""
vin_list = [90,115,180,230,277,300]
vin_list = [115,230]
vout = 50
iout_nom = 1.5
duty_list = [0.005,10,25,50,75,99]
# duty_list = [10,75]
frequency_list = [2000, 400]
led_list = [50,36,24]
zoom_enable = False


"""SCOPE SETTINGS"""
trig_channel = 3
trig_level = 5
trig_edge = 'POS'

time_position = 10
time_scale = 0.5

"""
MEASUREMENT SETTINGS OPTIONS: "MAX,MIN,RMS,MEAN,PDELta"
"""

ch1_enable = 'ON'
ch2_enable = 'ON'
ch3_enable = 'ON'
ch4_enable = 'ON'

"""CHANNEL 1"""
ch1_scale = 5
ch1_position = -2
ch1_bw = 500
ch1_rel_x_position = 10
ch1_label = "HBP"
ch1_measure = "MAX,MIN,MEAN,PDELta,RMS"
ch1_color = "BLUE"
ch1_coupling = "DCLimit"

"""CHANNEL 2"""
ch2_scale = 10
ch2_position = -1
ch2_bw = 20
ch2_rel_x_position = 20
ch2_label = "VOUT"
ch2_measure = "MAX,MIN,RMS"
ch2_color = "LIGHT_BLUE"
ch2_coupling = "DCLimit"

"""CHANNEL 3"""
ch3_scale = 10
ch3_position = -4
ch3_bw = 20
ch3_rel_x_position = 30
ch3_label = "DIM+"
ch3_measure = "MAX,MIN"
ch3_color = "YELLOW"
ch3_coupling = "DCLimit"

"""CHANNEL 4"""
ch4_scale = 0.2
ch4_position = -4.5
ch4_bw = 20
ch4_rel_x_position = 40
ch4_label = "IOUT"
ch4_measure = "MAX,MIN,PDelta,RMS"
ch4_color = "GREEN"
ch4_coupling = "DCLimit"


"""DO NOT EDIT BELOW THIS LINE"""
####################################################################################################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
sig_gen = Tektronix_SigGen_AFG31000(sig_gen_address)
led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output(times):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(1)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    
    scope.channel_settings(state=ch1_enable, channel=1, scale=ch1_scale, position=ch1_position, label=ch1_label,
                            color=ch1_color, rel_x_position=ch1_rel_x_position, bandwidth=ch1_bw, coupling=ch1_coupling, offset=0)
        
    scope.channel_settings(state=ch2_enable, channel=2, scale=ch2_scale, position=ch2_position, label=ch2_label,
                            color=ch2_color, rel_x_position=ch2_rel_x_position, bandwidth=ch2_bw, coupling=ch2_coupling, offset=0)
    
    scope.channel_settings(state=ch3_enable, channel=3, scale=ch3_scale, position=ch3_position, label=ch3_label,
                            color=ch3_color, rel_x_position=ch3_rel_x_position, bandwidth=ch3_bw, coupling=ch3_coupling, offset=0)
    
    scope.channel_settings(state=ch4_enable, channel=4, scale=ch4_scale, position=ch4_position, label=ch4_label,
                            color=ch4_color, rel_x_position=ch4_rel_x_position, bandwidth=ch4_bw, coupling=ch4_coupling, offset=0)
    
    if ch1_enable != 'OFF': scope.measure(1, ch1_measure)
    if ch2_enable != 'OFF': scope.measure(2, ch2_measure)
    if ch3_enable != 'OFF': scope.measure(3, ch3_measure)
    if ch4_enable != 'OFF': scope.measure(4, ch4_measure)

    scope.record_length(50E6)
    scope.time_position(time_position)
    scope.time_scale(time_scale)

    # scope.remove_zoom()
    # if zoom_enable == True:
    #     scope.add_zoom(rel_pos=50, rel_scale=0.5)
    
    trigger_channel = trig_channel
    trigger_level = trig_level
    trigger_edge = trig_edge
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1

def roundup(x):
    return int(math.ceil(x / 100.0)) * 100

#################### DATA EXPORT CODE ###########################################################
def create_header_list(df_header_list):
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list
    print(df)
    return df

def export_to_excel(df, output_list, excel_name, sheet_name, anchor):
    df.loc[len(df)] = output_list
    print(output_list)

    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    df_to_excel(wb, sheet_name, df, anchor)
    wb.save(dst)

def export_screenshot_to_excel(excel_name, waveforms_folder, sheet_name, filename, anchor):

    src = f"{os.getcwd()}/pwm_dimming_transient_template.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    try:
        if not os.path.exists(dst): shutil.copyfile(src, dst)
    except:
        pass

    wb = load_workbook(dst)
    image_to_excel(wb, sheet_name, filename=filename, folder_path=waveforms_folder, anchor=anchor)
    wb.save(dst)

##################################################################################################


def set_pwm(frequency, duty):
    if duty == 100: duty = 99
    if duty == 0: duty = 0.005
    sig_gen.set_load_impedance(channel = 1, impedance = 'INF') 
    sig_gen.out_cont_pulse(channel = 1, frequency = frequency, phase = '0 DEG', low = '0V', high = '10V', units = 'VPP', duty = duty, width = 'ABC')
    sig_gen.channel_state(channel = 1, state ='ON')
    soak(2)

def reset_pwm(frequency):
    set_pwm(frequency, 0.005)
    soak(3)

def pwm_width(f=2000, pwm_start=0, pwm_end=100):
    
    T = 1/int(f)

    Ton1 = T*(pwm_start/100)
    Ton2 = T*(pwm_end/100)

    if Ton1 < Ton2:
        width = Ton1
        width_range = 'LONG'
    else:
        width = Ton2
        width_range = 'WITH'

    return width, width_range




def pwm_dimming_transient(start_duty, end_duty, vin, led, condition, pwm_frequency, col, row):

    set_pwm(pwm_frequency, start_duty)

    ac.voltage = vin
    ac.turn_on()

    sleep(3)

    scope.run_single()

    sleep(1)

    set_pwm(pwm_frequency, end_duty)

    sleep(6)

    scope.stop()

    filename = f'{led}V, {vin}Vac {pwm_frequency}Hz, {start_duty}-{end_duty}% PWM Dimming, f={pwm_frequency}Hz.png' if condition=="NA" else f'{led}V, {vin}Vac {pwm_frequency}Hz, {start_duty}-{end_duty}% PWM Dimming, f={pwm_frequency}Hz, {condition}.png'
    screenshot(filename, path)
    excel_name = f"{test} f={pwm_frequency}Hz"
    export_screenshot_to_excel(excel_name, waveforms_folder, f"{led}V, {vin}Vac", filename, f"{col}{row}")
    

def operation():

    global row

    for pwm_frequency in frequency_list:

        for led in led_list:

            led_ctrl.voltage(led)


            for vin in vin_list:

                row = 2
                duty_list.sort()
                for i in range(len(duty_list)):
                    for duty in duty_list:
                        if duty_list[i] < duty:
                                start_duty = duty_list[i]
                                end_duty = duty
                                # print(start_duty, end_duty)

                                """Adjust Trigger settings"""
                                width, width_range = pwm_width(pwm_frequency, start_duty, end_duty)
                                delta = width*0.03
                                scope.width_trigger(trigger_channel=dim_channel, width_polarity='POS', width_range=width_range, width=width, delta=delta)

                                pwm_dimming_transient(start_duty, end_duty, vin, led, condition="First Power Up Discharged Cout", pwm_frequency=pwm_frequency, col='B', row=row)
                                pwm_dimming_transient(start_duty, end_duty, vin, led, condition="No AC reset (Not Discharged Cout)", pwm_frequency=pwm_frequency, col='S', row=row)
                                row += 37

                                discharge_output(times=2)
                                
                row = 2
                duty_list.sort(reverse=True)
                for i in range(len(duty_list)):
                    for duty in duty_list:
                        if duty_list[i] > duty:
                                start_duty = duty_list[i]
                                end_duty = duty
                                print(start_duty, end_duty)


                                """adjust trigger"""
                                width, width_range = pwm_width(pwm_frequency, start_duty, end_duty)
                                delta = width*0.03
                                scope.width_trigger(trigger_channel=dim_channel, width_polarity='POS', width_range=width_range, width=width, delta=delta)

                                pwm_dimming_transient(start_duty, end_duty, vin, led, condition="NA", pwm_frequency=pwm_frequency, col='AJ', row=row)
                                row += 37
                                discharge_output(times=2)

def main():
    global waveform_counter
    discharge_output(times=1)
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)