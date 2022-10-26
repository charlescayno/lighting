"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.default_scope_settings import *
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	

import getpass
username = getpass.getuser().lower()
####################################################################################################################################################################



test = "AC Cycling (no zener)"
operation = "CV"
pulse_count = 5
condition = sys.argv[1]
pulse = float(sys.argv[2])
load_type = sys.argv[3]
excel_name = f"{test}, {operation}, {condition}, {pulse}s on-off, {pulse_count} pulses, NL"


waveforms_folder = f'C:/Users/{username}/Desktop/APPS EVAL/Single Stage GAN/Test Data ({operation})/{test}/{condition}/{load_type}/{pulse} s'
path = path_maker(f'{waveforms_folder}')

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.129"
sig_gen_address = '11'
dimming_power_meter_address = 21
eload_channel = 5

"""USER INPUT"""



vin_list = [90,115,130,180,230,265,277,300]
# vin_list = [265,277,300]
vout = 50
iout_nom = 1.5

# iout_list = [0.25*iout_nom,0.5*iout_nom,0.75*iout_nom,iout_nom]
iout_list = [0]

# cr_load = vout/iout_nom
# cr_list = [cr_load/0.25,cr_load/0.5,cr_load/0.75,cr_load]
cr_list = [0]

iout_channel = 1
vin_channel = 2
ids_channel = 3
vout_channel = 4



"""BROWN-IN SETTINGS"""
start_voltage = 0
end_voltage = 277
slew_rate = 0.5
frequency = 60
time_fixvoltage = 60



"""SCOPE SETTINGS"""
time_division = pulse
time_scale = 2*time_division
start_soak = 5*time_scale
off_time = pulse
on_time = pulse
end_soak = 4*time_scale
delay = start_soak + end_soak + (pulse_count*(off_time + on_time)) + off_time


trig_channel = 3
trig_level = 4
trig_edge = 'NEG'

time_position = 20

"""ZOOM SETTINGS"""
zoom_enable = False
zoom_pos = 50
zoom_rel_scale = 1


"""
MEASUREMENT SETTINGS OPTIONS: "MAX,MIN,RMS,MEAN,PDELta"
"""

ch1_enable = 'OFF'
ch2_enable = 'ON'
ch3_enable = 'ON'
ch4_enable = 'ON'

"""CHANNEL 1"""
ch1_scale = 0.3
ch1_position = -4
ch1_bw = 20
ch1_rel_x_position = 20
ch1_label = "IOUT"
ch1_measure = "MAX,MIN"
ch1_color = "GREEN"
ch1_coupling = "DCLimit"

"""CHANNEL 2"""
ch2_scale = 300
ch2_position = 3
ch2_bw = 20
ch2_rel_x_position = 40
ch2_label = "VIN"
ch2_measure = "MAX,RMS"
ch2_color = "YELLOW"
ch2_coupling = "AC"

"""CHANNEL 3"""
ch3_scale = 1
ch3_position = -4
ch3_bw = 500
ch3_rel_x_position = 60
ch3_label = "PRI_IDS"
ch3_measure = "MAX"
ch3_color = "LIGHT_BLUE"
ch3_coupling = "DCLimit"

"""CHANNEL 4"""
ch4_scale = 2
ch4_position = 0
ch4_bw = 20
ch4_rel_x_position = 80
ch4_label = "VOUT"
ch4_measure = "MAX,MIN"
ch4_color = "PINK"
ch4_coupling = "DCLimit"
ch4_offset = 52


"""DO NOT EDIT BELOW THIS LINE"""
####################################################################################################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
# led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output(times):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(1)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    # INCLUDE
    
    scope.channel_settings(state=ch1_enable, channel=1, scale=ch1_scale, position=ch1_position, label=ch1_label,
                            color=ch1_color, rel_x_position=ch1_rel_x_position, bandwidth=ch1_bw, coupling=ch1_coupling, offset=0)
        
    scope.channel_settings(state=ch2_enable, channel=2, scale=ch2_scale, position=ch2_position, label=ch2_label,
                            color=ch2_color, rel_x_position=ch2_rel_x_position, bandwidth=ch2_bw, coupling=ch2_coupling, offset=0)
    
    scope.channel_settings(state=ch3_enable, channel=3, scale=ch3_scale, position=ch3_position, label=ch3_label,
                            color=ch3_color, rel_x_position=ch3_rel_x_position, bandwidth=ch3_bw, coupling=ch3_coupling, offset=0)
    
    scope.channel_settings(state=ch4_enable, channel=4, scale=ch4_scale, position=ch4_position, label=ch4_label,
                            color=ch4_color, rel_x_position=ch4_rel_x_position, bandwidth=ch4_bw, coupling=ch4_coupling, offset=ch4_offset)
    
    if ch1_enable != 'OFF': scope.measure(1, ch1_measure)
    if ch2_enable != 'OFF': scope.measure(2, ch2_measure)
    if ch3_enable != 'OFF': scope.measure(3, ch3_measure)
    if ch4_enable != 'OFF': scope.measure(4, ch4_measure)

    scope.record_length(50E6)
    scope.time_position(time_position)
    scope.time_scale(time_scale)

    if zoom_enable:
        scope.remove_zoom()
        scope.add_zoom(rel_pos=zoom_pos, rel_scale=zoom_rel_scale)
    
    trigger_channel = trig_channel
    trigger_level = trig_level
    trigger_edge = trig_edge
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1

def browning(start, end, slew, frequency):

    if start > end:
        print(f"brownout: {start} -> {end} Vac")
        for voltage in np.arange(start, end+1, -slew):
            ac.voltage = voltage
            ac.frequency = frequency
            ac.turn_on()
            sleep(1)
    if start < end:
        print(f"brownin: {start} -> {end} Vac")
        for voltage in np.arange(start, end+1, slew):
            ac.voltage = voltage
            ac.frequency = frequency
            ac.turn_on()
            sleep(1)

def fixvoltage(voltage, soak):
    ac.voltage = voltage
    ac.frequency = ac.set_freq(voltage)
    ac.turn_on()
    sleep(soak)

def roundup(x):
    return int(math.ceil(x / 100.0)) * 100

def brown_in(led):

    test_time = ((end_voltage-start_voltage)/slew_rate)+time_fixvoltage
    scope_time = roundup(test_time)
    delay = scope_time-test_time
    time_scale = scope_time/10
    print(f"Estimated time: {scope_time/60} mins.")
    scope.time_scale(time_scale)
    scope.run()
    soak(int(delay))

    # start of test
    browning(start_voltage, end_voltage, slew_rate, frequency)
    fixvoltage(end_voltage,time_fixvoltage)
    scope.stop()

    discharge_output()

    filename = f"{led}V, {start_voltage}-{end_voltage} Vac, {slew_rate}V per s.png"
    screenshot(filename, path)

def brown_out(led):

    test_time = ((end_voltage-start_voltage)/slew_rate)+time_fixvoltage
    scope_time = roundup(test_time)
    delay = scope_time-test_time
    time_scale = scope_time/10
    print(f"Estimated time: {scope_time/60} mins.")
    scope.time_scale(time_scale)
    scope.run()
    
    # start of test
    fixvoltage(end_voltage, time_fixvoltage)
    browning(end_voltage, start_voltage, slew_rate, frequency)
    soak(int(delay))
    scope.stop()
    discharge_output()

    filename = f"{led}V, {end_voltage}-{start_voltage} Vac, {slew_rate}V per s.png"
    screenshot(filename, path)

def brown_in_brown_out(led):

    test_time = 2*((end_voltage-start_voltage)/slew_rate)
    scope_time = roundup(test_time)
    delay = scope_time-test_time
    time_scale = scope_time/10
    print(f"Estimated time: {scope_time/60} mins.")
    scope.time_scale(time_scale)
    scope.run()
    soak(int(delay/2))
    # start of test
    browning(start_voltage, end_voltage, slew_rate, frequency)
    browning(end_voltage, start_voltage, slew_rate, frequency)
    soak(int(delay/2))
    scope.stop()
    discharge_output()

    filename = f"{led}V, {start_voltage}-{end_voltage}-{start_voltage} Vac, {slew_rate}V per s.png"
    screenshot(filename, path)

def brown_out_brown_in(led):

    test_time = 2*((end_voltage-start_voltage)/slew_rate)
    scope_time = roundup(test_time)
    delay = scope_time-test_time
    time_scale = scope_time/10
    print(f"Estimated time: {scope_time/60} mins.")
    scope.time_scale(time_scale)
    scope.run()
    t_fixvoltage = int(delay/2)
    # start of test
    fixvoltage(end_voltage, t_fixvoltage)
    browning(end_voltage, start_voltage, slew_rate, frequency)
    browning(start_voltage, end_voltage, slew_rate, frequency)
    fixvoltage(end_voltage, t_fixvoltage)
    scope.stop()
    discharge_output()

    filename = f"{led}V, {end_voltage}-{start_voltage}-{end_voltage} Vac, {slew_rate}V per s.png"
    screenshot(filename, path)

def loadtransient_cc(iout_nom, start_percent, end_percent, transient_frequency):

    start = iout_nom*start_percent/100
    end = iout_nom*end_percent/100
    ton = 1 / transient_frequency
    toff = ton
    scope.time_scale(ton)
    eload.channel[eload_channel].dynamic(start, end, ton, toff)
    eload.channel[eload_channel].turn_on()

def create_output_excel_result(df):
    
    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    df_to_excel(wb, "Sheet1", df, 'B2')
    wb.save(dst)

def collect_data(vin, iout):

    ## EDIT THIS FUNCTION DEPENDING ON TEST
    
    """COLLECT DATA"""
    vac = vin


    labels, values = scope.get_measure(vout_channel)
    vout_max = float(values[0])
    vout_min = float(values[1])

    iout_max = 0

    labels, values = scope.get_measure(ids_channel)
    ids_max = float(values[0])


    vout_overshoot = abs(100*(vout_max-vout)/vout)
    vout_undershoot = abs(100*(vout-vout_min)/vout)

    if vout_overshoot > 5: vout_overshoot_result = 'FAIL'
    else: vout_overshoot_result = 'PASS'
    if vout_undershoot > 5: vout_undershoot_result = 'FAIL'
    else: vout_undershoot_result = 'PASS'


    if load_type == 'cc':
        load_percent = int(iout*100/iout_nom)
        load_value = float(f"{iout:.2f}")

    if load_type == 'cr':
        load_percent = 0
        load_value = float(f"{iout:.2f}")

    output_list = [pulse, load_percent, load_value, vac,
                    vout_max, vout_overshoot, vout_overshoot_result,
                    vout_min, vout_undershoot, vout_undershoot_result,
                    iout_max, ids_max]
    
    return output_list


def export_to_excel(df, vin, iout):

    ## EDIT THIS FUNCTION DEPENDING ON TEST

    output_list = collect_data(vin, iout)
    df.loc[len(df)] = output_list
    print(output_list)
    print()
    create_output_excel_result(df)



def operation():


    



    if load_type == 'cr':

        df_header_list = ['Pulse Time (s)', 'CR Load (%)', 'CR Load (ohms)', 'Vin (VAC)',
                        'Vout_max (V)', 'Overshoot (%)', 'PASS/FAIL',
                        'Vout_min (V)', 'Undershoot (%)', 'PASS/FAIL',
                        'Iout_max (A)', 'Ids_max (A)']
        df = pd.DataFrame(columns = df_header_list)
        df.loc[len(df)] = df_header_list



        for cr in cr_list:
            
            for vin in vin_list:


                if vin == 90 or vin == 115:
                    scope.channel_offset(vout_channel, 50)
                else:
                    scope.channel_offset(vout_channel, 52)
                if vin > 265:
                    scope.edge_trigger(ids_channel, 4, 'POS')
                else:
                    scope.edge_trigger(ids_channel, 2, 'POS')

                eload.channel[eload_channel].turn_off()

                
                ac.voltage = vin
                ac.turn_on()

                soak(3)


                scope.run_single()


                ac.ac_cycling(pulse_count, vin, start_soak, off_time, on_time, end_soak)
                # scope.stop()
                discharge_output(times=3)

                filename = f"AC Cyling, {pulse}s On-Off, {pulse_count} pulses - {vin}Vac, {cr}Ohms (0 CR load).png"
                screenshot(filename, path)

                export_to_excel(df, vin, cr)


                

    if load_type == 'cc':

        df_header_list = ['Pulse Time (s)', 'CC Load (%)', 'CC Load (A)', 'Vin (VAC)',
                        'Vout_max (V)', 'Overshoot (%)', 'PASS/FAIL',
                        'Vout_min (V)', 'Undershoot (%)', 'PASS/FAIL',
                        'Iout_max (A)', 'Ids_max (A)']
        df = pd.DataFrame(columns = df_header_list)
        df.loc[len(df)] = df_header_list


        for iout in iout_list:

            for vin in vin_list:

                if vin == 90 or vin == 115:
                    scope.channel_offset(vout_channel, 50)
                else:
                    scope.channel_offset(vout_channel, 52)
                if vin > 230:
                    scope.edge_trigger(ids_channel, 3, 'POS')
                else:
                    scope.edge_trigger(ids_channel, 2, 'POS')

                eload.channel[eload_channel].turn_off()

                soak(1)
                print("off eload")
                
                
                ac.voltage = vin
                ac.turn_on()

                soak(3)

                scope.run_single()
                ac.voltage = vin
                ac.turn_on()
                ac.ac_cycling(pulse_count, vin, start_soak, off_time, on_time, end_soak)
                # scope.stop()
                discharge_output(times=3)

                percent = int(iout*100/iout_nom)
                filename = f"AC Cyling, {pulse}s On-Off, {pulse_count} pulses - {vin}Vac, {iout}A ({percent} CC load).png"
                screenshot(filename, path)

                export_to_excel(df, vin, iout)

            

def main():
    global waveform_counter
    discharge_output(times=1)
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)