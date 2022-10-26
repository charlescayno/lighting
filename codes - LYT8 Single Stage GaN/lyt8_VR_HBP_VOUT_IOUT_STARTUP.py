"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from tracemalloc import start
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.default_scope_settings import *
import shutil
import numpy as np
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
date = now.strftime('%m%d')	
import getpass
username = getpass.getuser().lower()

from powi.equipment import create_header_list, export_to_excel, export_screenshot_to_excel
####################################################################################################################################################################

# condition = input("Input test condition: ")
# condition = "CC point = 1.82 A, 20 V Zener, with linear regulator, no output zener, Lmag = 106 uH, CR Load"
condition = input(">> Condition: ")
load_type = sys.argv[1]
excel_name = f"{condition}"
# excel_name = "summary"
test = f"Startup VOUT HBP VR IDS"
operation = "CV"
waveforms_folder = f'C:/Users/{username}/Desktop/APPS EVAL/Single Stage GAN/Test Data ({operation})/{date}/{test}/{condition}'
path = path_maker(f'{waveforms_folder}')

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.129"
sig_gen_address = '11'
dimming_power_meter_address = 21
eload_channel = 5


iout_channel = 1
vin_channel = 2
ids_channel = 3
vout_channel = 4


"""USER INPUT"""
vout = 50
iout_nom = 1.5

vin_list = [90,115,130,180,230,265,277,300]
percent_list = [1, 0.75, 0.5, 0.25, 0.1, 0]

von_list = [0]
zoom_enable = True

iout_list = [percent * iout_nom for percent in percent_list]
cr_load = vout/iout_nom
cr_list = [(cr_load/percent if percent != 0 else 0) for percent in percent_list]







"""SCOPE SETTINGS"""
trig_channel = 3
trig_level = 2
trig_edge = 'POS'

time_position = 10
time_scale = 0.1

zoom_enable = False


ch1_enable = 'ON'
ch2_enable = 'ON'
ch3_enable = 'ON'
ch4_enable = 'ON'

"""CHANNEL 1"""
ch1_scale = 10
ch1_position = -4
ch1_bw = 20
ch1_rel_x_position = 20
ch1_label = "VOUT" 
ch1_measure = "MAX,MIN,MEAN"
ch1_color = "GREEN"
ch1_coupling = "DCLimit"

"""CHANNEL 2"""
ch2_scale = 5
ch2_position = -4
ch2_bw = 20
ch2_rel_x_position = 40
ch2_label = "HBP"
ch2_measure = "MAX,MIN,MEAN"
ch2_color = "YELLOW"
ch2_coupling = "DCLimit"

"""CHANNEL 3"""
ch3_scale = 10
ch3_position = -4
ch3_bw = 20
ch3_rel_x_position = 60
ch3_label = "VR"
ch3_measure = "MAX,MIN,MEAN"
ch3_color = "LIGHT_BLUE"
ch3_coupling = "DCLimit"

"""CHANNEL 4"""
ch4_scale = 2
ch4_position = -4
ch4_bw = 500
ch4_rel_x_position = 80
ch4_label = "IDS"
ch4_measure = "MAX,MIN,MEAN"
ch4_color = "PINK"
ch4_coupling = "DCLimit"



"""DO NOT EDIT BELOW THIS LINE"""
####################################################################################################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)

"""GENERIC FUNCTIONS"""

def discharge_output(times):    
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cr = 10
            eload.channel[i].turn_on()
        sleep(0.5)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(0.5)

        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].short_on()
            eload.channel[i].turn_on()
        sleep(0.5)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(0.5)

def scope_settings():
    
    scope.channel_settings(state=ch1_enable, channel=1, scale=ch1_scale, position=ch1_position, label=ch1_label,
                            color=ch1_color, rel_x_position=ch1_rel_x_position, bandwidth=ch1_bw, coupling=ch1_coupling, offset=0)
        
    scope.channel_settings(state=ch2_enable, channel=2, scale=ch2_scale, position=ch2_position, label=ch2_label,
                            color=ch2_color, rel_x_position=ch2_rel_x_position, bandwidth=ch2_bw, coupling=ch2_coupling, offset=0)
    
    scope.channel_settings(state=ch3_enable, channel=3, scale=ch3_scale, position=ch3_position, label=ch3_label,
                            color=ch3_color, rel_x_position=ch3_rel_x_position, bandwidth=ch3_bw, coupling=ch3_coupling, offset=0)
    
    scope.channel_settings(state=ch4_enable, channel=4, scale=ch4_scale, position=ch4_position, label=ch4_label,
                            color=ch4_color, rel_x_position=ch4_rel_x_position, bandwidth=ch4_bw, coupling=ch4_coupling, offset=0)
    
    if ch1_enable != 'OFF': scope.measure(1, ch1_measure)
    if ch2_enable != 'OFF': scope.measure(2, ch2_measure)
    if ch3_enable != 'OFF': scope.measure(3, ch3_measure)
    if ch4_enable != 'OFF': scope.measure(4, ch4_measure)

    scope.record_length(50E6)
    scope.time_position(time_position)
    scope.time_scale(time_scale)

    # scope.remove_zoom()
    # scope.add_zoom(rel_pos=50, rel_scale=0.04)
    
    trigger_channel = trig_channel
    trigger_level = trig_level
    trigger_edge = trig_edge
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1

def operation():


    if load_type == 'cc':
        for von in von_list:

            for iout in iout_list:
                
                for vin in vin_list:

                    eload.channel[eload_channel].von = von

                    if iout == 0:
                        eload.channel[eload_channel].turn_off()

                    else:
                        eload.channel[eload_channel].von = von
                        eload.channel[eload_channel].cc = iout
                        eload.channel[eload_channel].turn_on()

                    scope.run_single()
                    sleep(3)

                    ac.voltage = vin
                    ac.turn_on()
                    sleep(5)
                    
                    discharge_output(times=2)

                    percent = int(iout*100/iout_nom)



                    filename = f"Startup - {vin}Vac, {iout}A ({percent}percent CC load), Von = {von} V.png"
                    
                    screenshot(filename, path)

    if load_type == 'cr':

        for von in von_list:

            for cr in cr_list:

                for vin in vin_list:

                    if cr == 0:
                        eload.channel[eload_channel].turn_off()

                    else:
                        eload.channel[eload_channel].cr = cr
                        eload.channel[eload_channel].turn_on()

                    

                    scope.run_single()
                    sleep(3)

                    ac.voltage = vin
                    ac.turn_on()
                    sleep(5)
                    
                    discharge_output(times=2)

                    try: 
                        percent = int(cr_load*100/cr)
                    except:
                        percent = 0
                    crr = float(f"{cr:.2f}")
                    filename = f"Startup - {vin}Vac, {crr}Ohms ({percent} CR load).png"
                    
                    screenshot(filename, path)

def main():
    global waveform_counter
    discharge_output(times=1)
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)