# input("WARNING CODE IS NOT FINISHED.")


"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from tracemalloc import start
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.default_scope_settings import *
import shutil
import openpyxl
import numpy as np
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	
import getpass
username = getpass.getuser().lower()
####################################################################################################################################################################


operation = "CV"
test = "Startup Short RFBL"
condition = "Disabled CC (Rls = 7.1 MOhms)"

load_type = sys.argv[1]
excel_name = f"{test}, {operation}, {condition}, {load_type}"
waveforms_folder = f'C:/Users/{username}/Desktop/APPS EVAL/Single Stage GAN/Test Data ({operation})/{test}/{condition}/{load_type}'
path = path_maker(f'{waveforms_folder}')

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.129"
sig_gen_address = '11'
dimming_power_meter_address = 21
eload_channel = 5

"""USER INPUT"""

vout = 50
iout_nom = 1.5
iout_list = [0, 0.25*iout_nom,0.5*iout_nom,0.75*iout_nom,iout_nom]
iout_list = [iout_nom]
iout_list = [0]


cr_load = vout/iout_nom
cr_list = [0, cr_load/0.25,cr_load/0.5,cr_load/0.75,cr_load]
cr_list = [cr_load]
# cr_list = [0, cr_load]


iout_channel = 1
vin_channel = 2
ids_channel = 3
vout_channel = 4


"""SCOPE SETTINGS"""
trig_channel = 4
trig_level = 40
trig_edge = 'NEG'

time_position = 10
time_scale = 1

"""ZOOM SETTINGS"""
zoom_enable = False
zoom_pos = 50
zoom_rel_scale = 1


"""BROWN-IN SETTINGS"""
start_voltage = 0
end_voltage = 277
slew_rate = 0.5
frequency = 60
time_fixvoltage = 60




"""
MEASUREMENT SETTINGS OPTIONS: "MAX,MIN,RMS,MEAN,PDELta"
"""

ch1_enable = 'ON'
ch2_enable = 'ON'
ch3_enable = 'ON'
ch4_enable = 'ON'

"""CHANNEL 1"""
ch1_scale = 0.5
ch1_position = -2
ch1_bw = 20
ch1_rel_x_position = 20
ch1_label = "IOUT"
ch1_measure = "MAX,MIN"
ch1_color = "GREEN"
ch1_coupling = "DCLimit"
ch1_offset = 0

"""CHANNEL 2"""
ch2_scale = 150
ch2_position = 0
ch2_bw = 20
ch2_rel_x_position = 40
ch2_label = "VIN"
ch2_measure = "MAX,RMS"
ch2_color = "YELLOW"
ch2_coupling = "AC"
ch2_offset = 0


"""CHANNEL 3"""
ch3_scale = 2
ch3_position = -4
ch3_bw = 500
ch3_rel_x_position = 60
ch3_label = "PRI_IDS"
ch3_measure = "MAX,MIN"
ch3_color = "LIGHT_BLUE"
ch3_coupling = "DCLimit"
ch3_offset = 0

"""CHANNEL 4"""
ch4_scale = 10
ch4_position = -4
ch4_bw = 20
ch4_rel_x_position = 80
ch4_label = "VOUT"
ch4_measure = "MAX,MIN"
ch4_color = "PINK"
ch4_coupling = "DCLimit"
ch4_offset = 0


"""DO NOT EDIT BELOW THIS LINE"""
####################################################################################################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
# led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output(times):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(1)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    # INCLUDE
    
    scope.channel_settings(state=ch1_enable, channel=1, scale=ch1_scale, position=ch1_position, label=ch1_label,
                            color=ch1_color, rel_x_position=ch1_rel_x_position, bandwidth=ch1_bw, coupling=ch1_coupling, offset=ch1_offset)
        
    scope.channel_settings(state=ch2_enable, channel=2, scale=ch2_scale, position=ch2_position, label=ch2_label,
                            color=ch2_color, rel_x_position=ch2_rel_x_position, bandwidth=ch2_bw, coupling=ch2_coupling, offset=ch2_offset)
    
    scope.channel_settings(state=ch3_enable, channel=3, scale=ch3_scale, position=ch3_position, label=ch3_label,
                            color=ch3_color, rel_x_position=ch3_rel_x_position, bandwidth=ch3_bw, coupling=ch3_coupling, offset=ch3_offset)
    
    scope.channel_settings(state=ch4_enable, channel=4, scale=ch4_scale, position=ch4_position, label=ch4_label,
                            color=ch4_color, rel_x_position=ch4_rel_x_position, bandwidth=ch4_bw, coupling=ch4_coupling, offset=ch4_offset)
    
    if ch1_enable != 'OFF': scope.measure(1, ch1_measure)
    if ch2_enable != 'OFF': scope.measure(2, ch2_measure)
    if ch3_enable != 'OFF': scope.measure(3, ch3_measure)
    if ch4_enable != 'OFF': scope.measure(4, ch4_measure)

    scope.record_length(50E6)
    scope.time_position(time_position)
    scope.time_scale(time_scale)

    if zoom_enable:
        scope.remove_zoom()
        scope.add_zoom(rel_pos=zoom_pos, rel_scale=zoom_rel_scale)
    
    trigger_channel = trig_channel
    trigger_level = trig_level
    trigger_edge = trig_edge
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    # print(filename)
    waveform_counter += 1

def browning(start, end, slew, frequency):

    if start > end:
        print(f"brownout: {start} -> {end} Vac")
        for voltage in np.arange(start, end+1, -slew):
            ac.voltage = voltage
            ac.frequency = frequency
            ac.turn_on()
            sleep(1)
    if start < end:
        print(f"brownin: {start} -> {end} Vac")
        for voltage in np.arange(start, end+1, slew):
            ac.voltage = voltage
            ac.frequency = frequency
            ac.turn_on()
            sleep(1)

def fixvoltage(voltage, soak):
    ac.voltage = voltage
    ac.frequency = ac.set_freq(voltage)
    ac.turn_on()
    sleep(soak)

def roundup(x):
    return int(math.ceil(x / 100.0)) * 100

def create_output_excel_result(df):
    
    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    df_to_excel(wb, "Sheet1", df, 'B2')
    wb.save(dst)

def collect_data(end_voltage, load):

    ## EDIT THIS FUNCTION DEPENDING ON TEST

    """COLLECT DATA"""

    input_step = f"{end_voltage}"

    labels, values = scope.get_measure(vout_channel)
    vout_max = round(float(values[0]), 4)
    vout_min = round(float(values[1]), 4)

    labels, values = scope.get_measure(iout_channel)
    iout_max = round(float(values[0]), 4)

    labels, values = scope.get_measure(ids_channel)
    ids_max = round(float(values[0]), 4)

    vout_overshoot = abs(100*(vout_max-vout)/vout)
    vout_undershoot = abs(100*(vout-vout_min)/vout)

    if vout_overshoot > 8: vout_overshoot_result = 'FAIL'
    else: vout_overshoot_result = 'PASS'
    if vout_undershoot > 8: vout_undershoot_result = 'FAIL'
    else: vout_undershoot_result = 'PASS'

    if load_type == 'cc':
        load_percent = int(load*100/iout_nom)
        load_value = float(f"{load:.2f}")

    if load_type == 'cr':
        try:
            load_percent = int(cr_load*100/load)
            load_value = float(f"{load:.2f}")
        except:
            load_percent = 0
            load_value = 0

    output_list = [input_step, load_percent, load_value,
                    vout_max, vout_overshoot, vout_overshoot_result,
                    vout_min, vout_undershoot, vout_undershoot_result,
                    iout_max, ids_max]
    
    return output_list

def export_to_excel(df, end_voltage, load):

    ## EDIT THIS FUNCTION DEPENDING ON TEST

    output_list = collect_data(end_voltage, load)
    df.loc[len(df)] = output_list
    print(output_list)
    create_output_excel_result(df)

def startup_short_rfbl(df, end_voltage, load):

    scope.run_single()


    if load == 0:
        eload.channel[eload_channel].turn_off()
        filename = f"{test}, {end_voltage}Vac, NL.png"
    else:
        if load_type == 'cc':
            eload.channel[eload_channel].von = 0
            eload.channel[eload_channel].cc = load
            eload.channel[eload_channel].turn_on()
            percent = int(load*100/iout_nom)
            filename = f"{test}, {end_voltage}Vac, {load} A ({percent} CC load).png"
        if load_type == 'cr':
            eload.channel[eload_channel].cr = load
            eload.channel[eload_channel].turn_on()
            percent = int(cr_load*100/load)
            filename = f"{test}, {end_voltage}Vac, {load:.2f} Ohms ({percent} CR load).png"


    fixvoltage(end_voltage, 12)
    ac.turn_off()
    soak(5)

    screenshot(filename, path)


def line_ovp_startup_at_high_input(df, end_voltage, load):

    scope.time_position(10)
    scope.time_scale(0.2)
    scope.run_single()
    scope.edge_trigger(vin_channel, end_voltage, 'POS')

    if load == 0:
        eload.channel[eload_channel].turn_off()
        filename = f"{test}, {end_voltage}Vac, NL.png"
    else:
        if load_type == 'cc':
            eload.channel[eload_channel].von = 0
            eload.channel[eload_channel].cc = load
            eload.channel[eload_channel].turn_on()
            percent = int(load*100/iout_nom)
            filename = f"{test}, {end_voltage}Vac, {load} A ({percent} CC load).png"
        if load_type == 'cr':
            eload.channel[eload_channel].cr = load
            eload.channel[eload_channel].turn_on()
            percent = int(cr_load*100/load)
            filename = f"{test}, {end_voltage}Vac, {load:.2f} Ohms ({percent} CR load).png"

    soak(3)
    fixvoltage(end_voltage, 3)
    ac.turn_off()
    soak(5)

    input(">> Capture waveform?")
    screenshot(filename, path)
    
    i = 0
    while '' == input(">> Press ENTER to continue capturing screenshot, other keys to end loop."):
        if i == 0: filename = filename.split('.png')[0] + f"_{i}.png"
        else: filename = filename.split('_')[0] + f"_{i}.png"
        print(filename)
        i += 1
        screenshot(filename, path)

    export_to_excel(df, end_voltage, load)
    discharge_output(times=2)




def operation():

    df_header_list = ['Vin Sweep (VAC)', 'Load (%)', 'Load',
                        'Vout_max (V)', 'Overshoot (%)', 'PASS/FAIL',
                        'Vout_min (V)', 'Undershoot (%)', 'PASS/FAIL',
                        'Iout_max (A)', 'Ids_max (A)']
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list
    print(df)
    
    if load_type == 'cc':
        for iout in iout_list:
            line_ovp_startup_at_high_input(df, end_voltage=270, load=iout)
            line_ovp_startup_at_high_input(df, end_voltage=275, load=iout)
            line_ovp_startup_at_high_input(df, end_voltage=280, load=iout)
            line_ovp_startup_at_high_input(df, end_voltage=281, load=iout)
            line_ovp_startup_at_high_input(df, end_voltage=282, load=iout)
            line_ovp_startup_at_high_input(df, end_voltage=283, load=iout)
            line_ovp_startup_at_high_input(df, end_voltage=284, load=iout)
            line_ovp_startup_at_high_input(df, end_voltage=285, load=iout)
            line_ovp_startup_at_high_input(df, end_voltage=286, load=iout)
            line_ovp_startup_at_high_input(df, end_voltage=290, load=iout)


    if load_type == 'cr':
        for cr in cr_list:
            line_ovp_startup_at_high_input(df, end_voltage=270, load=cr)
            line_ovp_startup_at_high_input(df, end_voltage=275, load=cr)
            line_ovp_startup_at_high_input(df, end_voltage=280, load=cr)
            line_ovp_startup_at_high_input(df, end_voltage=281, load=cr)
            line_ovp_startup_at_high_input(df, end_voltage=282, load=cr)
            line_ovp_startup_at_high_input(df, end_voltage=283, load=cr)
            line_ovp_startup_at_high_input(df, end_voltage=284, load=cr)
            line_ovp_startup_at_high_input(df, end_voltage=285, load=cr)
            line_ovp_startup_at_high_input(df, end_voltage=286, load=cr)

            line_ovp_startup_at_high_input(df, end_voltage=290, load=cr)


def main():
    global waveform_counter
    discharge_output(times=1)
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)