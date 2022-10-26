"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from tracemalloc import start
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.default_scope_settings import *
import shutil
import numpy as np
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	
import getpass
username = getpass.getuser().lower()
####################################################################################################################################################################

test = "PRIMARY SWITCHING (STEADY-STATE)"
operation = "CV"
condition = input(">> Type test condition: ")

excel_name = f"{condition}"
waveforms_folder = f'C:/Users/{username}/Desktop/APPS EVAL/Single Stage GAN/Test Data ({operation})/{test}/{condition}/'
path = path_maker(f'{waveforms_folder}')

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.129"
sig_gen_address = '11'
dimming_power_meter_address = 21
eload_channel = 5


iout_channel = 1
vin_channel = 2
ids_channel = 3
vout_channel = 4


"""USER INPUT"""
vin_list = [90,100,110,115,120,132,180,200,230,265,277,300]
vout = 50
iout_nom = 1.5
iout_list = [0.10*iout_nom, 0.25*iout_nom,0.5*iout_nom,0.75*iout_nom,iout_nom]
iout_list = [iout_nom]
zoom_enable = True


"""SCOPE SETTINGS"""
trig_channel = 3
trig_level = 0.4
trig_edge = 'POS'

time_position = 50
time_scale = 0.001

"""
MEASUREMENT SETTINGS OPTIONS: "MAX,MIN,RMS,MEAN,PDELta"
"""

ch1_enable = 'ON'
ch2_enable = 'OFF'
ch3_enable = 'ON'
ch4_enable = 'ON'

"""CHANNEL 1"""
ch1_scale = 0.2
ch1_position = -4
ch1_bw = 20
ch1_rel_x_position = 20
ch1_label = "IOUT"
ch1_measure = "MAX,MIN,MEAN"
ch1_color = "GREEN"
ch1_coupling = "DCLimit"

"""CHANNEL 2"""
ch2_scale = 100
ch2_position = -4
ch2_bw = 500
ch2_rel_x_position = 40
ch2_label = "VDS"
ch2_measure = "MAX,MIN"
ch2_color = "YELLOW"
ch2_coupling = "DCLimit"

"""CHANNEL 3"""
ch3_scale = 1
ch3_position = -4
ch3_bw = 500
ch3_rel_x_position = 60
ch3_label = "PRI_IDS"
ch3_measure = "MAX"
ch3_color = "LIGHT_BLUE"
ch3_coupling = "DCLimit"

"""CHANNEL 4"""
ch4_scale = 10
ch4_position = -4
ch4_bw = 20
ch4_rel_x_position = 80
ch4_label = "VOUT"
ch4_measure = "MAX,MIN,MEAN"
ch4_color = "PINK"
ch4_coupling = "DCLimit"


"""DO NOT EDIT BELOW THIS LINE"""
####################################################################################################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
# led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output(times):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(1)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    
    scope.channel_settings(state=ch1_enable, channel=1, scale=ch1_scale, position=ch1_position, label=ch1_label,
                            color=ch1_color, rel_x_position=ch1_rel_x_position, bandwidth=ch1_bw, coupling=ch1_coupling, offset=0)
        
    scope.channel_settings(state=ch2_enable, channel=2, scale=ch2_scale, position=ch2_position, label=ch2_label,
                            color=ch2_color, rel_x_position=ch2_rel_x_position, bandwidth=ch2_bw, coupling=ch2_coupling, offset=0)
    
    scope.channel_settings(state=ch3_enable, channel=3, scale=ch3_scale, position=ch3_position, label=ch3_label,
                            color=ch3_color, rel_x_position=ch3_rel_x_position, bandwidth=ch3_bw, coupling=ch3_coupling, offset=0)
    
    scope.channel_settings(state=ch4_enable, channel=4, scale=ch4_scale, position=ch4_position, label=ch4_label,
                            color=ch4_color, rel_x_position=ch4_rel_x_position, bandwidth=ch4_bw, coupling=ch4_coupling, offset=0)
    
    if ch1_enable != 'OFF': scope.measure(1, ch1_measure)
    if ch2_enable != 'OFF': scope.measure(2, ch2_measure)
    if ch3_enable != 'OFF': scope.measure(3, ch3_measure)
    if ch4_enable != 'OFF': scope.measure(4, ch4_measure)

    scope.record_length(50E6)
    scope.time_position(time_position)
    scope.time_scale(time_scale)

    # scope.remove_zoom()
    if zoom_enable == True:
        scope.add_zoom(rel_pos=50, rel_scale=0.5)
    
    trigger_channel = trig_channel
    trigger_level = trig_level
    trigger_edge = trig_edge
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1

def browning(start, end, slew, frequency):

    if start > end:
        print(f"brownout: {start} -> {end} Vac")
        for voltage in np.arange(start, end+1, -slew):
            ac.voltage = voltage
            ac.frequency = frequency
            ac.turn_on()
            sleep(1)
    if start < end:
        print(f"brownin: {start} -> {end} Vac")
        for voltage in np.arange(start, end+1, slew):
            ac.voltage = voltage
            ac.frequency = frequency
            ac.turn_on()
            sleep(1)

def fixvoltage(voltage, soak):
    ac.voltage = voltage
    ac.frequency = ac.set_freq(voltage)
    ac.turn_on()
    sleep(soak)

def roundup(x):
    return int(math.ceil(x / 100.0)) * 100

def brown_in(led):

    test_time = ((end_voltage-start_voltage)/slew_rate)+time_fixvoltage
    scope_time = roundup(test_time)
    delay = scope_time-test_time
    time_scale = scope_time/10
    print(f"Estimated time: {scope_time/60} mins.")
    scope.time_scale(time_scale)
    scope.run()
    soak(int(delay))

    # start of test
    browning(start_voltage, end_voltage, slew_rate, frequency)
    fixvoltage(end_voltage,time_fixvoltage)
    scope.stop()

    discharge_output()

    filename = f"{led}V, {start_voltage}-{end_voltage} Vac, {slew_rate}V per s.png"
    screenshot(filename, path)

def brown_out(led):

    test_time = ((end_voltage-start_voltage)/slew_rate)+time_fixvoltage
    scope_time = roundup(test_time)
    delay = scope_time-test_time
    time_scale = scope_time/10
    print(f"Estimated time: {scope_time/60} mins.")
    scope.time_scale(time_scale)
    scope.run()
    
    # start of test
    fixvoltage(end_voltage, time_fixvoltage)
    browning(end_voltage, start_voltage, slew_rate, frequency)
    soak(int(delay))
    scope.stop()
    discharge_output()

    filename = f"{led}V, {end_voltage}-{start_voltage} Vac, {slew_rate}V per s.png"
    screenshot(filename, path)

def brown_in_brown_out(led):

    test_time = 2*((end_voltage-start_voltage)/slew_rate)
    scope_time = roundup(test_time)
    delay = scope_time-test_time
    time_scale = scope_time/10
    print(f"Estimated time: {scope_time/60} mins.")
    scope.time_scale(time_scale)
    scope.run()
    soak(int(delay/2))
    # start of test
    browning(start_voltage, end_voltage, slew_rate, frequency)
    browning(end_voltage, start_voltage, slew_rate, frequency)
    soak(int(delay/2))
    scope.stop()
    discharge_output()

    filename = f"{led}V, {start_voltage}-{end_voltage}-{start_voltage} Vac, {slew_rate}V per s.png"
    screenshot(filename, path)

def brown_out_brown_in(led):

    test_time = 2*((end_voltage-start_voltage)/slew_rate)
    scope_time = roundup(test_time)
    delay = scope_time-test_time
    time_scale = scope_time/10
    print(f"Estimated time: {scope_time/60} mins.")
    scope.time_scale(time_scale)
    scope.run()
    t_fixvoltage = int(delay/2)
    # start of test
    fixvoltage(end_voltage, t_fixvoltage)
    browning(end_voltage, start_voltage, slew_rate, frequency)
    browning(start_voltage, end_voltage, slew_rate, frequency)
    fixvoltage(end_voltage, t_fixvoltage)
    scope.stop()
    discharge_output()

    filename = f"{led}V, {end_voltage}-{start_voltage}-{end_voltage} Vac, {slew_rate}V per s.png"
    screenshot(filename, path)

def find_trigger(channel, trigger_delta):

    scope.edge_trigger(channel, 0, 'POS')

    # finding trigger level
    scope.run_single()
    scope.trigger_mode('AUTO')
    soak(5)
    scope.trigger_mode('NORM')
    

    # get initial peak-to-peak measurement value
    labels, values = scope.get_measure(channel)
    max_value = float(values[0])
    max_value = float(f"{max_value:.4f}")

    # set max_value as initial trigger level
    trigger_level = max_value
    scope.edge_trigger(channel, trigger_level, 'POS')

    # check if it triggered within 3 seconds
    scope.run_single()
    soak(3)
    trigger_status = scope.trigger_status()

    # increase trigger level until it reaches the maximum trigger level
    while (trigger_status == 1):
        trigger_level += trigger_delta
        scope.edge_trigger(channel, trigger_level, 'POS')

        # check trigger status
        scope.run_single()
        soak(3)
        trigger_status = scope.trigger_status()

    # decrease trigger level below to get the maximum trigger possible
    trigger_level -= 1*trigger_delta
    scope.edge_trigger(channel, trigger_level, 'POS')

    scope.run_single()
    soak(3)
    trigger_status = scope.trigger_status()
    while (trigger_status != 1):
        # decrease trigger level below to get the maximum trigger possible
        trigger_level -= 1*trigger_delta
        scope.edge_trigger(channel, trigger_level, 'POS')

        # check trigger status
        scope.run_single()
        soak(3)
        trigger_status = scope.trigger_status()




#################### DATA EXPORT CODE ###########################################################
def create_header_list(df_header_list):
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list
    print(df)
    return df

def export_to_excel(df, output_list, excel_name, sheet_name, anchor):
    df.loc[len(df)] = output_list
    print(output_list)

    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    df_to_excel(wb, sheet_name, df, anchor)
    wb.save(dst)

def export_screenshot_to_excel(excel_name, waveforms_folder, sheet_name, filename, anchor):

    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{excel_name}.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    image_to_excel(wb, sheet_name, filename=filename, folder_path=waveforms_folder, anchor=anchor)
    wb.save(dst)

def collect_data(input_voltage, load):

    ################ SUBJECT TO CHANGE ################
    input_step = input_voltage

    labels, values = scope.get_measure(vout_channel)
    vout_max = round(float(values[0]),2)
    vout_min = round(float(values[1]),2)
    vout_mean = round(float(values[2]),2)

    labels, values = scope.get_measure(iout_channel)
    iout_max = round(float(values[0]),2)
    iout_min = round(float(values[1]),2)
    iout_mean = round(float(values[2]),2)

    labels, values = scope.get_measure(ids_channel)
    ids_max = round(float(values[0]),2)

    vout_overshoot = round(abs(100*(vout_max-vout)/vout),2)
    vout_undershoot = round(abs(100*(vout-vout_min)/vout),2)

    if vout_overshoot > 8: vout_overshoot_result = 'FAIL'
    else: vout_overshoot_result = 'PASS'
    if vout_undershoot > 8: vout_undershoot_result = 'FAIL'
    else: vout_undershoot_result = 'PASS'

    load_percent = int(load*100/iout_nom)
    load_value = round(load, 2)


    ton = round(float(scope.get_cursor(1)["delta x"])/1E-6, 2)
    fsw = round((1/float(scope.get_cursor(2)["delta x"]))/1000, 2)

    output_list = [input_step, load_percent, load_value,
                    vout_max, vout_overshoot, vout_overshoot_result,
                    vout_min, vout_undershoot, vout_undershoot_result,
                    iout_max, ids_max, iout_mean, vout_mean, ton, fsw]
    
    return output_list


##################################################################################################


def operation():


    ################### COPY PASTE THIS CODE IN HEADER OF MAIN ###################
    df_header_list = ['Input Voltage (VAC)', 'Load (%)', 'Load',
                        'Vout_max (V)', 'Overshoot (%)', 'PASS/FAIL',
                        'Vout_min (V)', 'Undershoot (%)', 'PASS/FAIL',
                        'Iout_max (A)', 'Ids_max (A)',
                        'Iout_mean (A)', 'Vout_mean (V)', 'Ton (us)', 'Fsw (kHz)']
    df = create_header_list(df_header_list)
    ##############################################################################

    for iout in iout_list:
        
        row = 2

        for vin in vin_list:
            
            eload.channel[eload_channel].von = 0
            eload.channel[eload_channel].cc = iout
            eload.channel[eload_channel].turn_on()

            ac.voltage = vin
            ac.turn_on()
            sleep(5)

            find_trigger(channel=ids_channel, trigger_delta=0.0271)

            scope.run_single()
            sleep(3)

            ac.voltage = vin
            ac.turn_on()
            sleep(5)

            prompt("Adjust cursor.")

            percent = int(iout*100/iout_nom)
            iout = round(iout, 2)
            filename = f"Steady-state - {vin}Vac, {iout}A ({percent}percent CC load).png"
            screenshot(filename, path)

            output_list = collect_data(vin, iout)
            export_to_excel(df, output_list, excel_name, sheet_name="Ton vs Fsw", anchor="B2")
            export_screenshot_to_excel(excel_name, waveforms_folder, "Ton vs Fsw", filename, f"V{row}")
            row += 34
            
            discharge_output(times=3)
            

def main():
    global waveform_counter
    discharge_output(times=1)
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)