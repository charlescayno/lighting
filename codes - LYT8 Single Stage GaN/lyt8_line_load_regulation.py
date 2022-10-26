"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
import numpy as np
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Keithley_DC_2230G
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
date = now.strftime('%m%d')
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

import getpass
username = getpass.getuser().lower()

from powi.equipment import create_header_list, export_to_excel, export_screenshot_to_excel
##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address_1 = 20 
load_power_meter_address = 2
dimming_power_meter_address = 21
eload_address = 8
scope_address = "10.125.11.10"
dc_source_address = '4'
eload_channel = 5


##################################################################################

vin_list = [90,100,110,115,120,132,180,200,230,265,277,300]
vout = 50
iout_nom = 1.5
soak_time = 60 # s
cr_load = vout/iout_nom

test = f"{date} Parametrics"
condition = input(">> Condition: ")
excel_name = f"{condition}"
waveforms_folder = f'C:/Users/{username}/Desktop/APPS EVAL/Single Stage GAN/Test Data (CV)/{test}/'
path = path_maker(f'{waveforms_folder}')

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
pml1 = PowerMeter(load_power_meter_address_1)
eload = ElectronicLoad(eload_address)

"""DISCHARGE FUNCTION"""
def discharge_output(times=1):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(2)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)



def collect_data(vin):


    ################ SUBJECT TO CHANGE ################
    vac = vin
    freq = ac.set_freq(vin)
    vac_actual = float(f"{pms.voltage:.6f}")
    iin = float(f"{pms.current*1000:.6f}")
    pin = float(f"{pms.power:.6f}")
    pf = float(f"{pms.pf:.6f}")
    thd = float(f"{pms.thd:.6f}")
    vo1 = float(f"{pml.voltage:.6f}")
    io1 = float(f"{pml.current*1000:.6f}")
    po1 = float(f"{pml.power:.6f}")
    vreg1 = float(f"{100*(float(vo1)-vout)/float(vo1):.6f}")
    iout1 = (float(io1)/1000)
    try: ireg1 = float(f"{100*(iout1-iout_nom)/iout1:.6f}")
    except: ireg1 = 0
    eff = float(f"{100*(float(po1))/float(pin):.6f}")
    # vpfc = float(f"{pml1.voltage:.6f}")

   
    output_list = [vac, freq, vac_actual, iin, pin, pf, thd, vo1, io1, po1, vreg1, ireg1, eff]

    return output_list


def operation():


    ################### COPY PASTE THIS CODE IN HEADER OF MAIN ###################
    df_header_list = ['Vin', 'Freq (Hz)', 'Vac (VAC)', 'Iin (mA)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (mA)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)', 'Io (mA)'] 
    df = create_header_list(df_header_list)
    ##############################################################################

    eload.channel[eload_channel].cr = cr_load
    eload.channel[eload_channel].turn_on()

    ac.voltage = vin_list[0]
    ac.turn_on()

    soak(3)

    eload.channel[eload_channel].cc = iout_nom
    eload.channel[eload_channel].turn_on()


    for vin in vin_list:

        for load_percent in load_percent_list:
            eload.channel[eload_channel].cc = iout_nom * (load_percent / 100)
            eload.channel[eload_channel].turn_on()

            ac.voltage = vin
            ac.turn_on()
        
            soak(soak_time)

            output_list = collect_data(vin)
            export_to_excel(df, waveforms_folder, output_list, excel_name, sheet_name=excel_name, anchor="B2")

    print(f"\n\nFinal Data: ")
    print(df)

    discharge_output(2)

    
def main():
    global waveform_counter
    discharge_output(1)
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)
