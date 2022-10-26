##################################################################################
"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
import numpy as np
import shutil
import os
import pandas as pd

# from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Keithley_DC_2230G
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.equipment import create_header_list, export_to_excel, export_screenshot_to_excel, export_df_to_excel
from powi.equipment import path_maker, remove_file
from powi.equipment import create_scatter_chart, append_series, reset_chartsheet, save_chartsheet, create_bar_chart
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import getpass
username = getpass.getuser().lower()

from datetime import datetime
now = datetime.now()
date = now.strftime('%m%d')
##################################################################################

unit = sys.argv[1]

waveforms_folder = f"C:/Users/ccayno/Desktop/APPS EVAL/Single Stage GAN/Test Data (CV)/1014/No Load/"
path = path_maker(f'{waveforms_folder}')
excel_name = f"No Load_1007 rework"
full_path = path + excel_name + '.xlsx'
print(full_path)

def anchor_picker(ws_name, chart, data_length=101, led_row=[3,106,209], x_axis_col=['C', 'T', 'AK', 'BB'], y_axis_col=['M', 'AD', 'AU', 'BL']):
    
    """
        led_row: sequential reference starting from 48 V to 24 V
        x_axis_col: data anchors for x-axis chart; sequential reference starting from 90 VAC to 277 VAC
        y_axis_col: data anchors for y-axis chart; sequential reference starting from 90 VAC to 277 VAC    
    """

    if ws_name == 'Parametrics':
        led_list = [48, 36, 24]
        vin_list = [90, 115, 230, 277]
        for led in led_list:

            # row section per led
            if led == 48: row = led_row[0]
            if led == 36: row = led_row[1]
            if led == 24: row = led_row[2]

            # vin_list = [90,100,110,115,120,132,180,200,230,265,277]

            x_col = x_axis_col[0]
            y_col = y_axis_col[0]
            append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V", chart=chart)
    
    elif ws_name == excel_name:
        row = led_row[0]
        x_col = x_axis_col[0]
        y_col = y_axis_col[0]
        append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"No Load", chart=chart)

    elif ws_name == 'DALI':
        x_col = x_axis_col[0]
        y_col = y_axis_col[0]
        row = led_row[0]
        led = 48
        append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V", chart=chart)

    elif ws_name == 'Analog Dimming' or ws_name == 'Analog Deep Dimming':
        led_list = [48,24]
        vin_list = [115,230]

        for led in led_list:
            for i in range(len(led_list)):
                if led == led_list[i]: row = led_row[i]

            for vin in vin_list:

                for i in range(len(vin_list)):

                    if vin == vin_list[i]:
                        x_col = x_axis_col[i]
                        y_col = y_axis_col[i]

                append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V, {vin} VAC", chart=chart)
   
    elif ws_name == 'Resistor Dimming':
        led_list = [48,24]
        vin_list = [90,115,230]

        for led in led_list:
            for i in range(len(led_list)):
                if led == led_list[i]: row = led_row[i]

            for vin in vin_list:

                for i in range(len(vin_list)):

                    if vin == vin_list[i]:
                        x_col = x_axis_col[i]
                        y_col = y_axis_col[i]

                append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V, {vin} VAC", chart=chart)

    elif ws_name == 'PWM_300Hz' or ws_name == 'PWM_3000Hz':
        led_list = [48,24]
        vin_list = [115,230]

        for led in led_list:
            for i in range(len(led_list)):
                if led == led_list[i]: row = led_row[i]

            for vin in vin_list:

                for i in range(len(vin_list)):

                    if vin == vin_list[i]:
                        x_col = x_axis_col[i]
                        y_col = y_axis_col[i]

                append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V, {vin} VAC", chart=chart)

    else:
        pass


def no_load_chart(ws_name, chart_row_location = 2):
    
    #########################################
    ### user variables ######################
    title = f"No Load"
    x_title = 'Input Voltage (VAC)'
    led_row = [3]
    x_axis_col = ['B']
    y_axis_col_pin = ['F']

    x_min_scale = 90
    x_max_scale = 300
    x_major_unit = 15
    x_minor_unit = 15
    
    style = 2
    ### user variables ######################
    #########################################
    
    data_length = 12

    """ OUTPUT CURRENT REGULATION """
    chart = create_scatter_chart(title=title, style=style, x_title=x_title, y_title='No Load Input Power (W)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 0.25, y_max_scale = 0.5, y_major_unit = 0.05, y_minor_unit = 0.05)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pin)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"B{chart_row_location}")


def operation():

    global wb
    global chart_sheet

    wb = load_workbook(full_path)
    # wb.create_sheet("Chart")
    chart_sheet = reset_chartsheet(wb)
    no_load_chart(ws_name=excel_name, chart_row_location=2)
    wb.save(full_path)


def main():

    operation()
        
if __name__ == "__main__":
    headers("Chart")
    main()
    footers(waveform_counter)


