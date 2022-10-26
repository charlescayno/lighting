##################################################################################
"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
import numpy as np
import shutil
import os
import pandas as pd

# from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Keithley_DC_2230G
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.equipment import create_header_list, export_to_excel, export_screenshot_to_excel, export_df_to_excel
from powi.equipment import path_maker, remove_file
from powi.equipment import create_scatter_chart, append_series, reset_chartsheet, save_chartsheet, create_bar_chart
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import getpass
username = getpass.getuser().lower()

from datetime import datetime
now = datetime.now()
date = now.strftime('%m%d')
##################################################################################

waveforms_folder = f"C:/Users/ccayno/Desktop/APPS EVAL/Single Stage GAN/Test Data (CV)/1006 - Final Parametrics/"
path = path_maker(f'{waveforms_folder}')
excel_name = f"No Load"
full_path = path + excel_name + '.xlsx'
print(full_path)

def anchor_picker(ws_name, chart, data_length=101, led_row=[3,106,209], x_axis_col=['C', 'T', 'AK', 'BB'], y_axis_col=['M', 'AD', 'AU', 'BL']):

    row = led_row[0]
    x_col = x_axis_col[0]
    y_col = y_axis_col[0]
    append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=ws_name, chart=chart)


def create_chart_from_test_name_sheet(test_name, title, x_title, data_row, x_axis_col, y_axis_col,
                                            x_min_scale, x_max_scale, x_major_unit, x_minor_unit,
                                            y_min_scale, y_max_scale, y_major_unit, y_minor_unit,
                                            data_length, style, chart_col_location, chart_row_location):
    
    chart = create_scatter_chart(title=title, style=style, x_title=x_title, y_title=title,
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = y_min_scale, y_max_scale = y_max_scale, y_major_unit = y_major_unit, y_minor_unit = y_minor_unit)


    anchor_picker(test_name, chart, data_length=data_length, led_row=data_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col)
    save_chartsheet(chart_sheet, chart, chart_position=f"{chart_col_location}{chart_row_location}")

def operation():

    global wb
    global chart_sheet

    wb = load_workbook(full_path)
    sheet_list = wb.get_sheet_names()
    if "Chart" not in sheet_list: wb.create_sheet("Chart")
    chart_sheet = reset_chartsheet(wb)

    create_chart_from_test_name_sheet(test_name="No Load", title="Input Power (W)", x_title='Input Voltage (VAC)', data_row=[3], x_axis_col=['B'], y_axis_col=['F'],
                                            x_min_scale=90, x_max_scale=300, x_major_unit=15, x_minor_unit=15,
                                            y_min_scale=0.25, y_max_scale=0.5, y_major_unit=0.05, y_minor_unit=0.05,
                                            data_length=500, style=2, chart_col_location='B', chart_row_location=2)

    wb.save(full_path)


def main():

    operation()
        
if __name__ == "__main__":
    headers("Chart")
    main()
    footers(waveform_counter)


