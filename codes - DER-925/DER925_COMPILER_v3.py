try:
    import pandas as pd
    import os
    from openpyxl import Workbook, load_workbook
    import sys
    import openpyxl
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    from openpyxl.utils import get_column_letter

except:
    import pip
    pip.main(['install','pillow'])
    pip.main(['install','pandas'])
    pip.main(['install','openpyxl'])
    
    import pandas as pd
    import os
    from openpyxl import Workbook, load_workbook
    import openpyxl
    import sys
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    from openpyxl.utils import get_column_letter


def col_row_extractor(excel_coordinate):
    """
    extract col and row given an excel coordinate (i.e. 'B4' -> col = 2, row = 4)

    excel_coordinate : i.e. 'B4' (str)
    returns col, row (int)
    """
    coordinates = coordinate_from_string(excel_coordinate)
    col = column_index_from_string(coordinates[0])
    row = coordinates[1]
    return col, row

def excel_to_df(filename, sheet_name, start_corner, end_corner):
    """
    reading dataframe from excel.
    
    filename     : must include full filename path (cwd + path + file.extension)
    sheet_name   : sheet name in excel file
    start_corner : cell coordinate to start selection of data
    end_corner   : cell coordinate to end selection of data

    returns df
    """

    # print(f"reading dataframe from {filename} {sheet_name}")

    start_col, start_row = col_row_extractor(start_corner)
    end_col, end_row = col_row_extractor(end_corner)

    skiprows = start_row - 2
    usecols = f'{get_column_letter(start_col)}:{get_column_letter(end_col)}'
    nrows = end_row - start_row + 1

    return pd.read_excel(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)

def df_to_excel(wb, sheet_name, df, anchor):
    """
    writing dataframe to excel.

    wb          : workbook
    sheet_name  : sheet name in excel file
    df          : dataframe
    anchor      : anchor point in excel

    returns None
    """

    # print(f"writing in {wb} {sheet_name}")

    start_col, start_row = col_row_extractor(anchor)
    df_row_len, df_col_len = df.shape
    end_row = start_row + df_row_len - 1
    end_col = start_col + df_col_len - 1


    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            wb[sheet_name][f'{get_column_letter(col)}{row}'] = df.iloc[row-start_row, col-start_col]


def image_to_excel(wb, sheet_name, filename, folder_path, anchor):
    """
    writing image to excel.

    image size -> 39R, 16C
    wb          : workbook
    sheet_name  : sheet name in excel file
    filename    : filename of theh image
    folder_path : image location
    anchor      : anchor point in excel
    """

    file = os.getcwd() + folder_path + filename
    img = openpyxl.drawing.image.Image(file)
    img.anchor = anchor
    img.width = 1026
    img.height = 762

    wb[sheet_name].add_image(img)




## ENTER UNIT NUMBER ####
unit = sys.argv[1]
#########################

# replace existing file to the template
from shutil import copyfile
src = f"{os.getcwd()}/Unit X Data Pack V2.xlsx"
dst = f"{os.getcwd()}/Unit {unit} Data Pack.xlsx"
if os.path.exists(dst): os.remove(dst)
copyfile(src, dst)

# destination file
dest = f'Unit {unit} Data Pack.xlsx'

"""DATAFRAMES"""

"""(PF, THD, EFF, NL, Harmonics)"""

file = f'Unit {unit} (PF, THD, EFF, NL, Harmonics).xlsm'
path = f"/Electricals/"
filename = os.getcwd() + path + file

df_linereg_48V = excel_to_df(filename, 'LineReg 48V', 'E6', 'P17')
df_linereg_36V = excel_to_df(filename, 'LineReg 36V', 'E6', 'P17')
df_linereg_24V = excel_to_df(filename, 'LineReg 24V', 'E6', 'P17')

df_nl = excel_to_df(filename, 'NL', 'E6', 'P17')
df_nl_lightoff = excel_to_df(filename, 'NL lightoff', 'E6', 'P17')


df_harmonics_120Vac_24V = excel_to_df(filename, '120Vac 24V', 'E11', 'P37')
df_harmonics_230Vac_24V = excel_to_df(filename, '230Vac 24V', 'E11', 'P37')
df_harmonics_277Vac_24V = excel_to_df(filename, '277Vac 24V', 'E11', 'P37')

df_harmonics_120Vac_48V = excel_to_df(filename, '120Vac 48V', 'E11', 'P37')
df_harmonics_230Vac_48V = excel_to_df(filename, '230Vac 48V', 'E11', 'P37')
df_harmonics_277Vac_48V = excel_to_df(filename, '277Vac 48V', 'E11', 'P37')

"""ANALOG DIMMING (0.1V)"""

file = f"Unit {unit} (0-10V 0.1 V Step Analog Dimming).xlsm"
path = "/Dimming/Analog/"
filename = os.getcwd() + path + file

df_analog_120Vac_48V_0_1V = excel_to_df(filename, '48V 120Vac', 'D6', 'P106')
df_analog_277Vac_48V_0_1V = excel_to_df(filename, '48V 277Vac', 'D6', 'P106')

df_analog_120Vac_24V_0_1V = excel_to_df(filename, '24V 120Vac', 'D6', 'P106')
df_analog_277Vac_24V_0_1V = excel_to_df(filename, '24V 277Vac', 'D6', 'P106')

"""RESISTOR DIMMING"""

file = f"Unit {unit} (Resistor Dimming).xlsm"
path = "/Dimming/Resistor/"
filename = os.getcwd() + path + file

df_resistor_120Vac_48V = excel_to_df(filename, '48V 120Vac', 'E6', 'P16')
df_resistor_277Vac_48V = excel_to_df(filename, '48V 277Vac', 'E6', 'P16')

df_resistor_120Vac_24V = excel_to_df(filename, '24V 120Vac', 'E6', 'P16')
df_resistor_277Vac_24V = excel_to_df(filename, '24V 277Vac', 'E6', 'P16')

"""PWM DIMMING"""

file = f"Unit {unit} (2KHz, 400Hz PWM Dimming).xlsm"
path = "/Dimming/PWM/"
filename = os.getcwd() + path + file

df_pwm_2kHz_120Vac_48V = excel_to_df(filename, '48V 120Vac 2KHz', 'D6', 'P16')
df_pwm_2kHz_120Vac_24V = excel_to_df(filename, '24V 120Vac 2KHz', 'D6', 'P16')

df_pwm_400Hz_277Vac_48V = excel_to_df(filename, '48V 277Vac 400Hz', 'D6', 'P16')
df_pwm_400Hz_277Vac_24V = excel_to_df(filename, '24V 277Vac 400Hz', 'D6', 'P16')

"""DALI DIMMING"""

file = f"Unit {unit} (DALI Dimming).xlsm"
path = "/Dimming/DALI/"
filename = os.getcwd() + path + file

df_dali_277Vac_48V = excel_to_df(filename, '48V 277Vac', 'D6', 'P91')
df_dali_277Vac_24V = excel_to_df(filename, '24V 277Vac', 'D6', 'P91')


"""I2C DIMMING (Linear)"""

file = f"Unit {unit} (I2C Dimming Linear).xlsm"
path = "/Dimming/I2C Linear/"
filename = os.getcwd() + path + file

df_i2c_linear_230Vac_48V = excel_to_df(filename, '48V 230Vac', 'D6', 'P91')
df_i2c_linear_230Vac_24V = excel_to_df(filename, '24V 230Vac', 'D6', 'P91')


"""I2C DIMMING (Logarithmic)"""

file = f"Unit {unit} (I2C Dimming Logarithmic).xlsm"
path = "/Dimming/I2C Logarithmic/"
filename = os.getcwd() + path + file

df_i2c_logarithmic_120Vac_48V = excel_to_df(filename, '48V 120Vac', 'D6', 'P91')
df_i2c_logarithmic_120Vac_24V = excel_to_df(filename, '24V 120Vac', 'D6', 'P91')


"""RSET DIMMING"""

file = f"Unit {unit} (RSET Dimming).xlsm"
path = "/Dimming/RSET/"
filename = os.getcwd() + path + file

df_rset_120Vac_48V = excel_to_df(filename, '48V 120Vac', 'E6', 'P15')
df_rset_277Vac_48V = excel_to_df(filename, '48V 277Vac', 'E6', 'P15')

df_rset_120Vac_24V = excel_to_df(filename, '24V 120Vac', 'E6', 'P15')
df_rset_277Vac_24V = excel_to_df(filename, '24V 277Vac', 'E6', 'P15')


# input()













"""WRITING DATA TO EXCEL"""




wb = load_workbook(dest)

sheet_name = 'EMI'
image_to_excel(wb, sheet_name, f"Unit {unit} 115VAC_scan.png", "/EMI/Scan/", 'B8')
image_to_excel(wb, sheet_name, f"Unit {unit} 115VAC_peaklist.png", "/EMI/Peaklist/", 'T8')
image_to_excel(wb, sheet_name, f"Unit {unit} 230VAC_scan.png", "/EMI/Scan/", 'B55')
image_to_excel(wb, sheet_name, f"Unit {unit} 230VAC_peaklist.png", "/EMI/Peaklist/", 'T55')

sheet_name = 'PF, THD, EFF'
wb[sheet_name]["A3"] = f"DER-925 Unit {unit}"
df_to_excel(wb, sheet_name, df_linereg_48V, 'B12')
df_to_excel(wb, sheet_name, df_linereg_36V, 'B28')
df_to_excel(wb, sheet_name, df_linereg_24V, 'B44')


sheet_name = 'No Load Input Power'
df_to_excel(wb, sheet_name, df_nl, 'B10')
df_to_excel(wb, sheet_name, df_nl_lightoff, 'B28')

sheet_name = 'Harmonics'
df_to_excel(wb, sheet_name, df_harmonics_120Vac_48V, 'B10')
df_to_excel(wb, sheet_name, df_harmonics_230Vac_48V, 'B41')
df_to_excel(wb, sheet_name, df_harmonics_277Vac_48V, 'B72')

df_to_excel(wb, sheet_name, df_harmonics_120Vac_24V, 'AZ10')
df_to_excel(wb, sheet_name, df_harmonics_230Vac_24V, 'AZ41')
df_to_excel(wb, sheet_name, df_harmonics_277Vac_24V, 'AZ72')


sheet_name = 'Analog Dimming (0.1V step)'
df_to_excel(wb, sheet_name, df_analog_120Vac_48V_0_1V, 'B12')
df_to_excel(wb, sheet_name, df_analog_277Vac_48V_0_1V, 'AF12')

df_to_excel(wb, sheet_name, df_analog_120Vac_24V_0_1V, 'CP12')
df_to_excel(wb, sheet_name, df_analog_277Vac_24V_0_1V, 'DT12')

sheet_name = 'Resistor Dimming'
df_to_excel(wb, sheet_name, df_resistor_120Vac_48V, 'C12')
df_to_excel(wb, sheet_name, df_resistor_277Vac_48V, 'AG12')

df_to_excel(wb, sheet_name, df_resistor_120Vac_24V, 'CQ12')
df_to_excel(wb, sheet_name, df_resistor_277Vac_24V, 'DU12')

sheet_name = 'PWM Dimming 400 Hz'
df_to_excel(wb, sheet_name, df_pwm_400Hz_277Vac_48V, 'AF12')
df_to_excel(wb, sheet_name, df_pwm_400Hz_277Vac_24V, 'DT12')

sheet_name = 'PWM Dimming 2000 Hz'
df_to_excel(wb, sheet_name, df_pwm_2kHz_120Vac_48V, 'B12')
df_to_excel(wb, sheet_name, df_pwm_2kHz_120Vac_24V, 'CP12')


sheet_name = 'DALI I2C Dimming (Logarithmic)'
df_to_excel(wb, sheet_name, df_dali_277Vac_48V, 'AE10')
df_to_excel(wb, sheet_name, df_dali_277Vac_24V, 'BY10')

sheet_name = 'I2C Dimming (Linear)'
df_to_excel(wb, sheet_name, df_i2c_linear_230Vac_48V, 'P10')
df_to_excel(wb, sheet_name, df_i2c_linear_230Vac_24V, 'BJ10')


sheet_name = 'I2C Dimming (Logarithmic)'
df_to_excel(wb, sheet_name, df_i2c_logarithmic_120Vac_48V, 'A10')
df_to_excel(wb, sheet_name, df_i2c_logarithmic_120Vac_24V, 'AU10')


sheet_name = 'RSET Dimming'
df_to_excel(wb, sheet_name, df_rset_120Vac_48V, 'B10')
df_to_excel(wb, sheet_name, df_rset_277Vac_48V, 'AF10')
df_to_excel(wb, sheet_name, df_rset_120Vac_24V, 'AV10')
df_to_excel(wb, sheet_name, df_rset_277Vac_24V, 'BZ10')


wb.save(dest)
print(f"Unit {unit} Data Pack is complete.")

