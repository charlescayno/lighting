from functools import wraps
from math import isnan
from time import sleep, time
from abc import ABC, abstractmethod
import atexit
import os
import sys

try:
    from pyfirmata import Arduino, util
    import pyfirmata
    import pyvisa
    import numpy as np
    from gtts import gTTS
    import sys
    from playsound import playsound

except:
    import pip
    pip.main(['install','pyqt5'])
    pip.main(['install','pyinstaller'])
    pip.main(['install','pyautogui'])
    pip.main(['install','pyfirmata'])
    pip.main(['install', 'pyvisa'])
    pip.main(['install','pandas'])
    pip.main(['install','opencv-python'])
    pip.main(['install','matplotlib'])
    pip.main(['install','numpy'])
    pip.main(['install', 'gTTS'])
    pip.main(['install', 'playsound'])

    from pyfirmata import Arduino, util
    import pyfirmata
    import pyvisa
    import numpy as np



GPIB_NO = 0
PORT = 0

type_to_address = {
    'gpib': f'GPIB{GPIB_NO}',
    'lan': 'TCPIP',
    'usb': 'USB',
    'serial': 'ASRL'
    }

def reject_nan(func):
    @wraps(func)
    def wrapped(*args, **kwargs):
        for _ in range(10):
            response = func(*args, **kwargs)
            if not isnan(response):
                return response
            sleep(0.2)
        return None
    return wrapped

class Equipment(ABC):
    def __init__(self, address, comm_type='gpib', timeout=10000):
        rm = pyvisa.ResourceManager()
        # print(rm.list_resources()) # print out devices connected on the VISA COM
        if comm_type == 'serial':
            self.address = f'{type_to_address[comm_type]}{address}'
            print(self.address)
        else:
            self.address = f'{type_to_address[comm_type]}::{address}'
        self.timeout = timeout
        try:
            self.device = rm.open_resource(self.address)
            self.device.timeout = self.timeout
            atexit.register(self.cleanup)
        except Exception as e:
            raise pyvisa.VisaIOError(e)
            
        self._id = 0

    def __repr__(self):
        return (f'{self.__class__.__name__}'
                 '(address={self.address}, timeout={self.timeout})')

    @abstractmethod
    def cleanup(self):
        pass

    @property
    def id(self):
        self._id = self.write('*IDN?')
        return self._id

    def close(self):
        self.device.close()

    # def write(self, command):
    #     response = None
    #     try:
    #         if "?" in command:
    #             response = self.device.query(command).strip()
    #         else:
    #             self.device.write(command)
    #     except Exception as e:
    #         raise pyvisa.VisaIOError(e)

    #     return response
    
    def write(self, command):
        response = None
        reply_available = False # CDO
        try:
            if "BDMM" in command: # CDO
                reply_available = True
                command = command.split(':')[1]

            if "?" in command:
                response = self.device.query(command).strip()
            else:
                self.device.write(command)

            if reply_available: # CDO
                self.device.read()
                reply_available = False

        except Exception as e:
            raise pyvisa.VisaIOError(e)

        return response

class ACSource(Equipment):
    def __init__(self, address, comm_type='gpib', timeout=10000):
        super().__init__(address, comm_type, timeout)

        self._voltage = 0
        self._frequency = 0
        self._coupling = 'AC'

    def turn_on(self, phase=None):
        if self._coupling == 'AC':
            self.write(f'VOLT {self._voltage}')
            self.write('OUTP:COUP AC')
            self.write('VOLT:OFFS 0')

        elif self._coupling == 'DC':
            self.write('VOLT 0')
            self.write('OUTP:COUP DC')
            self.write(f'VOLT:OFFS {self._voltage}')
        
        self.write('OUTP ON')

    def turn_off(self):
        self.write('OUTP OFF')

    def set_freq(self, voltage):
        if voltage >= 180 and voltage <= 265: ac_freq = 50
        else: ac_freq = 60
        return ac_freq

    @property
    def voltage(self):
        return self._voltage

    @property
    def frequency(self):
        return self._frequency
    
    @property
    def coupling(self):
        return self._coupling

    @voltage.setter
    def voltage(self, voltage):
        self._voltage = voltage
        """automatic frequency setting"""
        self._frequency = self.set_freq(self._voltage)
        self.write(f'FREQ {self._frequency}')

    @frequency.setter
    def frequency(self, frequency):
        self._frequency = frequency # manual setting
        self.write(f'FREQ {self._frequency}')

    @coupling.setter
    def coupling(self, type):
        if type.upper() == 'DC':
            self._coupling = 'DC'
        else:
            self._coupling = 'AC'

    def ac_cycling(self, pulse_count, vin, start_soak, off_time, on_time, end_soak):
    
        freq = self.set_freq(vin)
        self.write(f"TRIG:TRAN:SOUR BUS")
        
        a = f"LIST:DWELL {start_soak}, "
        for i in range(pulse_count):
            a = a + f"{off_time}, {on_time}, "
        a = a + f"{off_time}, {end_soak}"
        # print(a)
        self.write(f"{a}")

        self.write(f"VOLT:MODE LIST")
        
        b = f"LIST:VOLT {vin}, "
        for i in range(pulse_count):
            b = b + f"0, {vin}, "
        b = b + f"0, {vin}"
        # print(b)
        self.write(b)

        self.write(f"VOLT:SLEW:MODE LIST")

        c = f"LIST:VOLT:SLEW 9.9e+037, "
        for i in range(pulse_count):
            c = c + f"9.9e+037, 9.9e+037, "
        c = c + f"9.9e+037, 9.9e+037"
        # print(c)
        self.write(c)

        self.write(f"FREQ:MODE LIST")
        d = f"LIST:FREQ {freq}, "
        for i in range(pulse_count):
            d = d + f"{freq}, {freq}, "
        d = d + f"{freq}, {freq}"
        # print(d)
        self.write(d)

        self.write(f"FREQ:SLEW:MODE LIST")

        e = f"LIST:FREQ:SLEW 9.9e+037, "
        for i in range(pulse_count):
            e = e + f"9.9e+037, 9.9e+037, "
        e = e + f"9.9e+037, 9.9e+037"
        # print(e)
        self.write(e)

        self.write(f"VOLT:OFFS:MODE FIX")
        self.write(f"VOLT:OFFS:SLEW:MODE FIX")
        self.write(f"PHAS:MODE LIST")

        f = f"LIST:PHAS 270, "
        for i in range(pulse_count):
            f = f + f"270, 270, "
        f = f + f"270, 270"
        # print(f)
        self.write(f)

        self.write(f"CURR:PEAK:MODE LIST")

        g = f"LIST:CURR 40.4, "
        for i in range(pulse_count):
            g = g + f"40.4, 40.4, "
        g = g + f"40.4, 40.4"
        # print(g)
        self.write(g)

        self.write(f"FUNC:MODE FIX")

        h = f"LIST:TTLT ON, "
        for i in range(pulse_count):
            h = h + f"OFF, OFF, "
        h = h + f"OFF, OFF"
        # print(h)
        self.write(h)

        self.write(f"LIST:STEP AUTO")
        self.write(f"OUTP:TTLT:STAT ON")
        self.write(f"OUTP:TTLT:SOUR LIST")
        self.write(f"TRIG:SYNC:SOUR PHASE")
        self.write(f"TRIG:SYNC:PHAS 0.0")
        self.write(f"TRIG:TRAN:DEL 0")
        self.write(f"Sens:Swe:Offs:Poin 0")
        self.write(f"TRIG:ACQ:SOUR TTLT")
        self.write(f"INIT:IMM:SEQ3")
        self.write(f"LIST:COUN 1")
        self.write(f"INIT:IMM:SEQ1")
        self.write(f"TRIG:TRAN:SOUR BUS")
        self.write(f"TRIG:IMM")

        delay = start_soak + end_soak + (pulse_count*(off_time + on_time)) + off_time
        sleep(delay)

    def cleanup(self):
        self.turn_off()
        self.close()

class DCSource(Equipment):
    def __init__(self, address, comm_type='gpib', timeout=10000):
        super().__init__(address, comm_type, timeout)

        self._voltage = 0
        self._current = 0
        self._cv = 0
        self._cc = 0

    def turn_on(self):
        self.write('CONF:OUTP ON')
        print("CONF:OUTP ON")

    def turn_off(self):
        self.write('CONF:OUTP OFF')
        print("CONF:OUTP OFF")

    @property
    def voltage(self):
        self._voltage = self.write('MEAS:VOLT?')
        return self._voltage

    @property
    def current(self):
        self._current = self.write('MEAS:CURR?')
        return self._current

    # @voltage.setter
    # def voltage(self, voltage):
    #     self._voltage = voltage

    @property
    def cv(self):
        return self._cv

    @property
    def cc(self):
        return self._cc

    @cv.setter
    def cv(self, cv):
        self._cv = cv
        self.write(f'VOLT {self._cv}')

    @cc.setter
    def cc(self, cc):
        self._cc = cc
        self.write(f'CURR {self._cc}')

    def cleanup(self):
        # self.turn_off()
        self.close()

class PowerMeter(Equipment):
    def __init__(self, address, comm_type='gpib', timeout=10000):
        super().__init__(address, comm_type, timeout)

        self._voltage = 0
        self._current = 0
        self._power = 0
        self._pf = 0
        self._thd = 0
        self._integtime = 0
        # self.reset()

    def integrate(self, integration_time=60):
        self.write('INTEGRATE:RESET')
        self.write('INTEGRATE:MODE NORMAL')
        self.write(f'INTEGRATE:TIMER 0, {integration_time // 60}, {integration_time%60} ')
        self.write('INTEGRATE:START')

    def integration_done(self):
        return self.write('INTEGRATE:STATE?') == 'TIM'

    @property
    def integtime(self, integtime):
        return self._integtime
    
    @integtime.setter
    def integtime(self, integtime):
        self._integtime = integtime

    @property
    @reject_nan
    def voltage(self):
        self.write('NUMERIC:ITEM1 U, 1')
        self._voltage = float(self.write('NUM:NORM:VAL? 1'))
        return self._voltage
        
    @property
    @reject_nan
    def current(self):
        self.write('NUMERIC:ITEM2 I, 1')
        self._current = float(self.write('NUM:NORM:VAL? 2'))
        return self._current

    @property
    @reject_nan
    def power(self):
        if self._integtime == 0:
            self.write('NUMERIC:ITEM3 P, 1')
            self._power = float(self.write('NUM:NORM:VAL? 3'))
        else:
            while not self.integration_done:
                sleep(0.5)
            print("integration successful")
            self.write('INTEGRATE:RESET')
            self.write('NUMERIC:ITEM6 WH, 1')
            self._power = float(self.write('NUM:NORM:VAL? 6')) * 3600 / self._integtime        
        return self._power

    @property
    @reject_nan
    def pf(self):
        self.write('NUMERIC:ITEM4 lambda, 1')
        self._pf = float(self.write('NUM:NORM:VAL? 4'))
        return self._pf

    @property
    @reject_nan
    def thd(self):
        self.write('NUMERIC:ITEM5 ITHD, 1')
        self._thd = float(self.write('NUM:NORM:VAL? 5'))
        return self._thd

    def get_harmonics(self):
        """
            returns: list of float harmonic content (mA)
        """
        self.write("HARMONICS:DISPLAY ON")
        self.write("NUMERIC:LIST:CLEAR ALL")
        self.write("NUMERIC:LIST:ITEM2 I,1")
        sleep(3)
        a = self.write("NUMeric:LIST:VALue? 2").split(',NAN,')[1].split(',')
        sleep(3)
        harmonic_content = []
        for i in a:
            harmonic_content.append(float(i)*1000)

        percent_content = []
        for i in range(len(harmonic_content)):
            percent_content.append(float(f"{(harmonic_content[i]*100/harmonic_content[0]):2f}"))
        print(percent_content)

        pin = float(f"{self.power:.6f}")

        return harmonic_content, percent_content

    

    def reset(self):
        self.write('*RST')

    def cleanup(self):
        self.close()

class Arduino(Equipment):

    """UNDER DEVELOPMENT"""


    def __init__(self, address, comm_type='serial', timeout=10000):
        super().__init__(address, comm_type, timeout)
        self.device.baud_rate = 115200
        self.device.flow_control = 0
        # self.device.write_termination = '\n'
        # self.device.read_termination = '\n'
        sleep(2)

    def read(self):
        print(self.device.read_raw())
        # return self.device.read_raw().rstrip().decode('utf-8')

    def cleanup(self):
        self.close()

class Oscilloscope(Equipment):
    def __init__(self, address, comm_type='lan', timeout=1E6):
        super().__init__(address, comm_type, timeout)

        self.write('SYST:DISP:UPD 1')


    def _extract_png(self, raw):
        offset = int(raw[1])-ord('0')
        return raw[offset+2:]

    def get_screenshot(self, filename="default.png", path=os.path.dirname(os.path.realpath(__file__))):
        self.write("HCOP:DEST \'MMEM\'")
        self.write("HCOP:DEV:LANG PNG")
        self.write("HCOP:DEV:INV ON")
        self.write("MMEM:NAME \'C:\\HCOPY.png\'")
        self.write("HCOP:IMMediate; *OPC?")

        self.device.write("MMEM:DATA? \'C:\\HCOPY.png\'")
        raw_image = self.device.read_raw()
        image = self._extract_png(raw_image)

        with open(path + "\\" + filename, "wb") as f:
            f.write(image)

    def probe_state(self, channel=1):
        return self.write(f'PROBe{channel}:SETup:STATe?')

    def probe_type(self, channel=1):   
        return self.write(f'PROBe{channel}:SETup:TYPE?')

    def probe_name(self, channel=1):
        return self.write(f'PROBe{channel}:SETup:NAME?')

    def probe_bandwidth(self, channel=1):
        return self.write(f'PROBe{channel}:SETup:BANDwidth?')

    def probe_attenuation(self, channel=1):
        return self.write(f'PROBe{channel}:SETup:ATTenuation[:AUTO]?')


    def auto_zero(self, channel):
        self.write(f"PROBe{channel}:SETup:OFFSet:AZERo")


    def get_measure(self, channel=1):
        channel_state = self.write(f'MEAS{channel}:ENAB?')
        if channel_state == '0':
            return None, None
        labels = []
        values = []
        self.write(f'MEAS{channel}:ARN ON')
        for item in self.write(f'MEAS{channel}:ARES?').split(','):
            label, value = item.split(':')
            labels.append(label.strip())
            values.append(float(value.strip()))        

        return labels, values

    def get_measure_all(self):
        result = []
        for i in range(1, 9):
            labels, values = self.get_measure(i)
            result.append({
                "channel": i,
                "labels": labels,
                "values": values
            })
        return result

    def get_vertical(self, channel=1):
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        result = {
            "channel": str(channel),
            "scale": float(self.write(f'CHAN{channel}:SCAL?')),
            "position": float(self.write(f'CHAN{channel}:POS?')),
            "offset": float(self.write(f'CHAN{channel}:OFFS?')),
            "coupling": self.write(f'CHAN{channel}:COUP?'),
            "bandwidth": self.write(f'CHAN{channel}:BAND?')
        }
        return result

    def get_horizontal(self):
        result = {
            "scale": float(self.write('TIM:SCAL?')),
            "position": float(self.write('TIM:HOR:POS?')),
            "resolution": float(self.write('ACQ:RES?')),
            "sample rate": float(self.write('ACQ:SRAT?'))
        }
        return result

    def get_cursor(self, cursor=1):
        cursor_state = self.write(f'CURS{cursor}:STAT?')
        if cursor_state == '0':
            return None

        result = {
            "x1 position": self.write(f'CURS{cursor}:X1P?'),
            "x2 position": self.write(f'CURS{cursor}:X2P?'),
            "y1 position": self.write(f'CURS{cursor}:Y1P?'),
            "y2 position": self.write(f'CURS{cursor}:Y2P?'),
            "delta x": self.write(f'CURS{cursor}:XDEL?'),
            "delta y": self.write(f'CURS{cursor}:YDEL?'),
            "source": self.write(f'CURS{cursor}:SOUR?')
        }
        return result

    def run(self):
        self.write('RUN')
        self.write('DISP:TRIG:LIN OFF')

    def run_single(self):
        self.trigger_mode(mode='NORM')
        # self.trigger_mode(mode='AUTO')
        self.write('RUNS')
        self.write('DISP:TRIG:LIN OFF')

    def stop(self):
        self.write('STOP')
        self.write('DISP:TRIG:LIN OFF')

    # CMC Code Update #################

    def record_length(self, record_length):
        """
            record_length : 1000 to 1 000 000 000
        """
        self.write(f'ACQ:POIN {record_length}')
        # print('Record Length: '+ str(record_length) + ' Sa') 

    def resolution(self, resolution):
        self.write(f'ACQ:RES {resolution}')
        print('Resolution: '+ str(resolution) + ' s') # 1E-15 to 0.5

    def time_position(self, time_position):
        self.write(f'TIM:REF {time_position}')
        self.write('TIM:HOR:POS 0')
    
    def time_scale(self, time_scale):
        """
        Parameters:
        <TimeScale> Range: 25E-12 to 50
                    Increment: 1E-12
                    *RST: 10E-9
                    Default unit: s/div
        """
        self.write(f'TIM:SCAL {time_scale}')

    def position_scale(self, time_position, time_scale):
        self.time_position(time_position)
        self.time_scale(time_scale)

    def remove_zoom(self):
        self.write("LAYout:ZOOM:REM 'Diagram1', 'Zoom1'")
        self.write("LAYout:ZOOM:REM 'Diagram1', 'Zoom2'")
        # self.remove_zoom_gate()

    def remove_zoom_gate(self):
        self.write(f"MEASurement1:GATE:ZCOupling OFF")

    def add_zoom_gate(self):
        self.write("MEASurement1:GATE ON")
        self.write(f"MEASurement1:GATE:ZCOupling ON")


    def add_zoom(self, rel_pos=50, rel_scale=10):
        self.remove_zoom()
        for i in range(5):
            self.write(f"LAYout:ZOOM:ADD 'Diagram{i}', VERT, OFF, -100e-6, 100e-6, -0.1, 0.05, 'Zoom1'")
            time_span = float(self.get_horizontal()['scale'])*10
            # print(f"Zoom scale: {((rel_scale/100)*time_span)/10} s/div.")
            self.write(f"LAYout:ZOOM:HORZ:REL:SPAN 'Diagram{i}', 'Zoom1', {rel_scale}")
            self.write(f"LAYout:ZOOM:HORZ:REL:POS 'Diagram{i}', 'Zoom1', {rel_pos}")
            self.write(f"LAYout:ZOOM:VERT:REL:SPAN 'Diagram{i}', 'Zoom1', 100")
            # self.write("MEASurement1:GATE ON")
            # self.write(f"MEASurement1:GATE:ZCOupling ON")
    
    def add_zoom_with_gate(self, rel_pos=50, rel_scale=10):
        self.remove_zoom()
        self.write("LAYout:ZOOM:ADD 'Diagram1', VERT, OFF, -100e-6, 100e-6, -0.1, 0.05, 'Zoom1'")
        time_span = float(self.get_horizontal()['scale'])*10
        # print(f"Zoom scale: {((rel_scale/100)*time_span)/10} s/div.")
        self.write(f"LAYout:ZOOM:HORZ:REL:SPAN 'Diagram1', 'Zoom1', {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:POS 'Diagram1', 'Zoom1', {rel_pos}")
        self.write("LAYout:ZOOM:VERT:REL:SPAN 'Diagram1', 'Zoom1', 100")
        self.add_zoom_gate()
        
    
    def add_zoom_1(self, rel_pos=50, zoom_scale=6):
        # self.remove_zoom()
        self.write("LAYout:ZOOM:ADD 'Diagram1', VERT, OFF, -100e-6, 100e-6, -0.1, 0.05, 'Zoom1'")
        time_span = float(self.get_horizontal()['scale'])*10
        print(f"Zoom scale: {zoom_scale} s/div.")
        rel_scale = (zoom_scale/time_span)*100*10
        print(f"Rel Scale: {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:SPAN 'Diagram1', 'Zoom1', {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:POS 'Diagram1', 'Zoom1', {rel_pos}")
        self.write("LAYout:ZOOM:VERT:REL:SPAN 'Diagram1', 'Zoom1', 100")
        # self.write("MEASurement1:GATE ON")
        # self.write(f"MEASurement1:GATE:ZCOupling ON")
    
    def add_zoom_2(self, rel_pos=50, zoom_scale=6):
        # self.remove_zoom()
        self.write("LAYout:ZOOM:ADD 'Diagram1', VERT, OFF, -100e-6, 100e-6, -0.1, 0.05, 'Zoom2'")
        time_span = float(self.get_horizontal()['scale'])*10
        print(f"Zoom scale: {zoom_scale} s/div.")
        rel_scale = (zoom_scale/time_span)*100*10
        print(f"Rel Scale: {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:SPAN 'Diagram1', 'Zoom2', {rel_scale}")
        self.write(f"LAYout:ZOOM:HORZ:REL:POS 'Diagram1', 'Zoom2', {rel_pos}")
        self.write("LAYout:ZOOM:VERT:REL:SPAN 'Diagram1', 'Zoom2', 100")
        # self.write("MEASurement1:GATE ON")
        # self.write(f"MEASurement1:GATE:ZCOupling ON")

    def edge_trigger(self, trigger_channel, trigger_level, trigger_edge):
        channel = 'CHAN' + str(trigger_channel) # 1 | 2 | 3 | 4

        self.trigger_mode(mode='NORM')
        
        # setting trigger to absolute without hysteresis
        self.write(f"TRIGger:LEVel{trigger_channel}:NOISe MANual")
        self.write(f'TRIGger:LEVel{trigger_channel}:NOISe:ABSolute 0')

        self.write(f'TRIG1:SOUR {channel}')
        self.write('TRIG1:TYPE EDGE')
        self.write(f'TRIG1:LEV{trigger_channel} {trigger_level}') # range: -10 to 10, increment: 1E-3
        self.write(f'TRIG1:EDGE:SLOP {trigger_edge}') # POS | NEG | EITH
    
    def width_trigger(self, trigger_channel, width_polarity='POS', width_range='LONG', width=100E-3, delta=0):
        channel = 'CHAN' + str(trigger_channel) # 1 | 2 | 3 | 4
        self.write(f'TRIG1:SOUR {channel}')
        self.write('TRIG1:TYPE WIDT')
        self.write(f'TRIG1:WIDT:POL {width_polarity}') # POS | NEG
        self.write(f'TRIG1:WIDT:RANG {width_range}') # WITHin | OUTSide | SHORter | LONGer
        self.write(f'TRIG1:WIDT:WIDT {width}')  ## Range: 100E-12 to 10000
                                                # Increment: 100E-9
                                                # *RST: 5E-9
                                                # Default unit: s
        self.write(f'TRIG1:WIDT:DELT {delta}')  ## Range: 0 to 432
                                                # Increment: 500E-12
                                                # *RST: 0
                                                # Default unit: s

    def timeout_trigger(self, trigger_channel, timeout_range='HIGH', timeout_time=1E-3):
        channel = 'CHAN' + str(trigger_channel) # 1 | 2 | 3 | 4
        self.write(f'TRIG1:SOUR {channel}')
        self.write(f'TRIG1:TYPE TIM')
        self.write(f'TRIG1:TIM:RANG {timeout_range}') # HIGH | LOW | EITHer
        self.write(f'TRIG1:TIM:TIME {timeout_time}') # Range: 100E-12 to 10000
                                                        # Increment: 100E-9
                                                        # *RST: 100E-9
                                                        # Default unit: s

    def trigger_level(self, trigger_channel, trigger_level):
        self.write(f'TRIG1:LEV{trigger_channel} {trigger_level}') # range: -10 to 10, increment: 1E-3

    def trigger_mode(self, mode):
        self.write(f'TRIG:MODE {mode}') # AUTO | NORMal | FREerun

    def force_trigger(self):
        self.write('TRIG:FORC')

    def trigger_status(self):
        status = self.write('ACQ:CURR?')
        status = int(status)
        return status

    ### Channel Settings ###

    def channel_BW(self, channel, channel_BW):
        """
        500: 'FULL'
        20: 'B20'
        200: 'B200'
        """
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        if channel_BW == 500:
            channel_BW = 'FULL'
        elif channel_BW == 20:
            channel_BW = 'B20'
        elif channel_BW == 200:
            channel_BW = 'B200'

        self.write(f'CHAN{channel}:BAND {channel_BW}')

    def channel_offset(self, channel, channel_offset):
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        self.write(f'CHAN{channel}:OFFS {channel_offset}')

    def channel_position(self, channel, channel_position):
        self.channel_offset(channel, 0)

        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        self.write(f'CHAN{channel}:POS {channel_position}')

    def channel_coupling(self, channel, channel_coupling):
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        self.write(f'CHAN{channel}:COUP {channel_coupling}') # DC | DCLimit | AC

    def channel_scale(self, channel, channel_scale):
        channel_state = self.write(f'CHAN{channel}:STAT?')
        if channel_state == '0':
            return None

        self.write(f'CHAN{channel}:SCAL {channel_scale}') # V/div

    def channel_state(self, channel, state='OFF'):
        self.write(f"CHANnel{channel}:STATe {state}")
        self.write(f"MEASurement{channel}:ENABle {state}")


    def channel_settings(self, state, channel=1, scale=1, position=0, label='IOUT', color='LIGHT_BLUE', rel_x_position=50, bandwidth=500, coupling='DCLimit', offset=0):
        """
        COLOR OPTIONS:

        - LIGHT_BLUE
        - YELLOW
        - PINK
        - GREEN
        - BLUE
        - ORANGE


        returns : state of the channel ('ON' or 'OFF')
        """

        self.channel_state(channel, state)

        self.channel_position(channel, position)
        self.channel_scale(channel, scale)
        
        if state == 'ON': print(f"CH{channel} - {label}")
        self.channel_label(channel, label, rel_x_position)
        self.channel_color(channel, color)

        self.channel_BW(channel, bandwidth)
        
        self.channel_coupling(channel, coupling)
        self.channel_offset(channel, offset)
        
        self.display_intensity()

        self.measure_enable(channel, state)

        return state

    def channel_label(self, channel, label, rel_x_position):

        self.write(f"DISPlay:SIGNal:LABel:REMove 'Label1', C{channel}W1")
        
        self.write(f"DISPlay:SIGNal:LABel:ADD 'Label1', C{channel}W1, '{label}', REL, {rel_x_position}, 0.1")

    def channel_color(self, channel, color):

        light_blue = int('ff26e9ff', 16)
        yellow = int('ffffff00', 16)
        pink = int('ffff49fb', 16)
        green = int('ff16e500', 16)
        blue = int('ff0080ff', 16)
        orange = int('ffff6000', 16)

        colors = {  'LIGHT_BLUE':light_blue,
                    'YELLOW': yellow,
                    'PINK': pink,
                    'GREEN': green,
                    'BLUE': blue,
                    'ORANGE': orange
        }

        signal = {
            '1': 2,
            '2': 5,
            '3': 8,
            '4': 11
        }

        self.write(f"DISPlay:COLor:SIGNal{signal[str(channel)]}:COLor {colors[color]}")
    
    def display_intensity(self, intensity=100):
        self.write(f"DISPlay:INTensity 0")
        self.write(f"DISPlay:INTensity {intensity}")

    def cursor(self, channel=1, cursor_set=1, X1=1, X2=1, Y1=0, Y2=0, type='VERT'):
        self.write(f"CURSor{cursor_set}:FUNCtion {type}") # HORizontal | VERTical | PAIRed
        self.write(f"CURS{cursor_set}:STAT ON")
        self.write(f"CURS{cursor_set}:SOUR C{channel}W1")
        self.write(f"CURS{cursor_set}:X1P {X1}")
        self.write(f"CURS{cursor_set}:X2P {X2}")
        self.write(f"CURS{cursor_set}:Y1P {Y1}")
        self.write(f"CURS{cursor_set}:Y2P {Y2}")

    """MEASURE"""

    def measure_enable(self, channel, state='ON'):
        self.write(f"MEASurement{channel}:ENABle {state}")

    def measure_source(self, channel):
        self.write(f"MEASurement{channel}:SOURce C{channel}W1")
        self.write(f"MEASurement{channel}:CATegory AMPTime")
    
    def measure_off(self, channel):
        self.write(f"MEASurement{channel}:AOFF")

    def measure(self, channel, measure_list):
        """
        HIGH | LOW | AMPLitude | MAXimum | MINimum | PDELta |
        MEAN | RMS | STDDev | POVershoot | NOVershoot | AREA |
        RTIMe | FTIMe | PPULse | NPULse | PERiod | FREQuency |
        PDCYcle | NDCYcle | CYCarea | CYCMean | CYCRms |
        CYCStddev | PULCnt | DELay | PHASe | BWIDth | PSWitching |
        NSWitching | PULSetrain | EDGecount | SHT | SHR | DTOTrigger |
        PROBemeter | SLERising | SLEFalling
        """

        # self.measure_enable(channel, state='ON')
        self.measure_source(channel)
        self.measure_off(channel)
        measure_list = measure_list.strip(" ").split(",")
        for type in measure_list:
            self.write(f"MEASurement{channel}:ADDitional {type}, ON")

    def cleanup(self):
        self.close()

class ElectronicLoad(Equipment):
    def __init__(self, address, comm_type='gpib', timeout=10000):
        super().__init__(address, comm_type, timeout)
        self.channel = [None] + [self.Channel(self, i) for i in range(1, 9)]

    def turn_on_all(self):
        for i in range(1, 9):
            self.channel[i].turn_on()

    def turn_off_all(self):
        for i in range(1, 9):
            self.channel[i].turn_off()

    

    def cleanup(self):
        self.turn_off_all()
        self.close()

    class Channel:
        def __init__(self, load, channel):
            self.load = load
            self.channel = channel
            self._cv = 0
            self._cc = 0
            self._cr = 0
            self._led_voltage = 0
            self._led_current = 0
            self._von = 0
            self._voltage = 0
            self._current = 0

        def dynamic(self, low=0, high=1, ton=0.0005, toff=0.0005): 
            self.load.write(f'CHAN {self.channel}')     
            self.load.write(f'CURR:DYN:L1 {low}')
            self.load.write(f'CURR:DYN:L2 {high}')
            self.load.write(f'CURR:DYN:RISE MAX')
            self.load.write(f'CURR:DYN:FALL MAX')
            self.load.write(f'CURR:DYN:T1 {ton}')
            self.load.write(f'CURR:DYN:T2 {toff}')

        @property
        def voltage(self):
            self.load.write(f'CHAN {self.channel}')
            self._voltage = float(self.load.write('FETC:VOLT?'))
            return self._voltage

        @property
        def current(self):
            self.load.write(f'CHAN {self.channel}')
            self._current = float(self.load.write('FETC:CURR?'))
            return self._current        

        @property
        def cv(self):
            return self._cv

        @property
        def cc(self):
            return self._cc

        @property
        def cr(self):
            return self._cr

        @property
        def led_voltage(self):
            return self._led_voltage
        
        @property
        def led_current(self):
            return self._led_current

        @property
        def von(self):
            return self._von


        @von.setter
        def von(self, voltage):
            self._von = voltage
            self.load.write(f'CHAN {self.channel}')
            self.load.write(f'CONF:VOLT:ON {self._von}')

        @cv.setter
        def cv(self, voltage):
            self._cv = voltage
            self.load.write(f'CHAN {self.channel}')
            self.load.write(f'VOLTAGE:L1 {self._cv}')
            self.load.write('MODE CV')

        @cc.setter
        def cc(self, current):
            self._cc = current
            self.load.write(f'CHAN {self.channel}')
            self.load.write(f'CURRENT:STATIC:L1 {self._cc}')
            self.load.write('MODE CC')

        @cr.setter
        def cr(self, resistance):
            self._cr = resistance
            self.load.write(f'CHAN {self.channel}')
            self.load.write(f'RESISTANCE:L1 {self._cr}')
            self.load.write(f'RESISTANCE:L2 {self._cr}')
            self.load.write('MODE CRH')

        @led_voltage.setter
        def led_voltage(self, voltage):
            self._led_voltage = voltage
            # not supported yet...
            self.load.write(f'CHAN {self.channel}')
            self.load.write(f'LED:VO {self._led_voltage}')
            self.load.write('MODE LEDH')

        @led_current.setter
        def led_current(self, current):
            self._led_current = current
            # not supported yet...
            self.load.write(f'CHAN {self.channel}')
            self.load.write(f'LED:IO {self._led_current}')
            self.load.write('MODE LEDH')

        def turn_on(self):
            self.load.write(f'CHAN {self.channel}')
            self.load.write('LOAD ON')

        def turn_off(self):
            self.load.write(f'CHAN {self.channel}')
            self.load.write('LOAD OFF')

        def short_on(self):
            self.load.write(f'CHAN {self.channel}')
            self.load.write('LOAD:SHOR ON')

        def short_off(self):
            self.load.write(f'CHAN {self.channel}')
            self.load.write('LOAD:SHOR OFF')

"""UNUSED CLASSES"""                 
class SignalGenerator(Equipment):
    def __init__(self, address, comm_type='usb', timeout=10000):
        super().__init__(address, comm_type, timeout)

    def cleanup(self):
        self.close()

class EMI(Equipment):
    def __init__(self, address, comm_type='lan', timeout=10000):
        super().__init__(address, comm_type, timeout)

    def cleanup(self):
        self.close()



"""CDO Classes"""
class Yokogawa_Oscilloscope_DLM5058(Equipment):
    def __init__(self, address, comm_type='lan', timeout=20000):
        super().__init__(address, comm_type, timeout)

 ##################### CH ON/OFF ###################    
    def channel_state(self, channel = 1, state ='ON'):
        self.write(f':CHAN{channel}:DISP {state}')

    def channel_state_all_off(self): 
        for channel in range(1, 9): 
            self.write(f':CHAN{channel}:DISP OFF')

 ##################### V&T /DIV #################### 
    
    def set_vdiv(self, channel = 1, vdiv = '1V'): #5mV to 100V
        self.write(f':CHAN{channel}:VDIV {vdiv}')

    def set_tdiv(self, tdiv = '10US'): #500ps to 500s
        self.write(f':TIM:TDIV {tdiv}')

 ##################### POSITION ####################

    def set_vert_position(self, channel = 1, vert_pos = 0): #-4div to 4div
        self.write(f':CHAN{channel}:POS {vert_pos}')

    def set_horiz_position(self, horiz_pos = 0): #Percentage 0% to 100%
        self.write(f':TRIG:POS {vert_pos}')

 ##################### ATTENUATION #################

    def set_probe_attenuation(self, channel = 1, attenuation = 10): #Add C{attenuation} if current
        self.write(f':CHAN{channel}:PROBE {attenuation}')

 ##################### TRIGGER #####################

    def set_trigger_mode(self, trig_mode = 'NORM'): #AUTO, NORMal, A(uto)LEVel, NSINgle
        self.write(f':TRIG:MODE {trig_mode}')

    def set_trigger_level(self, channel = 1, trig_lev = '1V'): 
        self.write(f':TRIG:SOUR:CHAN{channel}:LEV {trig_lev}')

 ##################### MEASURE #####################

    def disp_meas_off(self, channel = 1): 
        self.write(f':MEAS:CHAN{channel}:ALL OFF')

    def disp_meas_off_all(self): 
        for channel in range(1, 9): 
            self.write(f':MEAS:CHAN{channel}:ALL OFF')

    def disp_meas(self, channel = 1, parameter = 'FREQ'): 
        self.write(f':MEAS:CHAN{channel}:{parameter}:STATE ON') 
        ''' List of Parameters Drop-Down
            AMPLitude
            AVERage
            AVGFreq
            AVGPeriod
            BWIDth
            DELay
            DT
            DUTYcycle
            ENUMber
            FALL
            FREQuency
            HIGH
            LOW
            MAXimum
            MINimum
            NOVershoot
            NWIDth
            PERiod
            PNUMber
            POVershoot
            PTOPeak
            PWIDth
            RISE
            RMS
            SDEViation
            TY1Integ
            TY2Integ
            V1
            V2'''

    def get_meas_value(self, channel = 1, parameter = 'FREQ'):
        sleep(0.05) 
        result = self.write(f':MEAS:CHAN{channel}:{parameter}:VAL?').split(' ')[1]
        return result 

    def disp_meas_statistics(self, mode = 'CONT'): #OFF|ON|CONTinuous|CYCLe|HISTory
        self.write(f':MEAS:MODE {mode}')

    def set_sample_count(self, count = 1000): 
        self.write(f':ACQ:COUNT {count}')
        self.write(f':TRIG:MODE NORM')
        status = int(self.write(f':STATus:CONDition?'))
        while status > 0:
            status = int(self.write(f':STATus:CONDition?'))

    def get_min_mean_max_sd(self, channel = 1, parameter = 'FREQ'):
        result = {
            "maximum": self.write(f':MEASure:CHANnel{channel}:{parameter}:MAXimum?').split(' ')[1],
            "mean": self.write(f':MEASure:CHANnel{channel}:{parameter}:MEAN?').split(' ')[1],
            "minimum": self.write(f':MEASure:CHANnel{channel}:{parameter}:MINimum?').split(' ')[1],
            "sdeviation": self.write(f':MEASure:CHANnel{channel}:{parameter}:SDEViation?').split(' ')[1]
        }
        return result

 ##################### SCREENSHOT ###################
    def set_drive(self, drive = 'USB'): #FLAShmem|NETWork|USB 
        self.write(f' :IMAGE:SAVE:DRIVE {drive}')
    
    def set_background(self, background = 'GRAY'): #COLor|GRAY|OFF|REVerse 
        self.write(f' :IMAGE:TONE {background}')

    def set_filename(self,temperature = '25', test_id = ' ', condition = '5V_25V'): 
        self.write(f' :IMAGE:SAVE:NAME "{temperature}C_{test_id}_{condition}_"')

    def screenshot(self): 
        self.write(f' :IMAGE:EXECUTE')
    
 ##################### CLOSE ########################
    def stop(self):
        self.write(':STOP')

    def cleanup(self):
        self.close()

class Tektronix_SigGen_AFG31000(Equipment):
    def __init__(self, address, comm_type='gpib', timeout=20000):
        super().__init__(address, comm_type, timeout)

 ##################### CH ON/OFF ###################    
    def channel_state(self, channel = 1, state ='ON'):
        self.write(f':OUTP{channel}:STAT {state}')

    def channel_state_all_off(self): 
        for channel in range(1, 3): 
            self.write(f':OUTP{channel}:STAT OFF')

    def set_load_impedance(self, channel = 1, impedance = 'INF'): 
        self.write(f':OUTP{channel}:IMPedance {impedance}')

 ##################### CONTINUOUS ###################
    def out_cont_sine(self, channel = 1, frequency = '10kHz', phase = '0 DEG', low = '-500mV', high = '500mV', units = 'VPP'): 
        self.write(f' :SOURce{channel}:FUNCtion:SHAPe SIN')
        self.write(f' :SOURce{channel}:FREQuency:FIXed {frequency}')
        self.write(f' :SOURce{channel}:PHASe:ADJust {phase}')
        self.write(f' :SOURce{channel}:VOLTage:LEVel:IMMediate:LOW {low}')
        self.write(f' :SOURce{channel}:VOLTage:LEVel:IMMediate:HIGH {high}')
        self.write(f' :SOURce{channel}:VOLTage:UNIT {units}')

    def out_cont_square(self, channel = 1, frequency = '10kHz', phase = '0 DEG', low = '-500mV', high = '500mV', units = 'VPP'): 
        self.write(f' :SOURce{channel}:FUNCtion:SHAPe SQU')
        self.write(f' :SOURce{channel}:FREQuency:FIXed {frequency}')
        self.write(f' :SOURce{channel}:PHASe:ADJust {phase}')
        self.write(f' :SOURce{channel}:VOLTage:LEVel:IMMediate:LOW {low}')
        self.write(f' :SOURce{channel}:VOLTage:LEVel:IMMediate:HIGH {high}')
        self.write(f' :SOURce{channel}:VOLTage:UNIT {units}')

    def out_cont_ramp(self, channel = 1, frequency = '10kHz', phase = '0 DEG', low = '-500mV', high = '500mV', units = 'VPP', symmetry = 50.0): 
        self.write(f' :SOURce{channel}:FUNCtion:SHAPe RAMP')
        self.write(f' :SOURce{channel}:FREQuency:FIXed {frequency}')
        self.write(f' :SOURce{channel}:PHASe:ADJust {phase}')
        self.write(f' :SOURce{channel}:VOLTage:LEVel:IMMediate:LOW {low}')
        self.write(f' :SOURce{channel}:VOLTage:LEVel:IMMediate:HIGH {high}')
        self.write(f' :SOURce{channel}:VOLTage:UNIT {units}')
        self.write(f' :SOURce{channel}:FUNCtion:RAMP:SYMMetry {symmetry}')              

    def out_cont_pulse(self, channel = 1, frequency = '10kHz', phase = '0 DEG', low = '-500mV', high = '500mV', units = 'VPP', duty = 999, width = 'ABC'): #Choose duty or width
        self.write(f' :SOURce{channel}:FUNCtion:SHAPe PULSE')
        self.write(f' :SOURce{channel}:FREQuency:FIXed {frequency}')
        self.write(f' :SOURce{channel}:PHASe:ADJust {phase}')
        self.write(f' :SOURce{channel}:VOLTage:LEVel:IMMediate:LOW {low}')
        self.write(f' :SOURce{channel}:VOLTage:LEVel:IMMediate:HIGH {high}')
        self.write(f' :SOURce{channel}:VOLTage:UNIT {units}')
        
        if (duty != 999):
            self.write(f' :SOURce{channel}:PULSe:DCYCle {duty}')
            # print(duty)
        if (width != 'ABC'):
            self.write(f' :SOURce{channel}:PULSe:WIDTh {width}')
            print(width)
                
    def out_DC(self, channel = 1, voltage = '5V'): 
        self.write(f' :SOURce{channel}:FUNCtion:SHAPe DC')
        self.write(f' :SOURce{channel}:VOLTage:LEVel:IMMediate:OFFSet {voltage}')

 ##################### CLOSE ########################
    def stop(self):
        self.write(':STOP')

    def cleanup(self):
        self.close()

class Keithley_DC_2230G(Equipment):
    def __init__(self, address, comm_type='gpib', timeout=20000):
        super().__init__(address, comm_type, timeout)

 #################### SET VOL&CURR #################
    def set_volt_curr(self, channel ='CH1', voltage = 0.0, current = 0.0):
        self.write(f'APPL {channel}, {voltage}, {current}')

    def channel_state(self, channel ='CH1', state = 'ON'): #Individual
        self.write(f'INST {channel}')  
        self.write(f'CHAN:OUTP {state}')

    def channel_state_all(self, state = 'ON'): #For all channels
        self.write(f'OUTP {state}')          

 #################### CLOSE ########################
    def stop(self):
        self.write(':STOP')

    def cleanup(self):
        self.close()

class Fluke_BDMM_8808A(Equipment):
    def __init__(self, address, comm_type='serial', timeout=20000):
        super().__init__(address, comm_type, timeout)

 #################### BDMM #################
    def remote_mode(self):
        self.write(f'BDMM:REMS')
        

    def set_func(self, function = 'VDC'): #VDC, VAC, AAC, ADC, OHMS, CONTinuity, DIODE, FREQ
        self.write(f'BDMM:{function}')
        sleep(5)

    def get_meas_value(self):
        return self.write(f'BDMM:VAL?')

    def meas_hold(self):
        return self.write(f'BDMM:HOLD')

 #################### CLOSE ########################
    def stop(self):
        self.write(':STOP')

    def cleanup(self):
        self.close()


"""Charles' Custom Functions"""

# initialize variables
global Iout_index
global waveform_counter
global start
global waveforms_folder
waveform_counter = 0
Iout_index = 0
from datetime import datetime
from filemanager import path_maker, remove_file, move_file

def headers(test_name):
    global start
    print()
    print("="*80)
    print(f"Test: {test_name}")
    create_folder(test_name)
    start = datetime.now()
    print("="*80)

def create_folder(test_name):
    if not os.path.exists('waveforms'): os.mkdir('waveforms')
    waveforms_folder = f'waveforms/{test_name}'
    pathname = f"{os.getcwd()}/{waveforms_folder}"
    if not os.path.exists(pathname): os.mkdir(pathname)

def sfx():
    import winsound as ws
    ws.PlaySound("dingding.wav", ws.SND_ASYNC)
    sleep(2)  

def footers(waveform_counter):

    # from playsound import playsound
    # import winsound as ws
    # # playsound("test_complete.mp3")

    print("="*80)
    if waveform_counter > 0: print(f'{waveform_counter} waveforms captured.')
    print('test complete.')
    print()
    end = datetime.now()
    total_time = end-start
    temp = ""
    
    print(f"START TIME: {temp:>12}{start}")
    print(f"COMPLETION TIME: {temp:>7}{end}")
    print(f"TOTAL TESTING TIME: {temp:>16}{total_time}")
    print("="*80)
    
    sfx()    
    
def soak(soak_time):
    for seconds in range(soak_time, 0, -1):
        sleep(1)
        print(f"{seconds:5d}s", end="\r")
    print("       ", end="\r")

def convert_argv_to_int_list(a=[]):
    str_list = a.strip('[]').split(',')
    int_list = [int(x) for x in str_list]
    return int_list

def convert_argv_to_str_list(a=[]):
    str_list = a.strip('[]').split(',')
    return str_list

class LEDControl():

    """
    CLASS FUNCTION TO CONTROL LED LOADS
    
    """


    def __init__(self):

        from time import sleep, time
        import serial.tools.list_ports
        import math
        from pyfirmata import Arduino, util
        import pyfirmata

        # automatic detection of comm port
        ports = serial.tools.list_ports.comports()
        
        commPort = 'None'
        for i in range(0, len(ports)):

            port = str(ports[i])
            if 'USB Serial Device' in port:
                commPort = port[3:5]
                
        
        if commPort != 'None':
            try:
                # print(commPort)
                self.board = pyfirmata.Arduino(f'COM{commPort}')
                self.iterator = util.Iterator(self.board)
                self.iterator.start()
                print(f'LED Load Control: Connected to COM{commPort}')
                self.initialize_relays()
            except Exception as e:
                raise e
        else:
            raise ValueError


    def initialize_relays(self):
        self.RELAY1 = self.board.get_pin('d:10:o')
        self.RELAY2 = self.board.get_pin('d:9:o')
        self.RELAY3 = self.board.get_pin('d:8:o')

    def voltage(self, voltage: int):
        if voltage >= 46:
            self.RELAY1.write(1)
            self.RELAY2.write(0)
            self.RELAY3.write(0)
            print(f"LED set to {voltage}V.")
        elif voltage > 24 and voltage < 40:
            self.RELAY1.write(0)
            self.RELAY2.write(1)
            self.RELAY3.write(0)
            print(f"LED set to {voltage}V.")
        elif voltage <= 24 and voltage > 0:
            self.RELAY1.write(0)
            self.RELAY2.write(0)
            self.RELAY3.write(1)
            print(f"LED set to {voltage}V.")
        elif voltage == 0:
            self.RELAY1.write(0)
            self.RELAY2.write(0)
            self.RELAY3.write(0)
            print(f"LED set to NL.")
        else: print("Invalid LED.")

def tts(message):
    language = 'fr'
    myobj = gTTS(text=message, lang=language, slow=False)
    path_maker(f'{os.getcwd()}/sfx')
    if not os.path.exists(f"{os.getcwd()}/sfx/{message}.mp3"):
        myobj.save(f"{message}.mp3")
        move_file(f"{message}.mp3", f"{os.getcwd()}/sfx/{message}.mp3")
    playsound(f"{os.getcwd()}/sfx/{message}.mp3")
    print(f"{message}")



def prompt(message):
    language = 'en'
    myobj = gTTS(text=message, lang=language, slow=False)
    path_maker(f'{os.getcwd()}/sfx')
    if not os.path.exists(f"{os.getcwd()}/sfx/{message}.mp3"):
        myobj.save(f"{message}.mp3")
        move_file(f"{message}.mp3", f"{os.getcwd()}/sfx/{message}.mp3")
    playsound(f"{os.getcwd()}/sfx/{message}.mp3")
    input(f"{message}")









import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt

def get_anchor(col, row):
    """
    get anchor given a numerical col row  -> (col = 2, row = 4 -> 'B4')
    returns anchor (str)
    """
    anchor = f"{get_column_letter(col)}{row}"
    return anchor

def col_row_extractor(excel_coordinate):
    """
    extract col and row given an excel coordinate (i.e. 'B4' -> col = 2, row = 4)

    excel_coordinate : i.e. 'B4' (str)
    returns col, row (int)
    """
    coordinates = coordinate_from_string(excel_coordinate)
    col = column_index_from_string(coordinates[0])
    row = coordinates[1]
    return col, row

def excel_to_df(filename, sheet_name, start_corner, end_corner):
    """
    reading dataframe from excel.
    
    filename     : must include full filename path (cwd + path + file.extension)
    sheet_name   : sheet name in excel file
    start_corner : cell coordinate to start selection of data
    end_corner   : cell coordinate to end selection of data

    returns df
    """

    # print(f"reading dataframe from {filename} {sheet_name}")

    start_col, start_row = col_row_extractor(start_corner)
    end_col, end_row = col_row_extractor(end_corner)

    skiprows = start_row - 2
    usecols = f'{get_column_letter(start_col)}:{get_column_letter(end_col)}'
    nrows = end_row - start_row + 1

    return pd.read_excel(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)
    # return pd.read_csv(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)

def df_to_excel(wb, sheet_name, df, anchor):
    """
    writing dataframe to excel.

    wb          : workbook
    sheet_name  : sheet name in excel file
    df          : dataframe
    anchor      : anchor point in excel

    returns None
    """

    # print(f"writing in {wb} {sheet_name}")

    start_col, start_row = col_row_extractor(anchor)
    df_row_len, df_col_len = df.shape
    end_row = start_row + df_row_len - 1
    end_col = start_col + df_col_len - 1


    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            wb[sheet_name][f'{get_column_letter(col)}{row}'] = df.iloc[row-start_row, col-start_col]

def image_to_excel(wb, sheet_name, filename, folder_path, anchor):
    """
    writing image to excel.

    image size -> 39R, 16C
    wb          : workbook
    sheet_name  : sheet name in excel file
    filename    : filename of theh image
    folder_path : image location
    anchor      : anchor point in excel
    """

    file = os.getcwd() + folder_path + filename
    img = openpyxl.drawing.image.Image(file)
    img.anchor = anchor
    img.width = 1026
    img.height = 762

    wb[sheet_name].add_image(img)

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series


  


def create_scatter_chart(title="Efficiency (%)", style=2, x_title='Input Voltage (VAC)', y_title='Efficiency (%)',
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 5):
 

    chart = ScatterChart()
    chart.title = title
    chart.style = style
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.height = 10 # default is 7.5
    chart.width = 20 # default is 15

    chart.x_axis.scaling.min = x_min_scale
    chart.x_axis.scaling.max = x_max_scale
    chart.y_axis.scaling.min = y_min_scale
    chart.y_axis.scaling.max = y_max_scale

    chart.x_axis.majorUnit = x_major_unit
    chart.x_axis.minorUnit = x_minor_unit
    chart.y_axis.majorUnit = y_major_unit
    chart.y_axis.minorUnit = y_minor_unit


    return chart

def reset_chartsheet(wb):
    chart_sheet = wb["Chart"]
    try:
        no_of_existing_charts = len(chart_sheet._charts)
        for i in range(no_of_existing_charts):
            del chart_sheet._charts[(i-1)]
    except: pass
    
    return chart_sheet

def save_chartsheet(chart_sheet, chart, chart_position):
    chart_sheet.add_chart(chart, chart_position)

def append_series(path, wb, ws_name, x_anchor, last_row_anchor, series_title, chart):
    ## used in der-727_chart_compiler.py
    
    df = excel_to_df(path, ws_name, x_anchor, last_row_anchor)
    ws = wb[ws_name]
    xcol, xrow = col_row_extractor(x_anchor)
    last_col, last_row = col_row_extractor(last_row_anchor)
    xvalues = Reference(ws, min_col=xcol, min_row=xrow, max_row=last_row)
    values = Reference(ws, min_col=last_col, min_row=xrow, max_row=last_row)
    series = Series(values, xvalues, title=series_title)
    series.marker=openpyxl.chart.marker.Marker('auto')
    series.graphicalProperties.line.noFill=False
    chart.series.append(series) 

if __name__ == '__main__':
    pass


