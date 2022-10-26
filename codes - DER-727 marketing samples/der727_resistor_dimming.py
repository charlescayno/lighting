# June 23, 2022

"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_chartsheet
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_chartsheet
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	

##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.11.10"
sig_gen_address = '11'
dimming_power_meter_address = 21


"""USER INPUT"""

vin_list = [90,115,230,277]
led_list = [48,36,24]
resistor_list = [1,2,3,4,5,6,7,8,9,10,20,30,40,50,60,70,80,90,100] # kohms

vout = 48
iout = 1.59
soak_time = 10
iout_channel = 3

unit = input(">> UNIT #: ")
test = "Resistor Dimming"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/Marketing Samples/{unit}'
path = path_maker(f'{waveforms_folder}')

"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
# scope = Oscilloscope(scope_address)
led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output(times=1):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(2)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)


def create_new_excel_container():
    """

    INPUT FILE: "blank.xlsx" at automation/codes folder
    OUTPUT FILE: "{test}.xlsx" at {waveforms_folder} folder

    """
    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{test}_{unit}.xlsx"
    # while os.path.exists(dst): # if existing add next iteration number
    #     dst = dst.split('.xlsx')
    #     dst, iteration = dst[0].split('_')
    #     new_iteration = int(iteration) + 1
    #     dst = dst + f"_{new_iteration}.xlsx"
    # print(dst)
    # shutil.copyfile(src, dst)
    if not os.path.exists(dst): shutil.copyfile(src, dst)
    return dst



def create_output_excel_result(dst, df, anchor, led): 

    wb = load_workbook(dst)
    df_to_excel(wb, f"{led}", df, anchor)
    wb.save(dst)


def header_data(df_header_list):
    """
    header_data: creates dataframe with header specified by user
    
    NOTE:   - header_data and collect_data are partner functions
            - header_data need to be used first before collect_data
    
    input: df_header_list
    """
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list
    # print(df)

    return df

def collect_data(df, resistor, led, vin): 
    """
    collect_data: collects data. This function can be modified according to the user's needs
    
    output: print specific output list, and put it in the existing df
    """
    ####################################################################################
    res = resistor
    led_load = led
    vac = vin
    freq = ac.set_freq(vin)
    vac_actual = float(f"{pms.voltage:.2f}")
    iin = float(f"{pms.current*1000:.2f}")
    pin = float(f"{pms.power:.3f}")
    pf = float(f"{pms.pf:.4f}")
    thd = float(f"{pms.thd:.2f}") # AC
    vo1 = float(f"{pml.voltage:.3f}")
    io1 = float(f"{pml.current*1000:.2f}")
    po1 = float(f"{pml.power:.3f}")
    vreg1 = float(f"{100*(float(vo1)-vout)/float(vo1):.4f}")
    iout1 = float(io1)/1000
    try: ireg1 = float(f"{100*(iout1-iout)/iout1:.4f}")
    except: ireg1 = 0
    eff = float(f"{100*(float(po1))/float(pin):.4f}")

    output_list = [res, led_load, vac, freq, vac_actual, iin, pin, pf, thd, vo1, io1, po1, vreg1, ireg1, eff] # AC

    ####################################################################################

    df.loc[len(df)] = output_list
    print(df)
    return df

def operation():
    dst = create_new_excel_container()
    df_header_list = ['Resistor', 'led_load', 'Vin', 'Freq (Hz)', 'Vac (VAC)', 'Iin (mA)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (mA)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)'] 
    df = header_data(df_header_list)


    for resistor in resistor_list:

        input(f">> Change resistor to {resistor} kOhms.")

        for led in led_list:

            led_ctrl.voltage(led)

            for vin in vin_list:

                ac.voltage = vin
                ac.turn_on()
                if resistor == 0:
                    soak(240)
                else:
                    soak(soak_time)

                df = collect_data(df, resistor, led, vin)
 
                create_output_excel_result(dst, header_data(df_header_list), "B1", led)
                create_output_excel_result(dst, df[(df.Vin == 90) & (df.led_load == led)], "B2", led)
                
                create_output_excel_result(dst, header_data(df_header_list), "R1", led)
                create_output_excel_result(dst, df[(df.Vin == 115) & (df.led_load == led)], "R2", led)
            
                create_output_excel_result(dst, header_data(df_header_list), "AH1", led) 
                create_output_excel_result(dst, df[(df.Vin == 230) & (df.led_load == led)], "AH2", led)
            
                create_output_excel_result(dst, header_data(df_header_list), "AX1", led)
                create_output_excel_result(dst, df[(df.Vin == 277) & (df.led_load == led)], "AX2", led)

            discharge_output(1)

def main():
    global waveform_counter
    discharge_output(1)
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)