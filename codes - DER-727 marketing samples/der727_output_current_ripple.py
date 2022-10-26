## UPDATED AUGUST 08, 2022

"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	

##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.111"
sig_gen_address = '11'
dimming_power_meter_address = 21



"""USER INPUT"""

vin_list = [90,115,230,277]
led_list = [48, 36, 24]

iout_channel = 3

test = "Output Current Ripple"
unit = input(">> UNIT #: ")
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/Marketing Samples/{unit}'
path = path_maker(f'{waveforms_folder}')

"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output():
    ac.turn_off()
    for i in range(1):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(2)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    global condition
    global time_division
    
    scope.channel_settings(state='OFF', channel=1, scale=0.4, position=-4, label="Vfb",
                            color='YELLOW', rel_x_position=20, bandwidth=20, coupling='DCLimit', offset=0)
    
    
    scope.channel_settings(state='OFF', channel=2, scale=100, position=-4, label="Drain Voltage",
                            color='LIGHT_BLUE', rel_x_position=40, bandwidth=500, coupling='DCLimit', offset=0)
    
    
    scope.channel_settings(state='ON', channel=3, scale=0.02, position=-5, label="Output Current",
                            color='YELLOW', rel_x_position=50, bandwidth=20, coupling='DCLimit', offset=1.5)
    
    scope.channel_settings(state='OFF', channel=4, scale=100, position=0, label="Input Voltage",
                            color='YELLOW', rel_x_position=80, bandwidth=20, coupling='AC', offset=0)
    
    
    # scope.measure(1, "MAX,RMS")
    # scope.measure(2, "MAX,RMS")
    scope.measure(3, "MAX,MIN,RMS,MEAN,PDELta")
    # scope.measure(4, "MAX,MIN,RMS,MEAN,PDELta")

    # MAX,MIN,RMS,MEAN,PDELta

    scope.record_length(50E6)
    scope.time_position(50)
    scope.time_scale(200E-6)

    # scope.remove_zoom()
    # scope.add_zoom(rel_pos=50, rel_scale=25)
    
    trigger_channel = 3
    trigger_level = 0
    trigger_edge = 'POS'
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)
    scope.trigger_mode('AUTO')
    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1


def find_trigger(channel, trigger_delta):

    scope.edge_trigger(channel, 0, 'POS')

    # finding trigger level
    scope.run_single()
    scope.trigger_mode('AUTO')
    soak(5)
    scope.trigger_mode('NORM')
    

    # get initial peak-to-peak measurement value
    labels, values = scope.get_measure(channel)
    max_value = float(values[0])
    max_value = float(f"{max_value:.4f}")

    # set max_value as initial trigger level
    trigger_level = max_value
    scope.edge_trigger(channel, trigger_level, 'POS')

    # check if it triggered within 3 seconds
    scope.run_single()
    soak(3)
    trigger_status = scope.trigger_status()

    # increase trigger level until it reaches the maximum trigger level
    while (trigger_status == 1):
        trigger_level += trigger_delta
        scope.edge_trigger(channel, trigger_level, 'POS')

        # check trigger status
        scope.run_single()
        soak(3)
        trigger_status = scope.trigger_status()

    # decrease trigger level below to get the maximum trigger possible
    trigger_level -= 2*trigger_delta
    scope.edge_trigger(channel, trigger_level, 'POS')
    # print(f'Maximum trigger level found at: {trigger_level}')


def create_output_excel_result(df):
    
    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{test}.xlsx"
    if os.path.exists(dst):
        dst = dst.split('.xlsx')[0]
        try:
            iteration = float(dst.split('_')[1]) + 1
            dst = dst + f"_{iteration}.xlsx"
        except:
            dst = dst + '_1.xlsx'
    shutil.copyfile(src, dst)

    wb = load_workbook(dst)
    df_to_excel(wb, "Sheet1", df, 'B2')
    wb.save(dst)


def collect_data(led, v_input):
    
    """COLLECT DATA"""
    load = led
    vac = v_input
    labels, values = scope.get_measure(iout_channel)
    iout_pkpk = float(values[2])
    iout_mean = float(values[3])
    iout_ripple = 100*iout_pkpk/iout_mean
    iout_flicker = 100*iout_pkpk/(2*iout_mean)
    output_list = [load, vac, iout_pkpk, iout_mean, iout_ripple, iout_flicker]
    
    return output_list

def operation():
    df_header_list = ['LED', 'Vin', 'Ipk-pk (A)', 'Imean (A)', '%Ripple', 'Flicker(%)']
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list

    for led in led_list:

        led_ctrl.voltage(led)
        
        for vin in vin_list:

            ac.voltage = vin
            ac.turn_on()
            sleep(5)

            find_trigger(channel=iout_channel, trigger_delta=0.001)
            
            scope.run_single()
            sleep(3)
            

            discharge_output()
            filename = f"Iout Ripple - {led}V, {vin}Vac.png"
            screenshot(filename, path)


            output_list = collect_data(led, vin)
            df.loc[len(df)] = output_list

            print(output_list)

            create_output_excel_result(df)

    
    


def main():
    global waveform_counter
    discharge_output()
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)