"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
# date = now.strftime('%Y_%m_%d')
date = now.strftime('%Y_%m_%d_%H%M')


import shutil
import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt


from openpyxl import Workbook, load_workbook



def col_row_extractor(excel_coordinate):
    """
    extract col and row given an excel coordinate (i.e. 'B4' -> col = 2, row = 4)

    excel_coordinate : i.e. 'B4' (str)
    returns col, row (int)
    """
    coordinates = coordinate_from_string(excel_coordinate)
    col = column_index_from_string(coordinates[0])
    row = coordinates[1]
    return col, row

def excel_to_df(filename, sheet_name, start_corner, end_corner):
    """
    reading dataframe from excel.
    
    filename     : must include full filename path (cwd + path + file.extension)
    sheet_name   : sheet name in excel file
    start_corner : cell coordinate to start selection of data
    end_corner   : cell coordinate to end selection of data

    returns df
    """

    # print(f"reading dataframe from {filename} {sheet_name}")

    start_col, start_row = col_row_extractor(start_corner)
    end_col, end_row = col_row_extractor(end_corner)

    skiprows = start_row - 2
    usecols = f'{get_column_letter(start_col)}:{get_column_letter(end_col)}'
    nrows = end_row - start_row + 1

    return pd.read_excel(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)

def df_to_excel(wb, sheet_name, df, anchor):
    """
    writing dataframe to excel.

    wb          : workbook
    sheet_name  : sheet name in excel file
    df          : dataframe
    anchor      : anchor point in excel

    returns None
    """

    # print(f"writing in {wb} {sheet_name}")

    start_col, start_row = col_row_extractor(anchor)
    df_row_len, df_col_len = df.shape
    end_row = start_row + df_row_len - 1
    end_col = start_col + df_col_len - 1


    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            wb[sheet_name][f'{get_column_letter(col)}{row}'] = df.iloc[row-start_row, col-start_col]

def image_to_excel(wb, sheet_name, filename, folder_path, anchor):
    """
    writing image to excel.

    image size -> 39R, 16C
    wb          : workbook
    sheet_name  : sheet name in excel file
    filename    : filename of theh image
    folder_path : image location
    anchor      : anchor point in excel
    """

    file = os.getcwd() + folder_path + filename
    img = openpyxl.drawing.image.Image(file)
    img.anchor = anchor
    img.width = 1026
    img.height = 762

    wb[sheet_name].add_image(img)



##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.11.20"
sig_gen_address = '11'
dimming_power_meter_address = 21
"""USER INPUT"""
vin_list = [90,100,110,115,120,132,180,200,230,265,277]
iout_list = [1.59]
vout = 48

special_test_condition = sys.argv[1]
test = "LineRegulation_CR"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/{test}'
path = path_maker(f'{waveforms_folder}')


"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
# scope = Oscilloscope(scope_address)

def discharge_output():
    ac.turn_off()
    for i in range(1,9):
        eload.channel[i].cc = 1
        eload.channel[i].turn_on()
        eload.channel[i].short_on()
    sleep(1)
    for i in range(1,9):
        eload.channel[i].turn_off()
        eload.channel[i].short_off()
    sleep(1)

def operation():

    

    df = pd.DataFrame(columns = ['iout', 'Vin', 'Freq', 'Vac', 'Iin', 'Pin', 'PF', 'THD', 'Vo1', 'Io1', 'Po1', 'Vreg1', 'Ireg1', 'Eff'])
    df_header = df.columns.values.tolist()

    for iout in iout_list:

        eload.channel[5].cr = vout/iout
        eload.channel[5].turn_on()

        sleep(3)

        for vin in vin_list:
            ac.voltage = vin
            ac.turn_on()
        
            soak(15)

            cc_load = iout
            vac = vin
            freq = ac.set_freq(vin)
            vin = float(f"{pms.voltage:.2f}")
            iin = float(f"{pms.current*1000:.2f}")
            pin = float(f"{pms.power:.3f}")
            pf = float(f"{pms.pf:.4f}")
            thd = float(f"{pms.thd:.2f}")
            vo1 = float(f"{pml.voltage:.3f}")
            io1 = float(f"{pml.current*1000:.2f}")
            po1 = float(f"{pml.power:.3f}")
            vreg1 = float(f"{100*(float(vo1)-vout)/float(vo1):.4f}")
            iout1 = float(io1)/1000
            ireg1 = float(f"{100*(iout1-iout)/iout1:.4f}")
            eff = float(f"{100*(float(po1))/float(pin):.4f}")

            output_list = [cc_load, vac, freq, vin, iin, pin, pf, thd, vo1, io1, po1, vreg1, ireg1, eff]

            df.loc[len(df)] = output_list

            print(df)

    # replace existing file to the template
    src = f"{os.getcwd()}/template_line_regulation.xlsx"
    dst = f"{waveforms_folder}/{date}_LineRegulation_CR_{special_test_condition}.xlsx"
    shutil.copyfile(src, dst)


    wb = load_workbook(dst)
    df_to_excel(wb, "Line Regulation", df, 'A2')
    wb.save(dst)

    soak(3)


def main():
    global waveform_counter
    discharge_output()
    operation()
    discharge_output()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)