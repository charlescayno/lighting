"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Tektronix_SigGen_AFG31000, Keithley_DC_2230G
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	
import numpy as np
##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30
load_power_meter_address = 2
dimming_power_meter_address = 10
eload_address = 8
scope_address = "10.125.11.10"
dc_source_address = '4'
"""USER INPUT"""

vin_list = [90,115,230,277] # Vac
# vin_list = [90]
led_list = [36] # V
# led_list = [48]
vout = 48 # V
iout = 1.59 # A
soak_time = 2 # s

start_bit = 0
end_bit = 255
inc_bit = 1

test = "DALI"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/Final Data/Dimming/{test}'
path = path_maker(f'{waveforms_folder}')

"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
pmdc = PowerMeter(dimming_power_meter_address)
eload = ElectronicLoad(eload_address)
dc = Keithley_DC_2230G(dc_source_address)
led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output(times=1):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(2)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def create_new_excel_container():
    """
    INPUT FILE: "blank.xlsx" at automation/codes folder
    OUTPUT FILE: "{test}.xlsx" at {waveforms_folder} folder
    """
    src = f"{os.getcwd()}/blank.xlsx"
    dst = f"{waveforms_folder}/{test}_0.xlsx"
    while os.path.exists(dst): # if existing add next iteration number
        dst = dst.split('.xlsx')
        dst, iteration = dst[0].split('_')
        new_iteration = int(iteration) + 1
        dst = dst + f"_{new_iteration}.xlsx"
    print(dst)
    shutil.copyfile(src, dst)
    return dst

def create_output_excel_result(dst, df_header_list, df, led, vin): 
    global test
    wb = load_workbook(dst)

    if led == 48:
        row = 2
    if led == 36:
        row = 259
    if led == 24:
        row = 516
    
    if vin == 90:
        col = 'B'
    if vin == 115:
        col = 'T'
    if vin == 230:
        col = 'AL'
    if vin == 277:
        col = 'BD'

    header_anchor = f"{col}{row}"
    data_anchor = f"{col}{row+1}"

    df_header = header_data(df_header_list)
    df_data = df[(df.Vin == vin) & (df.led_load == led)]

    sheet_name = f"{test}"
    
    df_to_excel(wb, sheet_name, df_header, header_anchor)
    df_to_excel(wb, sheet_name, df_data, data_anchor)
    
    wb.save(dst)

def header_data(df_header_list):
    """
    header_data: creates dataframe with header specified by user
    
    NOTE:   - header_data and collect_data are partner functions
            - header_data need to be used first before collect_data
    
    input: df_header_list
    """
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list

    return df

def collect_data(df, bit, dim, led, vin): 
    """
    collect_data: collects data. This function can be modified according to the user's needs
    
    output: print specific output list, and put it in the existing df
    """
    ####################################################################################
    actual_dim_voltage = float(f"{pmdc.voltage:.6f}")
    led_load = led
    vac = vin
    freq = ac.set_freq(vin)
    try:
        vac_actual = float(f"{pms.voltage:.2f}")
    except:
        vac_actual = vin
    iin = float(f"{pms.current*1000:.2f}")
    pin = float(f"{pms.power:.3f}")
    pf = float(f"{pms.pf:.4f}")
    thd = float(f"{pms.thd:.2f}") # AC
    vo1 = float(f"{pml.voltage:.3f}")
    io1 = float(f"{pml.current*1000:.2f}")
    po1 = float(f"{pml.power:.3f}")
    vreg1 = float(f"{100*(float(vo1)-led)/float(vo1):.4f}")
    iout1 = float(io1)/1000
    try: ireg1 = float(f"{100*(iout1-iout)/iout1:.4f}")
    except: ireg1 = 0
    eff = float(f"{100*(float(po1))/float(pin):.4f}")

    output_list = [bit, dim, actual_dim_voltage, led_load, vac, freq, vac_actual, iin, pin, pf, thd, vo1, io1, po1, vreg1, ireg1, eff] # AC
    ####################################################################################

    df.loc[len(df)] = output_list
    return df
    
def set_dim_voltage(dim):
    dc.set_volt_curr(channel ='CH1', voltage = dim, current = 1.0)
    dc.channel_state(channel ='CH1', state = 'ON')
    soak(1)

def dali(bit):
    a = bit-1
    b = 253/3
    y = (10**(a/b - 1))/100 # dali equation in percentage
    dim = y*10 # dim voltage equivalent at 10 V

    if bit == 0:
        soak_time = 60
        dim = 0
    else:
        soak_time = 2
    soak(soak_time)
    set_dim_voltage(dim) # control dc source
    return dim

def reset_dc_source():
    set_dim_voltage(1)
    soak(2)
    set_dim_voltage(0)








def operation():

    dst = create_new_excel_container()
    df_header_list = ['bit', 'dim', 'dim_actual', 'led_load', 'Vin', 'Freq (Hz)', 'Vac (VAC)', 'Iin (mA)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (mA)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)'] 
    df_header = header_data(df_header_list)
    df = df_header
    
    reset_dc_source()

    for led in led_list:
        
        led_ctrl.voltage(led)
        soak(2)

        for vin in vin_list:

            ac.voltage = vin
            ac.turn_on()
        
            for bit in range(0, 255, 1):
                dim = dali(bit)

                df = collect_data(df, bit, dim, led, vin)
                print(df[(df.Vin == vin) & (df.led_load == led)])     
        
            create_output_excel_result(dst, df_header_list, df, led, vin)

            discharge_output(3)
            reset_dc_source()












def main():
    global waveform_counter
    discharge_output(2)
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)





           