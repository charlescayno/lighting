## UPDATED AUGUST 08, 2022

"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	

##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.111"
sig_gen_address = '11'
dimming_power_meter_address = 21



"""USER INPUT"""

vin_list = [90,115,230,277]
led_list = [48,36,24]

trig_channel = 4

# special_test_condition = input("Enter special test condition: ")
test = "Startup Profile"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/Final Data (REV C)/{test}'
path = path_maker(f'{waveforms_folder}')

"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output():
    ac.turn_off()
    for i in range(4):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(1)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(1)

def scope_settings():
    global condition
    global time_division
    
    scope.channel_settings(state='OFF', channel=1, scale=0.4, position=-4, label="Vfb",
                            color='YELLOW', rel_x_position=20, bandwidth=20, coupling='DCLimit', offset=0)
    
    
    scope.channel_settings(state='OFF', channel=2, scale=100, position=-4, label="Drain Voltage",
                            color='LIGHT_BLUE', rel_x_position=40, bandwidth=500, coupling='DCLimit', offset=0)
    
    
    scope.channel_settings(state='ON', channel=3, scale=0.2, position=-4, label="Output Current",
                            color='PINK', rel_x_position=40, bandwidth=20, coupling='DCLimit', offset=0)
    
    scope.channel_settings(state='ON', channel=4, scale=200, position=-2, label="Input Voltage",
                            color='YELLOW', rel_x_position=60, bandwidth=20, coupling='AC', offset=0)
    
    ## MEASURE OPTIONS: MAX,MIN,RMS,MEAN,PDELta
    
    # scope.measure(1, "MAX,RMS")
    # scope.measure(2, "MAX,RMS")
    scope.measure(3, "MAX,RMS")
    scope.measure(4, "MAX,RMS")
    

    scope.record_length(50E6)
    scope.time_position(10)
    scope.time_scale(100E-3)

    # scope.remove_zoom()
    # scope.add_zoom(rel_pos=50, rel_scale=0.04)
    
    trigger_channel = trig_channel
    trigger_level = 10
    trigger_edge = 'POS'
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1


def startup_90degPhase(vin):
    # ac.write("OUTP:STAT?")
    # ac.write("SYST:ERR?")
    ac.voltage = 0
    voltage = vin
    frequency = ac.set_freq(vin) 
    # ac.write("OUTP ON")
    ac.turn_on()
    ac.write("TRIG:TRAN:SOUR BUS")
    ac.write("ABORT")
    ac.write("LIST:DWEL 1, 1, 1")

    ac.write("VOLT:MODE LIST")
    ac.write(f"LIST:VOLT {voltage}, {voltage}, {voltage}")
    ac.write("VOLT:SLEW:MODE LIST")
    ac.write("LIST:VOLT:SLEW 9.9e+037, 9.9e+037, 9.9e+037")

    ac.write("FREQ:MODE LIST")
    ac.write(f"LIST:FREQ {frequency}, {frequency}, {frequency}")
    ac.write("FREQ:SLEW:MODE LIST")
    ac.write("LIST:FREQ:SLEW 9.9e+037, 9.9e+037, 9.9e+037")

    ac.write("VOLT:OFFS:MODE FIX")
    ac.write("VOLT:OFFS:SLEW:MODE FIX")

    ac.write("PHAS:MODE LIST")
    # ac.write("LIST:PHAS 270, 270, 270")
    ac.write("LIST:PHAS 90, 90, 90")

    ac.write("CURR:PEAK:MODE LIST")
    ac.write("LIST:CURR 40.4, 40.4, 40.4")

    ac.write("FUNC:MODE FIX")
    ac.write("LIST:TTLT ON,OFF,OFF")
    ac.write("LIST:STEP AUTO")
    # ac.write("SYST:ERR?")

    ac.write("OUTP:TTLT:STAT ON")
    ac.write("OUTP:TTLT:SOUR LIST")
    ac.write("TRIG:SYNC:SOUR PHASE")
    ac.write("TRIG:SYNC:PHAS 0.0")
    ac.write("TRIG:TRAN:DEL 0")
    ac.write("Sens:Swe:Offs:Poin 0")
    ac.write("TRIG:ACQ:SOUR TTLT")
    ac.write("INIT:IMM:SEQ3")
    ac.write("LIST:COUN 1")
    ac.write("INIT:IMM:SEQ1")
    ac.write("TRIG:TRAN:SOUR BUS")
    # ac.write("SYST:ERR?")

    ac.write("TRIG:IMM")

def operation():

    for led in led_list:

        led_ctrl.voltage(led)
        
        for vin in vin_list:

            scope.run_single()
            sleep(3)

            startup_90degPhase(vin)
            sleep(3)

            
            discharge_output()
            input("Adjust cursor. Press ENTER to continue...")
            
            
            filename = f"Startup Profile - {led}V, {vin}Vac.png"
            screenshot(filename, path)
            
            


def main():
    global waveform_counter
    discharge_output()
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)