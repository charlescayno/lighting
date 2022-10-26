"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
import numpy as np
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Keithley_DC_2230G
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
# date = now.strftime('%Y_%m_%d')
date = now.strftime('%Y_%m_%d_%H%M')
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_chartsheet
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
dimming_power_meter_address = 10
eload_address = 8
scope_address = "10.125.11.20"
dc_source_address = '4'

##################################################################################

vin_list = [90,100,110,115,120,132,180,200,230,265,277]
vout = 48
iout = 1.59
iout_list = [iout]
led_list = [48]
led_list = convert_argv_to_int_list(sys.argv[1])
test = "LED line regulation"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/{test}'
path = path_maker(f'{waveforms_folder}')

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
# scope = Oscilloscope(scope_address)

def discharge_output():
    ac.turn_off()
    for i in range(1,9):
        eload.channel[i].cc = 1
        eload.channel[i].turn_on()
        eload.channel[i].short_on()
    sleep(1)
    for i in range(1,9):
        eload.channel[i].turn_off()
        eload.channel[i].short_off()
    sleep(1)



def collect_data(v_input):
    
    """COLLECT DATA"""
    vac = v_input
    freq = ac.set_freq(v_input)
    vin = float(f"{pms.voltage:.2f}")
    iin = float(f"{pms.current*1000:.2f}")
    pin = float(f"{pms.power:.3f}")
    pf = float(f"{pms.pf:.4f}")
    thd = float(f"{pms.thd:.2f}") # AC
    vo1 = float(f"{pml.voltage:.3f}")
    io1 = float(f"{pml.current*1000:.2f}")
    po1 = float(f"{pml.power:.3f}")
    vreg1 = float(f"{100*(float(vo1)-vout)/float(vo1):.4f}")
    iout1 = float(io1)/1000
    try: ireg1 = float(f"{100*(iout1-iout)/iout1:.4f}")
    except: ireg1 = 0
    eff = float(f"{100*(float(po1))/float(pin):.4f}")

    output_list = [vac, freq, vin, iin, pin, pf, thd, vo1, io1, po1, vreg1, ireg1, eff] # AC
    
    return output_list



def operation():

    df = pd.DataFrame(columns = ['Vin', 'Freq', 'Vac', 'Iin', 'Pin', 'PF', 'THD', 'Vo1', 'Io1', 'Po1', 'Vreg1', 'Ireg1', 'Eff'])
    df_header = df.columns.values.tolist()

    for led in led_list:

        input(f">> Change LED to {led}V.")

        for v_input in vin_list:

            ac.voltage = v_input
            ac.turn_on()
        
            soak(15)

            output_list = collect_data(v_input)
            df.loc[len(df)] = output_list
            print(df)


    print()
    print(df)

    # replace existing file to the template
    src = f"{os.getcwd()}/template.xlsx"
    dst = f"{waveforms_folder}/{led}V LED load - Line Regulation.xlsx"
    shutil.copyfile(src, dst)


    wb = load_workbook(dst)
    df_to_excel(wb, "Line Regulation", df, 'B2')
 
    create_chartsheet(wb, df, title="Efficiency (%)", style=2, legend=None, x_title='Input Voltage (VAC)', y_title='Efficiency',
                        x_anchor="B1", y_anchor="N1",
                        x_min_scale = 90, x_max_scale = 277,  x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 85, y_max_scale = 100, y_major_unit = 1, y_minor_unit = 1)

    create_chartsheet(wb, df, title="Power Factor", style=14, legend=None, x_title='Input Voltage (VAC)', y_title='Power Factor',
                        x_anchor="B1", y_anchor="G1",
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 0.8, y_max_scale = 1, y_major_unit = 0.01, y_minor_unit = 0.01)

    create_chartsheet(wb, df, title="THD (%)", style=14, legend=None, x_title='Input Voltage (VAC)', y_title='THD (%)',
                        x_anchor="B1", y_anchor="H1",
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 0, y_max_scale = 15, y_major_unit = 1, y_minor_unit = 1)
    
    create_chartsheet(wb, df, title="Output Current (mA)", style=5, legend=None, x_title='Input Voltage (VAC)', y_title='Output Current (mA)',
                        x_anchor="B1", y_anchor="J1",
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = iout*0.95*1000, y_max_scale = iout*1.05*1000, y_major_unit = iout*0.05*100, y_minor_unit = iout*0.05*100)

    create_chartsheet(wb, df, title="Output Voltage (V)", style=5, legend=None, x_title='Input Voltage (VAC)', y_title='Output Voltage (V)',
                        x_anchor="B1", y_anchor="I1",
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = vout*0.95, y_max_scale = vout*1.05, y_major_unit = vout*0.005, y_minor_unit = vout*0.005)

    create_chartsheet(wb, df, title="Input Power (W)", style=5, legend=None, x_title='Input Voltage (VAC)', y_title='Input Power (W)',
                        x_anchor="B1", y_anchor="F1",
                        x_min_scale = 90, x_max_scale = 277, x_major_unit = 20, x_minor_unit = 10,
                        y_min_scale = 0, y_max_scale = 90, y_major_unit = 10, y_minor_unit = 10)

    wb.save(dst)





operation()