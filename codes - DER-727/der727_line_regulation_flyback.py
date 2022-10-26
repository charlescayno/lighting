"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	


import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt

def col_row_extractor(excel_coordinate):
    """
    extract col and row given an excel coordinate (i.e. 'B4' -> col = 2, row = 4)

    excel_coordinate : i.e. 'B4' (str)
    returns col, row (int)
    """
    coordinates = coordinate_from_string(excel_coordinate)
    col = column_index_from_string(coordinates[0])
    row = coordinates[1]
    return col, row

def excel_to_df(filename, sheet_name, start_corner, end_corner):
    """
    reading dataframe from excel.
    
    filename     : must include full filename path (cwd + path + file.extension)
    sheet_name   : sheet name in excel file
    start_corner : cell coordinate to start selection of data
    end_corner   : cell coordinate to end selection of data

    returns df
    """

    # print(f"reading dataframe from {filename} {sheet_name}")

    start_col, start_row = col_row_extractor(start_corner)
    end_col, end_row = col_row_extractor(end_corner)

    skiprows = start_row - 2
    usecols = f'{get_column_letter(start_col)}:{get_column_letter(end_col)}'
    nrows = end_row - start_row + 1

    return pd.read_excel(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)

def df_to_excel(wb, sheet_name, df, anchor):
    """
    writing dataframe to excel.

    wb          : workbook
    sheet_name  : sheet name in excel file
    df          : dataframe
    anchor      : anchor point in excel

    returns None
    """

    # print(f"writing in {wb} {sheet_name}")

    start_col, start_row = col_row_extractor(anchor)
    df_row_len, df_col_len = df.shape
    end_row = start_row + df_row_len - 1
    end_col = start_col + df_col_len - 1


    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            wb[sheet_name][f'{get_column_letter(col)}{row}'] = df.iloc[row-start_row, col-start_col]

def image_to_excel(wb, sheet_name, filename, folder_path, anchor):
    """
    writing image to excel.

    image size -> 39R, 16C
    wb          : workbook
    sheet_name  : sheet name in excel file
    filename    : filename of theh image
    folder_path : image location
    anchor      : anchor point in excel
    """

    file = os.getcwd() + folder_path + filename
    img = openpyxl.drawing.image.Image(file)
    img.anchor = anchor
    img.width = 1026
    img.height = 762

    wb[sheet_name].add_image(img)



##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 1 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.11.20"

"""USER INPUT"""
vin_list = [400]
iout_list = [0.1, 0.25, 0.5, 0.75, 1, 1.2, 1.3, 1.4, 1.56]
vout = 48

test = "Line Regulation (Flyback Stage Only) with dimming circuit"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/{test}'


"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
# scope = Oscilloscope(scope_address)

def discharge_output():
    ac.turn_off()
    for i in range(1,9):
        eload.channel[i].cc = 1
        eload.channel[i].turn_on()
        eload.channel[i].short_on()
    sleep(1)
    for i in range(1,9):
        eload.channel[i].turn_off()
        eload.channel[i].short_off()
    sleep(1)

def scope_settings():
    global condition
    
    # scope.channel_settings(state='ON', channel=1, scale=100, position=-1, label="Output Voltage", color='LIGHT_BLUE', rel_x_position=20, bandwidth=20, coupling='DC', offset=0)
    # scope.channel_settings(state='ON', channel=2, scale=2, position=3, label="VDS", color='YELLOW', rel_x_position=40, bandwidth=20, coupling='DCLimit', offset=0)
    # scope.channel_settings(state='OFF', channel=3, scale=1, position=-1, label="RelayON Pulse", color='LIGHT_BLUE', rel_x_position=60, bandwidth=20, coupling='DCLimit', offset=0)
    # scope.channel_settings(state='ON', channel=4, scale=1, position=-1, label="Input Current", color='PINK', rel_x_position=80, bandwidth=20, coupling='DCLimit', offset=0)
    
    # scope.measure(1, "MAX,RMS,FREQ")
    # scope.measure(2, "MAX,MIN")
    # scope.measure(3, "MAX,MIN")
    # scope.measure(4, "MAX,MIN")

    # scope.time_position(10)
    
    # scope.record_length(50E6)
    
    # scope.time_scale(1)

    # scope.remove_zoom()
    # scope.add_zoom(rel_pos=21.727, rel_scale=1)
    
    trigger_channel = 2
    trigger_level = 1
    trigger_edge = 'POS'
    # scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    # scope.stop()
    print()

def browning(start, end, slew, frequency):

    if start > end:
        print(f"brownout: {start} -> {end} Vac")
        for voltage in range(start, end+1, -slew):
            ac.voltage = voltage
            ac.frequency = frequency
            ac.turn_on()
            sleep(1)
    if start < end:
        print(f"brownin: {start} -> {end} Vac")
        for voltage in range(start, end+1, slew):
            ac.voltage = voltage
            ac.frequency = frequency
            ac.turn_on()
            sleep(1)

def operation():
    global waveform_counter

    # scope_settings()
    
    eload.channel[1].cc = iout
    eload.channel[1].turn_on()

    for vin in vin_list:
        ac.voltage = vin
        ac.coupling  = 'DC'
        ac.turn_on()

        input()

def test_line_regulation():

    plt.figure(figsize=(10,10))
    plt.xlabel('Iout (mA)', fontsize=24)
    plt.xticks(fontsize=18)
    plt.ylabel('Effeciency (%)', fontsize=24)
    plt.yticks(fontsize=18)
    xmin = 0
    xmax = 1400
    ymin = 0
    ymax = 96
    xmin, xmax, ymin, ymax = plt.axis()
    # plt.axis(xlim=(0, 1400), ylim=(0,96))
    effeciency_result_list = []
    iout_result_list = []


    df = pd.DataFrame(columns = ['Vin', 'Freq', 'Vdc', 'Iin', 'Pin', 'PF', 'Vo1', 'Io1', 'Po1', 'Vreg1', 'Ireg1', 'Eff'])

    for iout in iout_list:

        eload.channel[1].cc = iout
        eload.channel[1].turn_on()

        for vin in vin_list:
            ac.voltage = vin
            ac.coupling  = 'DC'
            ac.turn_on()
        
            soak(60)

            vac = vin
            freq = str(ac.set_freq(vin))
            vin = f"{pms.voltage:.2f}"
            iin = f"{pms.current*1000:.2f}"
            pin = f"{pms.power:.3f}"
            pf = f"{pms.pf:.4f}"
            # thd = f"{pms.thd:.2f}"
            vo1 = f"{pml.voltage:.3f}"
            io1 = f"{pml.current*1000:.2f}"
            po1 = f"{pml.power:.3f}"
            vreg1 = f"{100*(float(vo1)-vout)/float(vo1):.4f}"
            iout1 = float(io1)/1000
            ireg1 = f"{100*(iout1-iout)/iout1:.4f}"
            eff = f"{100*(float(po1))/float(pin):.4f}"

            output_list = [vac, freq, vin, iin, pin, pf, vo1, io1, po1, vreg1, ireg1, eff]
            # output_list = [vac, freq, vin, iin, pin, pf, vo1, io1, po1]

            df.loc[len(df)] = output_list
            # effeciency_result_list.append(df['Eff'].iloc[-1])
            # iout_result_list.append(df['Io1'].iloc[-1])
            # sleep(1)
            # plt.plot(iout_result_list, effeciency_result_list, color='blue', linewidth=10)
            # plt.pause(1)

            print(df)

    wb = Workbook()
    dest_filename = 'empty_book.xlsx'
    ws = wb.create_sheet()
    ws.title = "Line Regulation"
    df_to_excel(wb, "Line Regulation", df, 'A2')
    wb.save(f"{test}.xlsx")

    soak(3)


def main():
    global waveform_counter
    discharge_output()
    test_line_regulation()
    discharge_output()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)