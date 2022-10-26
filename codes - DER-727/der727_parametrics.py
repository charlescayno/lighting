## UPDATED AUGUST 08, 2022

"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
import numpy as np
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Keithley_DC_2230G
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
# date = now.strftime('%Y_%m_%d')
date = now.strftime('%Y_%m_%d_%H%M')
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series


##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
dimming_power_meter_address = 21
eload_address = 8
scope_address = "10.125.11.10"
dc_source_address = '4'

##################################################################################

vin_list = [90,100,110,115,120,132,180,200,230,265,277]
led_list = [48,36,24]
soak_time = 300 # 300
vout = 48
iout = 1.59
iout_list = [iout]



test = "Parametrics"

unit = input(">> Unit #: ")

waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/Final Data (REV C)/Summary/{unit}'
path = path_maker(f'{waveforms_folder}')

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
led_ctrl = LEDControl()

"""DISCHARGE FUNCTION"""
def discharge_output(times=1):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(2)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def header_data(df_header_list):
    """
    header_data: creates dataframe with header specified by user
    
    NOTE:   - header_data and collect_data are partner functions
            - header_data need to be used first before collect_data
    
    input: df_header_list
    """
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list

    return df


"""DATA MANIPULATION FUNCTIONS"""
def create_new_excel_container():
    """

    INPUT FILE: "blank.xlsx" at automation/codes folder
    OUTPUT FILE: "{test}.xlsx" at {waveforms_folder} folder

    """
    src = f"{os.getcwd()}/DATA_SUMMARY.xlsx"
    dst = f"{waveforms_folder}/DATA_SUMMARY.xlsx"
    if not os.path.exists(dst): shutil.copyfile(src, dst)

    return dst

def create_output_excel_result(dst, df_header_list, df, led, vin, test): 
    wb = load_workbook(dst)

    ###################################################################
    ####################### SUBJECT TO CHANGE #########################
    if led == 48: row = 2
    if led == 36: row = 15
    if led == 24: row = 28
    ###################################################################
    ###################################################################


    header_anchor = f"B{row}"
    data_anchor = f"B{row+1}"

    df_header = header_data(df_header_list)
    df_data = df[(df.led_load == led)]

    sheet_name = f"{test}"
    
    df_to_excel(wb, sheet_name, df_header, header_anchor)
    df_to_excel(wb, sheet_name, df_data, data_anchor)
    
    wb.save(dst)

def create_header_df(df_header_list):
    """
    create_header_df: creates dataframe with header specified by user
    
        - create_header_df and collect_data are partner functions
        - create_header_df need to be used first before collect_data
    
    input: df_header_list
    """
    df = pd.DataFrame(columns = df_header_list)
    df.loc[len(df)] = df_header_list

    return df

def collect_data(df, led, vin):

    """
    collect_data: collects data. This function can be modified according to the user's needs
    
    output: print specific output list, and put it in the existing df
    """
    ####################################################################################
    
    """COLLECT DATA"""
    led_load = led
    vac = vin
    freq = ac.set_freq(vin)
    vac_actual = float(f"{pms.voltage:.6f}")
    iin = float(f"{pms.current*1000:.6f}")
    pin = float(f"{pms.power:.6f}")
    pf = float(f"{pms.pf:.6f}")
    thd = float(f"{pms.thd:.6f}") # AC
    vo1 = float(f"{pml.voltage:.6f}")
    io1 = float(f"{pml.current*1000:.6f}")
    po1 = float(f"{pml.power:.6f}")
    vreg1 = float(f"{100*(float(vo1)-led)/float(vo1):.6f}")
    iout1 = float(io1)/1000
    try: ireg1 = float(f"{100*(iout1-iout)/iout1:.6f}")
    except: ireg1 = 0
    eff = float(f"{100*(float(po1))/float(pin):.6f}")

    output_list = [led_load, vac, freq, vac_actual, iin, pin, pf, thd, vo1, io1, po1, vreg1, ireg1, eff] # AC
    
    ####################################################################################

    df.loc[len(df)] = output_list
    return df




def operation():

    dst = create_new_excel_container()
    df_header_list = ['led_load', 'Vin', 'Freq (Hz)', 'Vac (VAC)', 'Iin (mA)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (mA)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)'] 
    df_header = create_header_df(df_header_list)
    df = df_header

    for led in led_list:

        led_ctrl.voltage(led)

        for vin in vin_list:

            ac.voltage = vin
            ac.turn_on()
        
            soak(soak_time)

            df = collect_data(df, led, vin)

            print(df[(df.led_load == led)])   
            create_output_excel_result(dst, df_header_list, df, led, vin, test)


        print(f"\n\nFinal Data at {led}V: ")
        print(df)

        discharge_output(2)

        


def main():
    global waveform_counter
    # discharge_output(3)
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)
