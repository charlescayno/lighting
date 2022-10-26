"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
# date = now.strftime('%Y_%m_%d')
date = now.strftime('%Y_%m_%d_%H%M')


import shutil
import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt


from openpyxl import Workbook, load_workbook



def col_row_extractor(excel_coordinate):
    """
    extract col and row given an excel coordinate (i.e. 'B4' -> col = 2, row = 4)

    excel_coordinate : i.e. 'B4' (str)
    returns col, row (int)
    """
    coordinates = coordinate_from_string(excel_coordinate)
    col = column_index_from_string(coordinates[0])
    row = coordinates[1]
    return col, row

def excel_to_df(filename, sheet_name, start_corner, end_corner):
    """
    reading dataframe from excel.
    
    filename     : must include full filename path (cwd + path + file.extension)
    sheet_name   : sheet name in excel file
    start_corner : cell coordinate to start selection of data
    end_corner   : cell coordinate to end selection of data

    returns df
    """

    # print(f"reading dataframe from {filename} {sheet_name}")

    start_col, start_row = col_row_extractor(start_corner)
    end_col, end_row = col_row_extractor(end_corner)

    skiprows = start_row - 2
    usecols = f'{get_column_letter(start_col)}:{get_column_letter(end_col)}'
    nrows = end_row - start_row + 1

    return pd.read_excel(filename, sheet_name, skiprows=skiprows, usecols=usecols, nrows=nrows)

def df_to_excel(wb, sheet_name, df, anchor):
    """
    writing dataframe to excel.

    wb          : workbook
    sheet_name  : sheet name in excel file
    df          : dataframe
    anchor      : anchor point in excel

    returns None
    """

    # print(f"writing in {wb} {sheet_name}")

    start_col, start_row = col_row_extractor(anchor)
    df_row_len, df_col_len = df.shape
    end_row = start_row + df_row_len - 1
    end_col = start_col + df_col_len - 1


    for row in range(start_row, end_row+1):
        for col in range(start_col, end_col+1):
            wb[sheet_name][f'{get_column_letter(col)}{row}'] = df.iloc[row-start_row, col-start_col]

def image_to_excel(wb, sheet_name, filename, folder_path, anchor):
    """
    writing image to excel.

    image size -> 39R, 16C
    wb          : workbook
    sheet_name  : sheet name in excel file
    filename    : filename of theh image
    folder_path : image location
    anchor      : anchor point in excel
    """

    file = os.getcwd() + folder_path + filename
    img = openpyxl.drawing.image.Image(file)
    img.anchor = anchor
    img.width = 1026
    img.height = 762

    wb[sheet_name].add_image(img)



##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 1 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.11.20"

"""USER INPUT"""
vin_list = [90,277]
iout_list = [1.59]
vout = 48


test_list = convert_argv_to_int_list(sys.argv[1]) # 1 - Startup, 2 - Normal
component = sys.argv[2]
special_test_condition = sys.argv[3]

for test_condition in test_list:
    if test_condition == 1: test = 'Startup'
    if test_condition == 2: test = 'Normal'


waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/CSA/{component}'
path = path_maker(f'{waveforms_folder}')

"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)

def discharge_output():
    ac.turn_off()
    for i in range(1,9):
        eload.channel[i].cc = 1
        eload.channel[i].turn_on()
        eload.channel[i].short_on()
    sleep(1)
    for i in range(1,9):
        eload.channel[i].turn_off()
        eload.channel[i].short_off()
    sleep(1)

def scope_settings():

    
    
    trigger_channel = 2
    trigger_level = 50
    trigger_edge = 'POS'
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    
    scope.channel_settings(state='OFF', channel=1)
    # scope.channel_settings(state='OFF', channel=2)
    scope.channel_settings(state='OFF', channel=3)
    scope.channel_settings(state='OFF', channel=4)

    scope.channel_settings(state='ON', channel=2, scale=100, position=-4, label=f"{component}_{special_test_condition}", color='YELLOW', rel_x_position=40, bandwidth=500, coupling='DCLimit', offset=0)
    scope.measure(2, "MAX")

    
    scope.stop()
    print()

def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1


def operation():
    global waveform_counter

    scope_settings()
    
    for iout in iout_list:

        for vin in vin_list:

            for test_condition in test_list:

                if test_condition == 1: # startup
                    scope.record_length(50E6)
                    scope.time_position(10)
                    scope.time_scale(100E-3)

                    cr = vout/iout
                    eload.channel[5].cr = cr
                    eload.channel[5].turn_on()
                    sleep(2)
                    scope.run_single()
                    sleep(2)
                    ac.voltage = vin
                    ac.turn_on()
                    sleep(3)
                    screenshot(f'Startup, CR_{cr:.2f}ohms, {vin}Vac, {ac.frequency}Hz, {special_test_condition}.png', path)            
                    discharge_output()

                if test_condition == 2: # normal
                    scope.record_length(50E6)
                    scope.time_position(50)
                    scope.time_scale(10E-6)

                    cr = vout/iout
                    eload.channel[5].cr = cr
                    eload.channel[5].turn_on()
                    sleep(2)
                    ac.voltage = vin
                    ac.turn_on()
                    sleep(3)
                    scope.run_single()
                    sleep(3)
                    screenshot(f'Normal, CR_{cr:.2f}ohms, {vin}Vac, {ac.frequency}Hz, {special_test_condition}.png', path)            
                    discharge_output()

def main():
    global waveform_counter
    discharge_output()
    operation()
    discharge_output()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)