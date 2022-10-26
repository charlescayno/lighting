##################################################################################
"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
import numpy as np
import shutil
import os
import pandas as pd

# from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Keithley_DC_2230G
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.equipment import create_header_list, export_to_excel, export_screenshot_to_excel, export_df_to_excel
from powi.equipment import path_maker, remove_file
from powi.equipment import create_scatter_chart, append_series, reset_chartsheet, save_chartsheet, create_bar_chart
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import getpass
username = getpass.getuser().lower()

from datetime import datetime
now = datetime.now()
date = now.strftime('%m%d')
##################################################################################

unit = sys.argv[1]
waveforms_folder = f"C:/Users/ccayno/Documents/Charles/Work/DER/DER-727/Marketing Samples/1005/Marketing Samples/{unit}/"
path = path_maker(f'{waveforms_folder}')
excel_name = f"DER727_UNIT_{unit}"
full_path = path + excel_name + '.xlsx'
print(full_path)

def anchor_picker(ws_name, chart, data_length=101, led_row=[3,106,209], x_axis_col=['C', 'T', 'AK', 'BB'], y_axis_col=['M', 'AD', 'AU', 'BL']):
    
    """
        led_row: sequential reference starting from 48 V to 24 V
        x_axis_col: data anchors for x-axis chart; sequential reference starting from 90 VAC to 277 VAC
        y_axis_col: data anchors for y-axis chart; sequential reference starting from 90 VAC to 277 VAC    
    """

    if ws_name == 'Parametrics':
        led_list = [48, 36, 24]
        vin_list = [90, 115, 230, 277]
        for led in led_list:

            # row section per led
            if led == 48: row = led_row[0]
            if led == 36: row = led_row[1]
            if led == 24: row = led_row[2]

            # vin_list = [90,100,110,115,120,132,180,200,230,265,277]

            x_col = x_axis_col[0]
            y_col = y_axis_col[0]
            append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V", chart=chart)
    
    elif ws_name == "No Load":
        row = led_row[0]
        x_col = x_axis_col[0]
        y_col = y_axis_col[0]
        append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"No Load", chart=chart)

    elif ws_name == 'DALI':
        x_col = x_axis_col[0]
        y_col = y_axis_col[0]
        row = led_row[0]
        led = 48
        append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V", chart=chart)

    elif ws_name == 'Analog Dimming' or ws_name == 'Analog Deep Dimming':
        led_list = [48,24]
        vin_list = [115,230]

        for led in led_list:
            for i in range(len(led_list)):
                if led == led_list[i]: row = led_row[i]

            for vin in vin_list:

                for i in range(len(vin_list)):

                    if vin == vin_list[i]:
                        x_col = x_axis_col[i]
                        y_col = y_axis_col[i]

                append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V, {vin} VAC", chart=chart)
   
    elif ws_name == 'Resistor Dimming':
        led_list = [48,24]
        vin_list = [90,115,230]

        for led in led_list:
            for i in range(len(led_list)):
                if led == led_list[i]: row = led_row[i]

            for vin in vin_list:

                for i in range(len(vin_list)):

                    if vin == vin_list[i]:
                        x_col = x_axis_col[i]
                        y_col = y_axis_col[i]

                append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V, {vin} VAC", chart=chart)

    elif ws_name == 'PWM_300Hz' or ws_name == 'PWM_3000Hz':
        led_list = [48,24]
        vin_list = [115,230]

        for led in led_list:
            for i in range(len(led_list)):
                if led == led_list[i]: row = led_row[i]

            for vin in vin_list:

                for i in range(len(vin_list)):

                    if vin == vin_list[i]:
                        x_col = x_axis_col[i]
                        y_col = y_axis_col[i]

                append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V, {vin} VAC", chart=chart)

    else:
        pass

def parametrics_chart(ws_name, chart_row_location = 2):
    
    #########################################
    ### user variables ######################
    title = f"Parametrics"
    x_title = 'Input Voltage (VAC)'
    led_row = [3,16,29] # 48, 36, 24 V respectively
    x_axis_col = ['C']
    y_axis_col_iout = ['K']
    y_axis_col_eff = ['O']
    y_axis_col_pf = ['H']
    y_axis_col_thd = ['I']

    x_min_scale = 90
    x_max_scale = 277
    x_major_unit = 17
    x_minor_unit = 17
    
    iout = 1590
    iout_percent_regulation = 5

    y_min_scale_iout_regulation = (1-iout_percent_regulation/100)*iout
    y_max_scale_iout_regulation = (1+iout_percent_regulation/100)*iout
    y_major_unit_iout_regulation = (iout_percent_regulation*2/1000)*iout
    y_minor_unit_iout_regulation = (iout_percent_regulation*2/1000)*iout

    y_min_scale_eff = 87
    y_max_scale_eff = 93
    y_major_unit_eff = 1
    y_minor_unit_eff = 1
    
    style = 2
    ### user variables ######################
    #########################################
    
    data_length = led_row[1] - led_row[0] - 2

    """ OUTPUT CURRENT REGULATION """
    chart = create_scatter_chart(title=title, style=style, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = y_min_scale_iout_regulation, y_max_scale = y_max_scale_iout_regulation, y_major_unit = y_major_unit_iout_regulation, y_minor_unit = y_minor_unit_iout_regulation)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"B{chart_row_location}")

    """ EFFICIENCY CHART """
    chart = create_scatter_chart(title="Efficiency (%)", style=style, x_title=x_title, y_title='Efficiency (%)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = y_min_scale_eff , y_max_scale = y_max_scale_eff, y_major_unit = y_major_unit_eff, y_minor_unit = y_minor_unit_eff)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    """ POWER FACTOR CHART """
    chart = create_scatter_chart(title="Power Factor", style=style, x_title=x_title, y_title='Power Factor',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 0.85, y_max_scale = 1, y_major_unit = 0.05, y_minor_unit = 0.05)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    """ THD (%) CHART """
    chart = create_scatter_chart(title="THD (%)", style=style, x_title=x_title, y_title='THD (%)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 0, y_max_scale = 20, y_major_unit = 2, y_minor_unit = 2)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")

def no_load_chart(ws_name, chart_row_location = 2):
    
    #########################################
    ### user variables ######################
    title = f"No Load"
    x_title = 'Input Voltage (VAC)'
    led_row = [3]
    x_axis_col = ['B']
    y_axis_col_pin = ['F']

    x_min_scale = 90
    x_max_scale = 277
    x_major_unit = 17
    x_minor_unit = 17
    
    style = 2
    ### user variables ######################
    #########################################
    
    data_length = 11

    """ OUTPUT CURRENT REGULATION """
    chart = create_scatter_chart(title=title, style=style, x_title=x_title, y_title='No Load Input Power (W)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 0.2, y_max_scale = 0.4, y_major_unit = 0.02, y_minor_unit = 0.02)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pin)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"B{chart_row_location}")

def analog_chart(ws_name, chart_row_location = 42):
    
    #########################################
    ### user variables ######################
    title = f"Analog Dimming"
    x_title = 'Dimming Voltage (V)'
    led_row = [3, 43]
    x_axis_col = ['C', 'T']
    y_axis_col_iout = ['M', 'AD']
    y_axis_col_eff= ['Q', 'AH']
    y_axis_col_pf = ['J', 'AA']
    y_axis_col_thd = ['K', 'AB']
    ### user variables ######################
    #########################################
    
    data_length = led_row[1] - led_row[0] - 2

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = 0, x_max_scale = 10, x_major_unit = 1, x_minor_unit = 0.5,
                        y_min_scale = 0, y_max_scale = 1600, y_major_unit = 100, y_minor_unit = 5)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")

    # """ EFFICIENCY CHART """
    # chart = create_scatter_chart(title="Efficiency (%)", style=2, x_title=x_title, y_title='Efficiency (%)',
    #                     x_min_scale = 0, x_max_scale = 10, x_major_unit = 1, x_minor_unit = 0.5,
    #                     y_min_scale = 0, y_max_scale = 95, y_major_unit = 10, y_minor_unit = 5)



    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    # save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    # """ POWER FACTOR CHART """
    # chart = create_scatter_chart(title="Power Factor", style=2, x_title=x_title, y_title='Power Factor',
    #                     x_min_scale = 0, x_max_scale = 10, x_major_unit = 1, x_minor_unit = 0.5,
    #                     y_min_scale = 0, y_max_scale = 1, y_major_unit = 0.1, y_minor_unit = 0.1)

    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    # save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    # """ THD (%) CHART """
    # chart = create_scatter_chart(title="THD (%)", style=2, x_title=x_title, y_title='THD (%)',
    #                     x_min_scale = 0, x_max_scale = 10, x_major_unit = 1, x_minor_unit = 0.5,
    #                     y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 1)


    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    # save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")

def deep_dimming_analog_chart(ws_name, chart_row_location = 42):
    
    #########################################
    ### user variables ######################
    title = f"Analog Deep Dimming"
    x_title = 'Dimming Voltage (V)'
    led_row = [3,66] # 48, 36, 24 V respectively
    x_axis_col = ['C', 'T']
    y_axis_col_iout = ['M', 'AD']
    y_axis_col_eff= ['Q', 'AH']
    y_axis_col_pf = ['J', 'AA']
    y_axis_col_thd = ['K', 'AB']

    x_min_scale = 0
    x_max_scale = 0.5
    x_major_unit = 0.05
    x_minor_unit = 0.05
    style = 2


    ### user variables ######################
    #########################################
    
    data_length = led_row[1] - led_row[0] - 2

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 0, y_max_scale = 90, y_major_unit = 10, y_minor_unit = 10)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


def resistor_chart(ws_name, chart_row_location = 22):
    
    #########################################
    ### user variables ######################
    title = f"Resistor Dimming"
    x_title = 'Dimming Resistance (kOhms)'
    led_row = [2,22] # 48, 36, 24 V respectively
    x_axis_col = ['B', 'R', 'AH']
    y_axis_col_iout = ['L', 'AB', 'AR']
    y_axis_col_eff= ['P', 'AF', 'AV']
    y_axis_col_pf = ['I', 'Y', 'AO']
    y_axis_col_thd = ['J', 'Z', 'AP']

    ### user variables ######################
    #########################################
    
    data_length = led_row[1] - led_row[0] - 1

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 1600, y_major_unit = 100, y_minor_unit = 5)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")

    # """ EFFICIENCY CHART """
    # chart = create_scatter_chart(title="Efficiency (%)", style=2, x_title=x_title, y_title='Efficiency (%)',
    #                     x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
    #                     y_min_scale = 0, y_max_scale = 95, y_major_unit = 10, y_minor_unit = 5)



    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    # save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    # """ POWER FACTOR CHART """
    # chart = create_scatter_chart(title="Power Factor", style=2, x_title=x_title, y_title='Power Factor',
    #                     x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
    #                     y_min_scale = 0, y_max_scale = 1, y_major_unit = 0.1, y_minor_unit = 0.1)

    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    # save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    # """ THD (%) CHART """
    # chart = create_scatter_chart(title="THD (%)", style=2, x_title=x_title, y_title='THD (%)',
    #                     x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
    #                     y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 1)

    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    # save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")



def pwm_chart(ws_name, chart_row_location = 82, frequency=3000):
    
    #########################################
    ### user variables ######################
    title = f"PWM Dimming (f = {frequency} Hz)"
    x_title = 'Duty Cycle (%)'
    led_row = [3,106] # 48, 24 V respectively
    x_axis_col = ['C', 'T']
    y_axis_col_iout = ['M', 'AD']
    y_axis_col_eff= ['Q', 'AH']
    y_axis_col_pf = ['J', 'AA']
    y_axis_col_thd = ['K', 'AB']

    ### user variables ######################
    #########################################
    # ws_name = f"{ws_name}_{frequency}Hz"
    
    data_length = led_row[1] - led_row[0] - 2

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 1600, y_major_unit = 100, y_minor_unit = 5)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    if frequency == 300: col = "B"
    if frequency == 3000: col = "N"
    save_chartsheet(chart_sheet, chart, chart_position=f"{col}{chart_row_location}")

    # """ EFFICIENCY CHART """
    # chart = create_scatter_chart(title="Efficiency (%)", style=2, x_title=x_title, y_title='Efficiency (%)',
    #                     x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
    #                     y_min_scale = 0, y_max_scale = 95, y_major_unit = 10, y_minor_unit = 5)


    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    # save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    # """ POWER FACTOR CHART """
    # chart = create_scatter_chart(title="Power Factor", style=2, x_title=x_title, y_title='Power Factor',
    #                     x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
    #                     y_min_scale = 0, y_max_scale = 1, y_major_unit = 0.1, y_minor_unit = 0.1)

    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    # save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    # """ THD (%) CHART """
    # chart = create_scatter_chart(title="THD (%)", style=2, x_title=x_title, y_title='THD (%)',
    #                     x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
    #                     y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 1)


    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    # save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")



def dali_chart(ws_name, chart_row_location = 2):
    
    #########################################
    ### user variables ######################
    title = f"DALI"
    x_title = 'DALI bits'
    led_row = [2] # 48V
    x_axis_col = ['B']
    y_axis_col_iout = ['N']
    y_axis_col_eff= ['R']
    y_axis_col_pf = ['K']
    y_axis_col_thd = ['L']

    ### user variables ######################
    #########################################
    
    # data_length = led_row[1] - led_row[0] - 2
    data_length = 256

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = 0, x_max_scale = 254, x_major_unit = 50, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 1600, y_major_unit = 100, y_minor_unit = 5)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")

    # """ EFFICIENCY CHART """
    # chart = create_scatter_chart(title="Efficiency (%)", style=2, x_title=x_title, y_title='Efficiency (%)',
    #                     x_min_scale = 0, x_max_scale = 254, x_major_unit = 50, x_minor_unit = 2,
    #                     y_min_scale = 0, y_max_scale = 95, y_major_unit = 10, y_minor_unit = 5)


    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    # save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    # """ POWER FACTOR CHART """
    # chart = create_scatter_chart(title="Power Factor", style=2, x_title=x_title, y_title='Power Factor',
    #                     x_min_scale = 0, x_max_scale = 254, x_major_unit = 50, x_minor_unit = 2,
    #                     y_min_scale = 0, y_max_scale = 1, y_major_unit = 0.1, y_minor_unit = 0.1)

    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    # save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    # """ THD (%) CHART """
    # chart = create_scatter_chart(title="THD (%)", style=2, x_title=x_title, y_title='THD (%)',
    #                     x_min_scale = 0, x_max_scale = 254, x_major_unit = 50, x_minor_unit = 2,
    #                     y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 1)

    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    # save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")




def _transfer_df_to_another_excel(filename, waveforms_folder, old_sheet_name, new_sheet_name, excel_name, 
                                    old_anchor, new_row, new_col):

    file = waveforms_folder + filename

    start_col, start_row = col_row_extractor(old_anchor)
    row = start_row
    col = 0
    usecols = get_column_letter(start_col)

    skiprows = start_row - 2
    usecols = f'{get_column_letter(start_col)}:{get_column_letter(1000)}'
    nrows = 1000 - start_row + 1
    
    try: data = pd.read_excel(file, old_sheet_name, usecols = usecols, header=skiprows, nrows=nrows)
    except:
        data = pd.read_excel(file, old_sheet_name, usecols = usecols, header=0, nrows=nrows) # this is for data not starting at 1st row of excel
        
    
    ##### finding the maximum row without nan #####
    while 1:
        try:
            dummy = data.iat[row, 0]
            if str(dummy) == 'nan': break
            row += 1
        except: break
    row_len = row

    ##### finding the maximum col without nan #####
    while 1:
        try:
            dummy = data.iat[0, col]
            if str(dummy) == 'nan': break
            col += 1
        except: break
    col_len = col

    end_col = start_col + col_len - 1
    end_row = start_row + row_len - 1
    end_corner = f"{get_column_letter(end_col)}{end_row}"

    old_end_corner = end_corner



    new_anchor = f'{get_column_letter(new_col)}{new_row}'


    df = excel_to_df(filename=file, sheet_name=old_sheet_name, start_corner=old_anchor, end_corner=old_end_corner)
    export_df_to_excel(df, waveforms_folder, excel_name=excel_name, sheet_name=new_sheet_name, anchor=new_anchor)

    df_row_len, df_col_len = df.shape
    return df_row_len, df_col_len


def transfer_df_to_another_excel(filename, old_sheet_name, new_sheet_name, excel_name, list_of_anchor_list):
    """
        This function needs to pair with _transfer_df_to_another_excel function for it to work.
        
        Transfer dataframe from old excel to another excel.
        Top down approach.
        List all for the first data set of anchors and will copy it from top to bottom

        Proceeding data set / anchor list will be the subsequent dataframe on the right of the previous data set / anchor list

        list_of_anchor_list = [anchor_list_1, anchor_list_2, etc.]
    """

    new_row=2
    new_col=2

    for anchor_list in list_of_anchor_list:
        for old_anchor in anchor_list:
            df_row_len, df_col_len = _transfer_df_to_another_excel(filename, waveforms_folder, old_sheet_name=old_sheet_name, new_sheet_name=new_sheet_name, excel_name=excel_name, old_anchor=old_anchor, new_row = new_row, new_col = new_col)
            new_row += df_row_len + 1
        new_col += df_col_len + 1
        new_row = 2




def operation():


    filename = f"Parametrics_{unit}.xlsx"
    old_sheet_name = "Parametrics"
    new_sheet_name = "Parametrics"
    anchor_list_1 = ["B2", "B15", "B28"]
    list_of_anchor_list = [anchor_list_1]
    transfer_df_to_another_excel(filename, old_sheet_name, new_sheet_name, excel_name, list_of_anchor_list)    

    filename = f"Analog Dimming_{unit}.xlsx"
    old_sheet_name = "Analog Dimming"
    new_sheet_name = "Analog Dimming"
    anchor_list_1 = ["S2", "S82"]
    anchor_list_2 = ["AJ2", "AJ82"]
    list_of_anchor_list = [anchor_list_1, anchor_list_2]
    transfer_df_to_another_excel(filename, old_sheet_name, new_sheet_name, excel_name, list_of_anchor_list)

    filename = f"Analog Dimming Deep Dimming_{unit}.xlsx"
    old_sheet_name = "Analog Dimming"
    new_sheet_name = "Analog Deep Dimming"
    anchor_list_1 = ["S2", "S400"]
    anchor_list_2 = ["AJ2", "AJ400"]
    list_of_anchor_list = [anchor_list_1, anchor_list_2]
    transfer_df_to_another_excel(filename, old_sheet_name, new_sheet_name, excel_name, list_of_anchor_list)

    filename = f"DALI_{unit}.xlsx"
    old_sheet_name = "DALI"
    new_sheet_name = "DALI"
    anchor_list_1 = ["AL2", "AL259", "AL516"]
    list_of_anchor_list = [anchor_list_1]
    transfer_df_to_another_excel(filename, old_sheet_name, new_sheet_name, excel_name, list_of_anchor_list)

    filename = f"Output Current Ripple_1.xlsx"
    old_sheet_name = "Sheet1"
    new_sheet_name = "Iout Ripple"
    anchor_list_1 = ["B2"]
    list_of_anchor_list = [anchor_list_1]
    transfer_df_to_another_excel(filename, old_sheet_name, new_sheet_name, excel_name, list_of_anchor_list)


    filename = f"Unit {unit} (NL, Harmonics).xlsm"
    old_sheet_name = "NL"
    new_sheet_name = "No Load"
    anchor_list_1 = ["E5"]
    list_of_anchor_list = [anchor_list_1]
    transfer_df_to_another_excel(filename, old_sheet_name, new_sheet_name, excel_name, list_of_anchor_list)

    filename = f"Unit {unit} (NL, Harmonics).xlsm"
    old_sheet_name_list = ["120Vac 48V", "230Vac 48V", "277Vac 48V", "120Vac 36V", "230Vac 36V", "277Vac 36V", "120Vac 24V", "230Vac 24V", "277Vac 24V"]
    new_sheet_name = "Harmonics"
    anchor_list_1 = ["E11", "E15"]
    list_of_anchor_list = [anchor_list_1]
    new_row=2
    new_col=2
    for anchor_list in list_of_anchor_list:
        for old_anchor in anchor_list:
            for old_sheet_name in old_sheet_name_list:
                df_row_len, df_col_len = _transfer_df_to_another_excel(filename, waveforms_folder, old_sheet_name=old_sheet_name, new_sheet_name=new_sheet_name, excel_name=excel_name, old_anchor=old_anchor, new_row = new_row, new_col = new_col)
                new_col += df_col_len + 1
            new_row += df_row_len + 1

    filename_list = [f"PWM_(300, 48)_{unit}.xlsx", f"PWM_(300, 24)_{unit}.xlsx"]
    old_sheet_name = "PWM_300Hz"
    new_sheet_name = "PWM_300Hz"
    new_row=2
    new_col=2
    for filename in filename_list:
        if filename == filename_list[0]:
            anchor_list_1 = ["S2"]
            anchor_list_2 = ["AJ2"]
            list_of_anchor_list = [anchor_list_1, anchor_list_2]
        if filename == filename_list[1]:
            anchor_list_1 = ["S208"]
            anchor_list_2 = ["AJ208"]
            list_of_anchor_list = [anchor_list_1, anchor_list_2]

        for anchor_list in list_of_anchor_list:
            for old_anchor in anchor_list:
                df_row_len, df_col_len = _transfer_df_to_another_excel(filename, waveforms_folder, old_sheet_name=old_sheet_name, new_sheet_name=new_sheet_name, excel_name=excel_name, old_anchor=old_anchor, new_row = new_row, new_col = new_col)
                new_col += df_col_len + 1 # paste data to the right                
        new_col=2 # reset col
        new_row += df_row_len + 1 # paste data at the bottom

    filename_list = [f"PWM_(3000, 48)_{unit}.xlsx", f"PWM_(3000, 24)_{unit}.xlsx"]
    old_sheet_name = "PWM_3000Hz"
    new_sheet_name = "PWM_3000Hz"
    new_row=2
    new_col=2
    for filename in filename_list:
        if filename == filename_list[0]:
            anchor_list_1 = ["S2"]
            anchor_list_2 = ["AJ2"]
            list_of_anchor_list = [anchor_list_1, anchor_list_2]
        if filename == filename_list[1]:
            anchor_list_1 = ["S208"]
            anchor_list_2 = ["AJ208"]
            list_of_anchor_list = [anchor_list_1, anchor_list_2]

        for anchor_list in list_of_anchor_list:
            for old_anchor in anchor_list:
                df_row_len, df_col_len = _transfer_df_to_another_excel(filename, waveforms_folder, old_sheet_name=old_sheet_name, new_sheet_name=new_sheet_name, excel_name=excel_name, old_anchor=old_anchor, new_row = new_row, new_col = new_col)
                new_col += df_col_len + 1 # paste data to the right                
        new_col=2 # reset col
        new_row += df_row_len + 1 # paste data at the bottom


    filename_list = [f"Resistor Dimming_48_{unit}.xlsx", f"Resistor Dimming_24_{unit}.xlsx"]
    old_sheet_name = "Resistor Dimming"
    new_sheet_name = "Resistor Dimming"
    anchor_list_1 = ["B1"]
    anchor_list_2 = ["R1"]
    anchor_list_3 = ["AH1"]
    list_of_anchor_list = [anchor_list_1, anchor_list_2, anchor_list_3]
    new_row=2
    new_col=2
    for filename in filename_list:
        for anchor_list in list_of_anchor_list:
            for old_anchor in anchor_list:
                df_header_list = ['Resistor', 'led_load', 'Vin', 'Freq (Hz)', 'Vac (VAC)', 'Iin (mA)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (mA)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)'] 
                df_header = create_header_list(df_header_list)
                export_df_to_excel(df_header, waveforms_folder, excel_name=excel_name, sheet_name=new_sheet_name, anchor=f"{get_column_letter(new_col)}1")

                df_row_len, df_col_len = _transfer_df_to_another_excel(filename, waveforms_folder, old_sheet_name=old_sheet_name, new_sheet_name=new_sheet_name, excel_name=excel_name, old_anchor=old_anchor, new_row = new_row, new_col = new_col)
                new_col += df_col_len + 1 # paste data to the right                
        new_col=2 # reset col
        new_row += df_row_len + 1 # paste data at the bottom
    

    new_row=2
    new_col=2
    for led in [48, 36, 24]:
        for vin in [90, 115, 230, 277]:
            filename = f"Vds Ids Startup Output Short - {led}V, {vin}Vac.png"
            export_screenshot_to_excel(excel_name, waveforms_folder, "Startup Output Short Waveforms", filename, f"{get_column_letter(new_col)}{new_row}")
            new_row += 34
        new_row = 2
        new_col += 17
    
    new_row=18
    new_col=2
    for led in [48, 36, 24]:
        for vin in [90, 115, 230, 277]:
            filename = f"Iout Ripple - {led}V, {vin}Vac.png"
            export_screenshot_to_excel(excel_name, waveforms_folder, "Iout Ripple", filename, f"{get_column_letter(new_col)}{new_row}")
            new_row += 34
        new_row=18
        new_col += 17


    filename = f"48V, 0-277-0 Vac, 0.5V per s.png"
    export_screenshot_to_excel(excel_name, waveforms_folder, "Brown In Brown Out", filename, "B2")





    global wb
    global chart_sheet

    wb = load_workbook(full_path)
    
    chart_sheet = reset_chartsheet(wb)
    parametrics_chart(ws_name="Parametrics", chart_row_location=2)
    no_load_chart(ws_name="No Load", chart_row_location=22)
    analog_chart(ws_name="Analog Dimming", chart_row_location=22)
    deep_dimming_analog_chart(ws_name="Analog Deep Dimming", chart_row_location=22)
    resistor_chart(ws_name="Resistor Dimming", chart_row_location=22)
    pwm_chart(ws_name="PWM_300Hz", chart_row_location=42, frequency=300)
    pwm_chart(ws_name="PWM_3000Hz", chart_row_location=42, frequency=3000)
    dali_chart(ws_name="DALI", chart_row_location=42)

    wb.save(full_path)


def main():
    print("10/05: This is the official DER-727 marketing samples data compiler.\n")
    print(f"Compiling data for DER-727 Unit {unit}...")
    operation()
    print(f"DER-727 Unit {unit} Data Pack is complete.")

        
if __name__ == "__main__":
    headers("Chart")
    main()
    footers(waveform_counter)


