# June 23, 2022

"""IMPORT DEPENDENCIES"""
from re import A
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from powi.default_scope_settings import *
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')

####################################################################################################################################################################
test = "AC Cycling"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/Final Data (REV C)/{test}'
path = path_maker(f'{waveforms_folder}')

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.111"
sig_gen_address = '11'
dimming_power_meter_address = 21

"""USER INPUT"""

vin_list = [90,115,230,277]
# vin_list = [230]
led_list = [48, 36, 24]
# led_list = [48]
pulse = float(input(">> Enter pulse duration: "))
pulse_count = 5



time_division = pulse
time_scale = 2*time_division

start_soak = 3*time_scale
off_time = pulse
on_time = pulse
end_soak = 3*time_scale
delay = start_soak + end_soak + (pulse_count*(off_time + on_time)) + off_time




"""SCOPE SETTINGS"""
trig_channel = 3
trig_level = 1.5
trig_edge = 'NEG'

time_position = 20

# measure settings: "MAX,MIN,RMS,MEAN,PDELta"

# """CHANNEL 1"""
# ch1_enable = 'OFF'
# ch1_scale = 0
# ch1_position = 0 
# ch1_label = "default"
# ch1_color = "BLUE"
# ch1_bw = 500
# ch1_coupling = "DCLimit"
# ch1_rel_x_position = 50
# ch1_measure = "MAX,MIN,RMS,MEAN,PDELta"

# """CHANNEL 2"""
# ch2_enable = 'ON'
# ch2_scale = 200
# ch2_position = -4 
# ch2_label = "Drain Voltage"
# ch2_color = "LIGHT_BLUE"
# ch2_bw = 500
# ch2_coupling = "DCLimit"
# ch2_rel_x_position = 40
# ch2_measure = "MAX,MIN,MEAN,PDELta"

"""CHANNEL 3"""
ch3_enable = 'ON'
ch3_scale = 0.2
ch3_position = -4
ch3_label = "Output Current"
ch3_color = "LIGHT_BLUE"
ch3_bw = 20
ch3_coupling = "DCLimit"
ch3_rel_x_position = 50
ch3_measure = "MAX,MIN"

# """CHANNEL 4"""
ch4_enable = 'ON'
ch4_scale = 100
ch4_position = 0 
ch4_label = "Input Voltage"
ch4_color = "YELLOW"
ch4_bw = 20
ch4_coupling = "DCLimit"
ch4_rel_x_position = 50
ch4_measure = "MAX,RMS"


"""DO NOT EDIT BELOW THIS LINE"""
####################################################################################################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output(times):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(2)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    
    scope.channel_settings(state=ch1_enable, channel=1, scale=ch1_scale, position=ch1_position, label=ch1_label,
                            color=ch1_color, rel_x_position=ch1_rel_x_position, bandwidth=ch1_bw, coupling=ch1_coupling, offset=0)
        
    scope.channel_settings(state=ch2_enable, channel=2, scale=ch2_scale, position=ch2_position, label=ch2_label,
                            color=ch2_color, rel_x_position=ch2_rel_x_position, bandwidth=ch2_bw, coupling=ch2_coupling, offset=0)
    
    scope.channel_settings(state=ch3_enable, channel=3, scale=ch3_scale, position=ch3_position, label=ch3_label,
                            color=ch3_color, rel_x_position=ch3_rel_x_position, bandwidth=ch3_bw, coupling=ch3_coupling, offset=0)
    
    scope.channel_settings(state=ch4_enable, channel=4, scale=ch4_scale, position=ch4_position, label=ch4_label,
                            color=ch4_color, rel_x_position=ch4_rel_x_position, bandwidth=ch4_bw, coupling=ch4_coupling, offset=0)
    
    if ch1_enable != 'OFF': scope.measure(1, ch1_measure)
    if ch2_enable != 'OFF': scope.measure(2, ch2_measure)
    if ch3_enable != 'OFF': scope.measure(3, ch3_measure)
    if ch4_enable != 'OFF': scope.measure(4, ch4_measure)

    scope.record_length(50E6)
    scope.time_position(time_position)
    scope.time_scale(time_scale)

    # scope.remove_zoom()
    # scope.add_zoom(rel_pos=50, rel_scale=0.04)
    
    trigger_channel = trig_channel
    trigger_level = trig_level
    trigger_edge = trig_edge
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    pml.set_current_range(2)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1

def operation():

    for led in led_list:

        led_ctrl.voltage(led)
        
        for vin in vin_list:
            scope.run_single()
            ac.voltage = vin
            ac.turn_on()
            ac.ac_cycling(pulse_count, vin, start_soak, off_time, on_time, end_soak)
            discharge_output(times=3)
            filename = f"AC Cycling, {pulse}s on-off - {led}V, {vin}Vac.png"
            screenshot(filename, path)



def main():
    global waveform_counter
    discharge_output(times=1)
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)