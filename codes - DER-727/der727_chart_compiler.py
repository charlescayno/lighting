"""IMPORT DEPENDENCIES"""
from dataclasses import dataclass
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Tektronix_SigGen_AFG31000, Keithley_DC_2230G
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_scatter_chart, append_series, reset_chartsheet, save_chartsheet
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	
import numpy as np




path = f"C:/Users/ccayno/Desktop/DER/DER-727/Test Data/Final Data/Dimming/Resistor Dimming/"
path = f"C:/Users/ccayno/Desktop/Apps Eval/Jon/"
# filename = f"DIMMING DATA.xlsx"
filename = f"PWM_0902.xlsx"
full_path = path + filename


def anchor_picker(ws_name, chart, data_length=101, led_row=[3,106,209], x_axis_col=['C', 'T', 'AK', 'BB'], y_axis_col=['M', 'AD', 'AU', 'BL']):
    
    """
        led_row: sequential reference starting from 48 V to 24 V
        x_axis_col: data anchors for x-axis chart; sequential reference starting from 90 VAC to 277 VAC
        y_axis_col: data anchors for y-axis chart; sequential reference starting from 90 VAC to 277 VAC    
    """
    
    led_list = [48, 36, 24]
    vin_list = [90, 115, 230, 277]
    if ws_name == 'Parametrics':
        vin_list = [90,100,110,115,120,132,180,200,230,265,277]

    for led in led_list:

        # row section per led
        if led == 48: row = led_row[0]
        if led == 36: row = led_row[1]
        if led == 24: row = led_row[2]

        if ws_name == 'Parametrics':
                x_col = x_axis_col[0]
                y_col = y_axis_col[0]
                append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V", chart=chart)
        
        else:
            for vin in vin_list:

                # x - axis column
                if vin == 90: x_col = x_axis_col[0]
                if vin == 115: x_col = x_axis_col[1]
                if vin == 230: x_col = x_axis_col[2]
                if vin == 277: x_col = x_axis_col[3]

                # y - axis column
                if vin == 90: y_col = y_axis_col[0]
                if vin == 115: y_col = y_axis_col[1]
                if vin == 230: y_col = y_axis_col[2]
                if vin == 277: y_col = y_axis_col[3]

                append_series(full_path, wb, ws_name, x_anchor=f"{x_col}{row}", last_row_anchor=f"{y_col}{data_length+row-1}", series_title=f"{led} V, {vin} VAC", chart=chart)

def parametrics_chart(ws_name, chart_row_location = 2):
    
    #########################################
    ### user variables ######################
    title = f"Parametrics"
    x_title = 'Input Voltage (VAC)'
    led_row = [3,16,29] # 48, 36, 24 V respectively
    x_axis_col = ['C']
    y_axis_col_iout = ['K']
    y_axis_col_eff = ['O']
    y_axis_col_pf = ['H']
    y_axis_col_thd = ['I']

    x_min_scale = 90
    x_max_scale = 277
    x_major_unit = 17
    x_minor_unit = 17
    style = 2

    ### user variables ######################
    #########################################
    
    data_length = led_row[1] - led_row[0] - 2

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=style, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 1510.5, y_max_scale = 1669.5, y_major_unit = 15.9, y_minor_unit = 15.9)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"B{chart_row_location}")

    """ EFFICIENCY CHART """
    chart = create_scatter_chart(title="Efficiency (%)", style=style, x_title=x_title, y_title='Efficiency (%)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 87, y_max_scale = 93, y_major_unit = 1, y_minor_unit = 1)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    """ POWER FACTOR CHART """
    chart = create_scatter_chart(title="Power Factor", style=style, x_title=x_title, y_title='Power Factor',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 0.85, y_max_scale = 1, y_major_unit = 0.05, y_minor_unit = 0.05)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    """ THD (%) CHART """
    chart = create_scatter_chart(title="THD (%)", style=style, x_title=x_title, y_title='THD (%)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 0, y_max_scale = 20, y_major_unit = 2, y_minor_unit = 2)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")



def dali_chart(ws_name, chart_row_location = 2):
    
    #########################################
    ### user variables ######################
    title = f"DALI"
    x_title = 'DALI bits'
    led_row = [2,259,516] # 48, 36, 24 V respectively
    x_axis_col = ['N', 'AF', 'AX', 'BD']
    y_axis_col_iout = ['N', 'AF', 'AX', 'BP']
    y_axis_col_eff= ['R', 'AJ', 'BB', 'BT']
    y_axis_col_pf = ['K', 'AC', 'AU', 'BM']
    y_axis_col_thd = ['L', 'AD', 'AV', 'BN']

    ### user variables ######################
    #########################################
    
    data_length = led_row[1] - led_row[0] - 2

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = 0, x_max_scale = 254, x_major_unit = 50, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 1600, y_major_unit = 100, y_minor_unit = 5)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"B{chart_row_location}")

    """ EFFICIENCY CHART """
    chart = create_scatter_chart(title="Efficiency (%)", style=2, x_title=x_title, y_title='Efficiency (%)',
                        x_min_scale = 0, x_max_scale = 254, x_major_unit = 50, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 95, y_major_unit = 10, y_minor_unit = 5)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    """ POWER FACTOR CHART """
    chart = create_scatter_chart(title="Power Factor", style=2, x_title=x_title, y_title='Power Factor',
                        x_min_scale = 0, x_max_scale = 254, x_major_unit = 50, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 1, y_major_unit = 0.1, y_minor_unit = 0.1)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    """ THD (%) CHART """
    chart = create_scatter_chart(title="THD (%)", style=2, x_title=x_title, y_title='THD (%)',
                        x_min_scale = 0, x_max_scale = 254, x_major_unit = 50, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 1)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")

def resistor_chart(ws_name, chart_row_location = 22):
    
    #########################################
    ### user variables ######################
    title = f"Resistor Dimming"
    x_title = 'Dimming Resistance (kOhms)'
    led_row = [3,24,45] # 48, 36, 24 V respectively
    x_axis_col = ['B', 'R', 'AH', 'AX']
    y_axis_col_iout = ['L', 'AB', 'AR', 'BH']
    y_axis_col_eff= ['P', 'AF', 'AV', 'BL']
    y_axis_col_pf = ['I', 'Y', 'AO', 'BE']
    y_axis_col_thd = ['J', 'Z', 'AP', 'BF']

    ### user variables ######################
    #########################################
    
    data_length = led_row[1] - led_row[0] - 2

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 1600, y_major_unit = 100, y_minor_unit = 5)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"B{chart_row_location}")

    """ EFFICIENCY CHART """
    chart = create_scatter_chart(title="Efficiency (%)", style=2, x_title=x_title, y_title='Efficiency (%)',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 95, y_major_unit = 10, y_minor_unit = 5)



    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    """ POWER FACTOR CHART """
    chart = create_scatter_chart(title="Power Factor", style=2, x_title=x_title, y_title='Power Factor',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 1, y_major_unit = 0.1, y_minor_unit = 0.1)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    """ THD (%) CHART """
    chart = create_scatter_chart(title="THD (%)", style=2, x_title=x_title, y_title='THD (%)',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 1)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")

def analog_chart(ws_name, chart_row_location = 42):
    
    #########################################
    ### user variables ######################
    title = f"Analog Dimming"
    x_title = 'Dimming Voltage (V)'
    led_row = [3,43,83] # 48, 36, 24 V respectively
    x_axis_col = ['D', 'T', 'AJ', 'BD']
    y_axis_col_iout = ['M', 'AC', 'AS', 'BL']
    y_axis_col_eff= ['Q', 'AG', 'AY', 'BP']
    y_axis_col_pf = ['J', 'AA', 'AR', 'BI']
    y_axis_col_thd = ['K', 'AB', 'AS', 'BJ']
    ### user variables ######################
    #########################################
    
    data_length = led_row[1] - led_row[0] - 2

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = 0, x_max_scale = 10, x_major_unit = 1, x_minor_unit = 0.5,
                        y_min_scale = 0, y_max_scale = 1600, y_major_unit = 100, y_minor_unit = 5)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"B{chart_row_location}")

    """ EFFICIENCY CHART """
    chart = create_scatter_chart(title="Efficiency (%)", style=2, x_title=x_title, y_title='Efficiency (%)',
                        x_min_scale = 0, x_max_scale = 10, x_major_unit = 1, x_minor_unit = 0.5,
                        y_min_scale = 0, y_max_scale = 95, y_major_unit = 10, y_minor_unit = 5)



    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    """ POWER FACTOR CHART """
    chart = create_scatter_chart(title="Power Factor", style=2, x_title=x_title, y_title='Power Factor',
                        x_min_scale = 0, x_max_scale = 10, x_major_unit = 1, x_minor_unit = 0.5,
                        y_min_scale = 0, y_max_scale = 1, y_major_unit = 0.1, y_minor_unit = 0.1)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    """ THD (%) CHART """
    chart = create_scatter_chart(title="THD (%)", style=2, x_title=x_title, y_title='THD (%)',
                        x_min_scale = 0, x_max_scale = 10, x_major_unit = 1, x_minor_unit = 0.5,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 1)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")

def deep_dimming_analog_chart(ws_name, chart_row_location = 42):
    
    #########################################
    ### user variables ######################
    title = f"Analog Deep Dimming"
    x_title = 'Dimming Voltage (V)'
    led_row = [3,65,127] # 48, 36, 24 V respectively
    x_axis_col = ['L', 'AB', 'AR', 'BH']
    y_axis_col_iout = ['K', 'AA', 'AQ', 'BG']
    y_axis_col_eff= ['P', 'AF', 'AV', 'BL']
    y_axis_col_pf = ['H', 'X', 'AN', 'BD']
    y_axis_col_thd = ['I', 'Y', 'AO', 'BE']

    x_min_scale = 0
    x_max_scale = 0.5
    x_major_unit = 0.05
    x_minor_unit = 0.05
    style = 2


    ### user variables ######################
    #########################################
    
    data_length = led_row[1] - led_row[0] - 2

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
                        y_min_scale = 0, y_max_scale = 90, y_major_unit = 10, y_minor_unit = 10)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"B{chart_row_location}")

    # """ EFFICIENCY CHART """
    # chart = create_scatter_chart(title="Efficiency (%)", style=2, x_title=x_title, y_title='Efficiency (%)',
    #                     x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
    #                     y_min_scale = 0, y_max_scale = 95, y_major_unit = 10, y_minor_unit = 5)



    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    # save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    # """ POWER FACTOR CHART """
    # chart = create_scatter_chart(title="Power Factor", style=2, x_title=x_title, y_title='Power Factor',
    #                     x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
    #                     y_min_scale = 0, y_max_scale = 1, y_major_unit = 0.1, y_minor_unit = 0.1)

    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    # save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    # """ THD (%) CHART """
    # chart = create_scatter_chart(title="THD (%)", style=2, x_title=x_title, y_title='THD (%)',
    #                     x_min_scale = x_min_scale, x_max_scale = x_max_scale, x_major_unit = x_major_unit, x_minor_unit = x_minor_unit,
    #                     y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 1)


    # anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    # save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")


def pwm_chart(ws_name, chart_row_location = 82, frequency=3000):
    
    #########################################
    ### user variables ######################
    title = f"PWM Dimming (f = {frequency} Hz)"
    x_title = 'Duty Cycle (%)'
    led_row = [3,106,209] # 48, 36, 24 V respectively
    x_axis_col = ['C', 'T', 'AK', 'BB']
    y_axis_col_iout = ['M', 'AD', 'AU', 'BL']
    y_axis_col_eff= ['Q', 'AH', 'AY', 'BP']
    y_axis_col_pf = ['J', 'AA', 'AR', 'BI']
    y_axis_col_thd = ['K', 'AB', 'AS', 'BJ']

    ### user variables ######################
    #########################################
    ws_name = f"{ws_name}_{frequency}Hz"
    
    data_length = led_row[1] - led_row[0] - 2

    """ DIMMING CHART """
    chart = create_scatter_chart(title=title, style=2, x_title=x_title, y_title='Output Current (mA)',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 1600, y_major_unit = 100, y_minor_unit = 5)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_iout)
    
    save_chartsheet(chart_sheet, chart, chart_position=f"B{chart_row_location}")

    """ EFFICIENCY CHART """
    chart = create_scatter_chart(title="Efficiency (%)", style=2, x_title=x_title, y_title='Efficiency (%)',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 95, y_major_unit = 10, y_minor_unit = 5)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_eff)

    save_chartsheet(chart_sheet, chart, chart_position=f"N{chart_row_location}")



    """ POWER FACTOR CHART """
    chart = create_scatter_chart(title="Power Factor", style=2, x_title=x_title, y_title='Power Factor',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 1, y_major_unit = 0.1, y_minor_unit = 0.1)

    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_pf)

    save_chartsheet(chart_sheet, chart, chart_position=f"Z{chart_row_location}")


    """ THD (%) CHART """
    chart = create_scatter_chart(title="THD (%)", style=2, x_title=x_title, y_title='THD (%)',
                        x_min_scale = 0, x_max_scale = 100, x_major_unit = 10, x_minor_unit = 2,
                        y_min_scale = 0, y_max_scale = 100, y_major_unit = 10, y_minor_unit = 1)


    anchor_picker(ws_name, chart, data_length=data_length, led_row=led_row, x_axis_col=x_axis_col, y_axis_col=y_axis_col_thd)

    save_chartsheet(chart_sheet, chart, chart_position=f"AL{chart_row_location}")


def operation():
    global wb
    global chart_sheet

    wb = load_workbook(full_path)


    chart_sheet = reset_chartsheet(wb)
    # parametrics_chart(ws_name="Parametrics", chart_row_location = 2)
    analog_chart(ws_name="Analog Dimming", chart_row_location=2)
    # pwm_chart(ws_name="PWM", chart_row_location=42, frequency=3000)
    # pwm_chart(ws_name="PWM", chart_row_location=62, frequency=300)
    # resistor_chart(ws_name="Resistor Dimming", chart_row_location=82)
    # dali_chart(ws_name="DALI", chart_row_location=102)
    # deep_dimming_analog_chart(ws_name="Deep Dimming Data", chart_row_location=2)
    
    wb.save(full_path)


def main():
    operation()
        
if __name__ == "__main__":
    headers("Chart")
    main()
    footers(waveform_counter)


