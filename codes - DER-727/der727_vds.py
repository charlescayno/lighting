# June 23, 2022

"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_chartsheet
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_chartsheet
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	

##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.11.10"
sig_gen_address = '11'
dimming_power_meter_address = 21


"""USER INPUT"""

vin_list = [90,115,230,277]
led_list = [48, 36, 24]

vds_channel = 2

# special_test_condition = input("Enter special test condition: ")
test = "LYTSwitch-6 Drain Voltage and Current "
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/{test}'
path = path_maker(f'{waveforms_folder}')

"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output():
    ac.turn_off()
    for i in range(1):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(2)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    global condition
    global time_division
    
    scope.channel_settings(state='OFF', channel=1, scale=0.4, position=-4, label="Vfb",
                            color='YELLOW', rel_x_position=20, bandwidth=20, coupling='DCLimit', offset=0)
    
    
    scope.channel_settings(state='ON', channel=2, scale=100, position=-4, label="Drain Voltage",
                            color='LIGHT_BLUE', rel_x_position=40, bandwidth=500, coupling='DCLimit', offset=0)
    
    
    scope.channel_settings(state='ON', channel=3, scale=0.4, position=-4, label="Drain Current",
                            color='YELLOW', rel_x_position=60, bandwidth=500, coupling='DCLimit', offset=0)
    
    scope.channel_settings(state='OFF', channel=4, scale=100, position=0, label="Input Voltage",
                            color='YELLOW', rel_x_position=80, bandwidth=20, coupling='AC', offset=0)
    
    
    # scope.measure(1, "MAX,RMS")
    scope.measure(2, "MAX,RMS")
    scope.measure(3, "MAX,RMS")
    # scope.measure(4, "MAX,MIN,RMS,MEAN,PDELta")

    scope.record_length(50E6)
    scope.time_position(50)
    scope.time_scale(10E-3)

    scope.remove_zoom()
    scope.add_zoom(rel_pos=50, rel_scale=0.04)
    
    trigger_channel = 2
    trigger_level = 0
    trigger_edge = 'POS'
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1


def find_trigger(channel, trigger_delta):

    scope.edge_trigger(channel, 0, 'POS')

    # finding trigger level
    scope.run_single()
    soak(5)

    # get initial peak-to-peak measurement value
    labels, values = scope.get_measure(channel)
    max_value = float(values[0])
    max_value = float(f"{max_value:.4f}")

    # set max_value as initial trigger level
    trigger_level = max_value
    scope.edge_trigger(channel, trigger_level, 'POS')

    # check if it triggered within 3 seconds
    scope.run_single()
    soak(3)
    trigger_status = scope.trigger_status()

    # increase trigger level until it reaches the maximum trigger level
    while (trigger_status == 1):
        trigger_level += trigger_delta
        scope.edge_trigger(channel, trigger_level, 'POS')

        # check trigger status
        scope.run_single()
        soak(3)
        trigger_status = scope.trigger_status()

    # decrease trigger level below to get the maximum trigger possible
    trigger_level -= 2*trigger_delta
    scope.edge_trigger(channel, trigger_level, 'POS')
    # print(f'Maximum trigger level found at: {trigger_level}')


def operation():

    for led in led_list:

        led_ctrl.voltage(led)
        
        for vin in vin_list:

            ac.voltage = vin
            ac.turn_on()
            sleep(5)

            find_trigger(channel=vds_channel, trigger_delta=1)
            
            scope.run_single()
            sleep(3)
            
            discharge_output()
            filename = f"Vds Ids Norm Op - {led}V, {vin}Vac.png"
            screenshot(filename, path)


def main():
    global waveform_counter
    discharge_output()
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)