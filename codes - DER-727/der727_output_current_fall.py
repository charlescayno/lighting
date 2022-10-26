## UPDATED AUGUST 08, 2022

"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	

##################################################################################

"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.10.111"
sig_gen_address = '11'
dimming_power_meter_address = 21



"""USER INPUT"""

vin_list = [90,115,230,277]
led_list = [48,36,24]

trig_channel = 3
iout = 1.59

# special_test_condition = input("Enter special test condition: ")
test = "Output Fall Profile"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/Final Data (REV C)/{test}'
path = path_maker(f'{waveforms_folder}')

"""DO NOT EDIT BELOW THIS LINE"""
##################################################################################

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)
led_ctrl = LEDControl()

"""GENERIC FUNCTIONS"""

def discharge_output():
    ac.turn_off()
    for i in range(1):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(2)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(2)

def scope_settings():
    global condition
    global time_division
    
    scope.channel_settings(state='OFF', channel=1, scale=0.4, position=-4, label="Vfb",
                            color='YELLOW', rel_x_position=20, bandwidth=20, coupling='DCLimit', offset=0)
    
    
    scope.channel_settings(state='OFF', channel=2, scale=100, position=-4, label="Drain Voltage",
                            color='LIGHT_BLUE', rel_x_position=40, bandwidth=500, coupling='DCLimit', offset=0)
    
    
    scope.channel_settings(state='ON', channel=3, scale=0.2, position=-4, label="Output Current",
                            color='PINK', rel_x_position=40, bandwidth=20, coupling='DCLimit', offset=0)
    
    scope.channel_settings(state='ON', channel=4, scale=200, position=-2, label="Input Voltage",
                            color='YELLOW', rel_x_position=60, bandwidth=20, coupling='AC', offset=0)
    
    ## MEASURE OPTIONS: MAX,MIN,RMS,MEAN,PDELta
    
    # scope.measure(1, "MAX,RMS")
    # scope.measure(2, "MAX,RMS")
    scope.measure(3, "MAX,RMS")
    scope.measure(4, "MAX,RMS")
    

    scope.record_length(50E6)
    scope.time_position(50)
    scope.time_scale(100E-3)

    # scope.remove_zoom()
    # scope.add_zoom(rel_pos=50, rel_scale=0.04)
    
    trigger_channel = trig_channel
    trigger_level = 0.1*iout
    trigger_edge = 'NEG'
    scope.edge_trigger(trigger_channel, trigger_level, trigger_edge)

    scope.stop()
    
def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1

def operation():

    for led in led_list:

        led_ctrl.voltage(led)
        
        for vin in vin_list:

            ac.voltage = vin
            ac.turn_on()
            sleep(3)
            scope.run_single()
            sleep(3)
            ac.turn_off()
            


            


            
            input("Adjust cursor. Press ENTER to continue...")
            discharge_output()
            
            filename = f"Output Fall Profile - {led}V, {vin}Vac.png"
            screenshot(filename, path)
            
            


def main():
    global waveform_counter
    discharge_output()
    scope_settings()
    operation()
        
if __name__ == "__main__":
    headers(test)
    main()
    footers(waveform_counter)