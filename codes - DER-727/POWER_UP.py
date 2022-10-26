"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_chartsheet
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_chartsheet
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0

from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	



"""COMMS"""
ac_source_address = 5
source_power_meter_address = 30 
load_power_meter_address = 2
eload_address = 8
scope_address = "10.125.11.10"
sig_gen_address = '11'
dimming_power_meter_address = 21

"""USER INPUT"""
load = int(sys.argv[1])
vin = int(sys.argv[2])
special_test_condition = sys.argv[3]

test = "Startup debugging 0V dimming"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/{test}/{special_test_condition}'
path = path_maker(f'{waveforms_folder}')

"""EQUIPMENT INITIALIZE"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)

"""GENERIC FUNCTIONS"""

def discharge_output(times):
    ac.turn_off()
    for i in range(times):
        for i in range(1,9):
            eload.channel[i].cc = 1
            eload.channel[i].turn_on()
            eload.channel[i].short_on()
        sleep(0.5)

        for i in range(1,9):
            eload.channel[i].turn_off()
            eload.channel[i].short_off()
        sleep(1)

def screenshot(filename, path):
    global waveform_counter

    scope.get_screenshot(filename, path)
    print(filename)
    waveform_counter += 1


def operation():

    discharge_output(times=1)

    led = LEDControl()
    


    led.voltage(load)
    scope.run_single()

    input(">> Press ENTER to turn on AC.")

    ac.voltage = vin
    ac.turn_on()

    input(">> Press ENTER to discharge.")
    

    filename = f"{load}V, {vin}Vac, Startup, {special_test_condition}.png"
    screenshot(filename, path)

    discharge_output(times=2)


operation()