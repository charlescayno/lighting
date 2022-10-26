from selenium import webdriver
import os
from openpyxl import load_workbook
from datetime import datetime,timedelta
from joblib import Parallel, delayed
import requests
from selenium.webdriver.chrome.options import Options

def get_price(address): #digikey, mouser, RS
    global wb
    global ws
    url = ws.cell(row=address,column=3).value

    options = webdriver.ChromeOptions()
    options.headless = True
    options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.84 Safari/537.36')

    chrome = webdriver.Chrome(options=options)
    # chrome = webdriver.Chrome()
    chrome.get(url)
    source = chrome.page_source

    index = source.find('"price"')
    x = index
    while ord(source[x]) < 48 or ord(source[x]) > 57:
        x=x+1
    index = x
    while ord(source[x]) != 34:
        x=x+1
    price=source[index:x]
    price = float(price.replace(',',''))    #price

    #stock
    if url.find("digikey") != -1:
        index = source.find('"qtyAvailable"')
        x = index
        if x==-1:
            stock = "0"
        else:
            while ord(source[x]) < 48 or ord(source[x]) > 57:
                x=x+1
            index = x
            while ord(source[x]) != 34:
                x=x+1
            stock=source[index:x]

    elif url.find("mouser") != -1:
        index = source.find('"inventoryLevel"')
        x = index
        while ord(source[x]) < 48 or ord(source[x]) > 57:
            x=x+1
        index = x
        while ord(source[x]) != 34:
            x=x+1
        stock=source[index:x]
    elif url.find("rs-online") != -1:
        index = source.find('"product_page_stock_volume"')
        x = index
        while ord(source[x]) < 48 or ord(source[x]) > 57:
            x=x+1
        index = x
        while ord(source[x]) != 34:
            x=x+1
        stock=source[index:x]

    else:
        stock = "0.0"

    chrome.close()
    return price,stock


wb = load_workbook(filename = 'sample.xlsx')
ws = wb.active

array = []

y=2
while ws.cell(row=y,column=1).value != None:
    array.insert(y-1,y)
    y=y+1
pricelist = Parallel(n_jobs=-1)(delayed(get_price)(address) for address in array)


y=2
while ws.cell(row=y,column=1).value != None:
    # ws.cell(row=y,column=7).value = pricelist[y-2][0]    #price
    ws.cell(row=y,column=4).value = pricelist[y-2][1]    #stock

    y=y+1

wb.save('sample.xlsx')
