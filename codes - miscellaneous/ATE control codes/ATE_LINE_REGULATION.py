"""COMMS"""
ac_source_address = 5
source_power_meter_address = 1 
load_power_meter_address = 2    
eload_address = 8
scope_address = "10.125.11.0"

"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_chartsheet
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor, create_chartsheet
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
date = now.strftime('%Y_%m_%d')	

"""INITIALIZE EQUIPMENT"""
ac = ACSource(ac_source_address)
pms = PowerMeter(source_power_meter_address)
pml = PowerMeter(load_power_meter_address)
eload = ElectronicLoad(eload_address)
scope = Oscilloscope(scope_address)

"""USER'S INPUT"""
vin_list = [230]
iout_reg = 1660
soak_time = 30
test = f""
test = "Line Regulation"
waveforms_folder = f'C:/Users/ccayno/Desktop/DER/DER-727/Test Data/{test}'
path = path_maker(f'{waveforms_folder}')


def discharge_output():
    ac.turn_off()
    for i in range(1,9):
        eload.channel[i].cc = 1
        eload.channel[i].turn_on()
        eload.channel[i].short_on()
    sleep(3)
    for i in range(1,9):
        eload.channel[i].turn_off()
        eload.channel[i].short_off()
    sleep(1)


def test_line_regulation(vin_list, soak_time):
    global waveform_counter
    

    for vin in vin_list:
        ac.voltage = vin
        ac.turn_on()

        soak(10)

        soak(soak_time)

        # create output list
        vac = str(vin)
        freq = str(ac.set_freq(vin))
        vac_actual = f"{pms.voltage:.2f}"
        iin = f"{pms.current*1000:.2f}"
        pin = f"{pms.power:.3f}"
        pf = f"{pms.pf:.4f}"
        thd = f"{pms.thd:.2f}"
        vo1 = f"{pml.voltage:.3f}"
        io1 = f"{pml.current*1000:.2f}"
        po1 = f"{pml.power:.3f}"
        ireg1 = f"{100*(float(io1)-iout_reg)/iout_reg:.4f}"
        eff = f"{100*(float(po1))/float(pin):.4f}"

        output_list = [vac, freq, vac_actual, iin, pin, pf, thd, vo1, io1, po1, ireg1, eff]

        print(','.join(output_list))

        discharge_output()

        soak(10)

    
    footers(waveform_counter)


if __name__ == "__main__":
    headers(test)
    discharge_output()
    test_line_regulation(vin_list, soak_time)
    discharge_output()
    footers(waveform_counter)


