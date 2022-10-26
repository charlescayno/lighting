"""IMPORT DEPENDENCIES"""
from time import time, sleep
import sys
import os
import math
import numpy as np
from powi.equipment import ACSource, PowerMeter, ElectronicLoad, Oscilloscope, LEDControl, Keithley_DC_2230G
from powi.equipment import headers, create_folder, footers, waveform_counter, soak, convert_argv_to_int_list, tts, prompt
from powi.equipment import excel_to_df, df_to_excel, image_to_excel, col_row_extractor, get_anchor
from filemanager import path_maker, remove_file
import winsound as ws
from playsound import playsound
waveform_counter = 0
from datetime import datetime
now = datetime.now()
# date = now.strftime('%Y_%m_%d')
date = now.strftime('%Y_%m_%d_%H%M')
import shutil
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

import getpass
username = getpass.getuser().lower()




from powi.equipment import create_header_list, export_to_excel, export_screenshot_to_excel
##################################################################################

test = "EXCEL DATA AUTOMATION"
condition = "test"
excel_name = f"{condition}"
waveforms_folder = f'C:/Users/{username}/Desktop/APPS EVAL/Single Stage GAN/Test Data (CV)/0915 - Testing DCDC/'
path = path_maker(f'{waveforms_folder}')



def collect_data(vin):


    ################ SUBJECT TO CHANGE ################
    vac = 1
    freq = 1
    vac_actual = 1
    iin = 1
    pin = 1
    pf = 1
    thd = 1
    vo1 = 1
    io1 = 1
    po1 = 1
    vreg1 = 1
    iout1 = 1
    try: ireg1 = 1
    except: ireg1 = 0
    eff = 1

   
    output_list = [vac, freq, vac_actual, iin, pin, pf, thd, vo1, io1, po1, vreg1, ireg1, eff]

    return output_list


def operation():

    scope_address = "10.125.10.129"
    scope = Oscilloscope(scope_address)

    scope.stop()
    # input()
    a = scope.save_channel_data(1)
    print(a)



    # file = f'TCI Preliminary Data - Copy.xlsx'
    # path = f"C:/Users/ccayno/Desktop/APPS EVAL/Single Stage GAN/Test Data (CV)/0915 - Testing DCDC/"
    # filename = path + file
    # print(filename)
    # cols_to_use = ['Vin', 'Freq (Hz)', 'Vac (VAC)', 'Iin (mA)', 'Pin (W)', 'PF', 'THD (%)', 'Vo (V)', 'Io1 (mA)', 'Po (W)', 'Vreg (%)', 'Ireg (%)', 'Eff (%)'] 
    # df = pd.read_excel(filename, header=1, usecols=lambda x: x in cols_to_use)
    
    # # print(df.loc[df['Vin'] == 190])
    # print(df['Vin'])
    # print("to be continued")


    
def main():
    operation()
        
if __name__ == "__main__":
    main()

